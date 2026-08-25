import json
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dashboard.logic.classification_validity_checks import subject_source_cohort
from dashboard.logic.feature_correlation_analysis import ANALYSIS_SCENARIOS
from dashboard.logic.feature_family_followup_analysis import (
    CLINICAL_OUTCOMES,
    COVARIATE_COLUMNS,
    _benjamini_hochberg,
    _candidate_identity,
    _diagnosis_levels,
    _diagnosis_specific_slopes,
    _fit_gee,
    _prepare_analysis_data,
    _standardize,
    _wald_test,
)
from mysite.settings import MEDIA_ROOT

RESULTS_ROOT = Path(MEDIA_ROOT) / "association-sensitivity" / "hc-vs-predlb"
PRIMARY_COVARIATES = ("gender", "education")
SCENARIO = next(
    scenario
    for scenario in ANALYSIS_SCENARIOS
    if scenario["key"] == "predlb-vs-hc"
)

CANONICAL_SPECS = (
    {
        "dataset": "dataset-clinical",
        "label": "Core sleep/diary features",
        "run_dir": (
                Path(MEDIA_ROOT)
                / "analysis-preparation"
                / "dataset-clinical"
                / "20260629_132528_560853"
        ),
    },
    {
        "dataset": "dataset-clinical-acc",
        "label": "Extended sleep/diary/activity features",
        "run_dir": (
                Path(MEDIA_ROOT)
                / "analysis-preparation"
                / "dataset-clinical-acc"
                / "20260629_132533_149540"
        ),
    },
)

MODEL_VARIANTS = (
    {
        "key": "primary",
        "label": "Primary: diagnosis + sex + education",
        "numeric_covariates": PRIMARY_COVARIATES,
        "categorical_covariates": (),
    },
    {
        "key": "age-adjusted",
        "label": "Age-adjusted: primary + age",
        "numeric_covariates": (*PRIMARY_COVARIATES, "age"),
        "categorical_covariates": (),
    },
    {
        "key": "collection-adjusted",
        "label": "Collection-adjusted: primary + NINR/NU20",
        "numeric_covariates": PRIMARY_COVARIATES,
        "categorical_covariates": ("Clinical collection",),
    },
    {
        "key": "age-collection-adjusted",
        "label": "Age- and collection-adjusted",
        "numeric_covariates": (*PRIMARY_COVARIATES, "age"),
        "categorical_covariates": ("Clinical collection",),
    },
    {
        "key": "ascertainment-stratum-adjusted",
        "label": "Ascertainment-stratum-adjusted",
        "numeric_covariates": PRIMARY_COVARIATES,
        "categorical_covariates": ("Ascertainment stratum",),
    },
    {
        "key": "age-ascertainment-stratum-adjusted",
        "label": "Age- and ascertainment-stratum-adjusted",
        "numeric_covariates": (*PRIMARY_COVARIATES, "age"),
        "categorical_covariates": ("Ascertainment stratum",),
    },
)

WITHIN_COLLECTION_VARIANTS = (
    {
        "key": "within-collection-primary",
        "label": "Within collection: diagnosis + sex + education",
        "numeric_covariates": PRIMARY_COVARIATES,
        "categorical_covariates": (),
    },
    {
        "key": "within-collection-age-adjusted",
        "label": "Within collection: primary + age",
        "numeric_covariates": (*PRIMARY_COVARIATES, "age"),
        "categorical_covariates": (),
    },
)

CATEGORY_ORDERS = {
    "Clinical collection": ("NU20", "NINR"),
    "Ascertainment stratum": (
        "COBEN",
        "HC/HC2",
        "pre-LBD/pre-LBD2",
    ),
}

REPORTED_FINDINGS = (
    {
        "dataset": "dataset-clinical-acc",
        "outcome": "UPDRS",
        "feature": "IQR.activity.Median Absolute Deviation",
        "role": "Shared association",
    },
    {
        "dataset": "dataset-clinical-acc",
        "outcome": "Attention",
        "feature": "SD.activity.Relative Interdencile Range",
        "role": "Shared association",
    },
    {
        "dataset": "dataset-clinical",
        "outcome": "Executive",
        "feature": "MAD.actigraphy_norm.Awakening > 5 minutes",
        "role": "Shared association",
    },
    {
        "dataset": "dataset-clinical",
        "outcome": "RBDq",
        "feature": "Max.actigraphy.Wake bouts",
        "role": "Diagnosis interaction",
    },
    {
        "dataset": "dataset-clinical",
        "outcome": "RBDq",
        "feature": "Median.actigraphy_norm.Sleep efficiency",
        "role": "Diagnosis interaction",
    },
    {
        "dataset": "dataset-clinical",
        "outcome": "RBDq",
        "feature": "Median.actigraphy.Wake after sleep onset",
        "role": "Diagnosis interaction",
    },
)


def run_hc_vs_predlb_association_sensitivity(output_dir=None, specs=None):
    """Run frozen-candidate age and source sensitivity analyses.

    Feature candidates are read from the canonical WASO-corrected follow-up
    workbooks. No feature screening, clustering, or representative selection is
    repeated in this sensitivity analysis.
    """
    output_dir = Path(output_dir) if output_dir else _default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)
    specs = tuple(specs or CANONICAL_SPECS)

    model_frames = []
    slope_frames = []
    count_frames = []
    canonical_tables = {}
    source_paths = []

    for spec in specs:
        run_dir = Path(spec["run_dir"])
        raw_path = run_dir / "raw" / "grouped_clinical_matrix_with_stats.xlsx"
        followup_path = (
                run_dir
                / "scenarios"
                / "predlb-vs-hc"
                / "correlation"
                / "feature_family_followup_analysis.xlsx"
        )
        _require_paths(raw_path, followup_path)
        source_paths.extend((raw_path, followup_path))

        grouped_df = pd.read_excel(raw_path)
        analysis_df = _attach_source_variables(
            _prepare_analysis_data(grouped_df, SCENARIO)
        )
        candidate_df = pd.read_excel(followup_path, sheet_name="candidate_pairs")
        canonical_tables[spec["dataset"]] = {
            "adjusted": pd.read_excel(
                followup_path,
                sheet_name="adjusted_associations",
            ),
            "interactions": pd.read_excel(
                followup_path,
                sheet_name="diagnosis_interactions",
            ),
        }

        count_frames.append(_source_counts(analysis_df, spec["dataset"]))
        for variant in MODEL_VARIANTS:
            models, slopes = _fit_candidate_set(
                analysis_df=analysis_df,
                candidates=candidate_df,
                dataset=spec["dataset"],
                dataset_label=spec["label"],
                variant=variant,
                analysis_subset="All collections",
            )
            model_frames.append(models)
            slope_frames.append(slopes)

        for collection in ("NINR", "NU20"):
            collection_df = analysis_df[
                analysis_df["Clinical collection"] == collection
                ].copy()
            for variant in WITHIN_COLLECTION_VARIANTS:
                models, slopes = _fit_candidate_set(
                    analysis_df=collection_df,
                    candidates=candidate_df,
                    dataset=spec["dataset"],
                    dataset_label=spec["label"],
                    variant=variant,
                    analysis_subset=collection,
                )
                model_frames.append(models)
                slope_frames.append(slopes)

    models_df = pd.concat(model_frames, ignore_index=True)
    slopes_df = pd.concat(slope_frames, ignore_index=True)
    counts_df = pd.concat(count_frames, ignore_index=True)
    models_df = _adjust_model_p_values(models_df)
    slopes_df = _adjust_slope_p_values(slopes_df)

    reproduction_df = _primary_reproduction_check(models_df, canonical_tables)
    reported_df = _reported_results(models_df)
    reported_slopes_df = _reported_results(slopes_df)
    settings_df = _settings(output_dir, source_paths)

    workbook_path = output_dir / "association_sensitivity_analysis.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        reported_df.to_excel(writer, sheet_name="reported_findings", index=False)
        reported_slopes_df.to_excel(
            writer,
            sheet_name="reported_diagnosis_slopes",
            index=False,
        )
        models_df.to_excel(writer, sheet_name="all_models", index=False)
        slopes_df.to_excel(writer, sheet_name="diagnosis_slopes", index=False)
        counts_df.to_excel(writer, sheet_name="source_counts", index=False)
        reproduction_df.to_excel(
            writer,
            sheet_name="primary_reproduction",
            index=False,
        )
        settings_df.to_excel(writer, sheet_name="settings", index=False)
        _style_sensitivity_workbook(writer.book)

    reproduction_passed = bool(
        not reproduction_df.empty
        and reproduction_df["Reproduced"].all()
    )
    summary = {
        "scenario": SCENARIO["label"],
        "run_dir": str(output_dir),
        "workbook_path": str(workbook_path),
        "candidate_model_count": int(len(models_df)),
        "diagnosis_slope_count": int(len(slopes_df)),
        "reported_model_count": int(len(reported_df)),
        "primary_reproduction_passed": reproduction_passed,
        "canonical_sources": [str(path) for path in source_paths],
    }
    summary_path = output_dir / "association_sensitivity_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    summary["summary_path"] = str(summary_path)
    return summary


def _fit_candidate_set(
        analysis_df,
        candidates,
        dataset,
        dataset_label,
        variant,
        analysis_subset,
):
    model_rows = []
    slope_rows = []
    for candidate in candidates.to_dict("records"):
        model_row, candidate_slopes = _fit_candidate(
            analysis_df,
            candidate,
            variant,
        )
        common = {
            "Dataset": dataset,
            "Dataset label": dataset_label,
            "Analysis subset": analysis_subset,
            "Model": variant["key"],
            "Model description": variant["label"],
        }
        model_rows.append({**common, **model_row})
        slope_rows.extend({**common, **row} for row in candidate_slopes)
    return pd.DataFrame(model_rows), pd.DataFrame(slope_rows)


def _fit_candidate(analysis_df, candidate, variant):
    identity = _candidate_identity(candidate)
    outcome_field = CLINICAL_OUTCOMES[candidate["Clinical outcome"]]
    feature = candidate["Representative feature"]
    numeric_covariates = tuple(variant["numeric_covariates"])
    categorical_covariates = tuple(variant["categorical_covariates"])
    numeric_columns = [
        COVARIATE_COLUMNS[name]
        for name in numeric_covariates
        if COVARIATE_COLUMNS[name] in analysis_df.columns
    ]
    required = [
        feature,
        outcome_field,
        "Diagnosis",
        "Person ID",
        *numeric_columns,
        *categorical_covariates,
    ]
    if feature not in analysis_df.columns:
        return _failed_candidate(identity, variant, "feature missing"), []

    data = _complete_cases(
        analysis_df,
        numeric_columns=[feature, outcome_field, *numeric_columns],
        categorical_columns=[
            "Diagnosis",
            "Person ID",
            *categorical_covariates,
        ],
    )
    diagnosis_levels = _diagnosis_levels(data)
    if (
            len(data) < 5
            or len(diagnosis_levels) < 2
            or data[feature].nunique() < 2
            or data[outcome_field].nunique() < 2
    ):
        return (
            _failed_candidate(
                identity,
                variant,
                "insufficient, constant, or single-diagnosis data",
                data,
            ),
            [],
        )

    design = pd.DataFrame(
        {"Feature (per SD)": _standardize(data[feature])},
        index=data.index,
    )
    diagnosis_columns = []
    for diagnosis in diagnosis_levels[1:]:
        column = f"Diagnosis: {diagnosis}"
        design[column] = (data["Diagnosis"] == diagnosis).astype(float)
        diagnosis_columns.append(column)

    for name, column in zip(numeric_covariates, numeric_columns):
        values = data[column].to_numpy(dtype=float)
        if name == "gender":
            design["Gender"] = values - np.nanmin(values)
        else:
            design[name.title()] = _standardize(values)

    category_references = []
    for column in categorical_covariates:
        categories = _ordered_categories(data[column], column)
        if not categories:
            continue
        category_references.append(f"{column}: {categories[0]}")
        for category in categories[1:]:
            design[f"{column}: {category}"] = (
                    data[column] == category
            ).astype(float)

    y = data[outcome_field].to_numpy(dtype=float)
    main_model = _fit_gee(y, design, data["Person ID"])
    interaction_design = design.copy()
    interaction_columns = []
    for diagnosis_column in diagnosis_columns:
        diagnosis = diagnosis_column.replace("Diagnosis: ", "", 1)
        interaction_column = f"Feature x {diagnosis}"
        interaction_design[interaction_column] = (
                interaction_design["Feature (per SD)"]
                * interaction_design[diagnosis_column]
        )
        interaction_columns.append(interaction_column)
    interaction_model = _fit_gee(y, interaction_design, data["Person ID"])
    wald_statistic, interaction_p = _wald_test(
        interaction_model,
        interaction_columns,
    )

    model_row = {
        **identity,
        "n": main_model["n"],
        "Person clusters": main_model["cluster_count"],
        "Diagnosis levels": ", ".join(diagnosis_levels),
        "Numeric covariates": ", ".join(numeric_covariates),
        "Categorical covariates": ", ".join(categorical_covariates),
        "Category references": "; ".join(category_references),
        "Feature beta per SD": main_model["coefficients"].get(
            "Feature (per SD)", np.nan
        ),
        "Feature robust SE": main_model["standard_errors"].get(
            "Feature (per SD)", np.nan
        ),
        "Feature 95% CI lower": main_model["ci_lower"].get(
            "Feature (per SD)", np.nan
        ),
        "Feature 95% CI upper": main_model["ci_upper"].get(
            "Feature (per SD)", np.nan
        ),
        "Feature p": main_model["p_values"].get(
            "Feature (per SD)", np.nan
        ),
        "Feature FDR p": np.nan,
        "Interaction Wald chi-square": wald_statistic,
        "Interaction p": interaction_p,
        "Interaction FDR p": np.nan,
        "Main model status": main_model["status"],
        "Interaction model status": interaction_model["status"],
    }
    slopes = _diagnosis_specific_slopes(
        candidate,
        interaction_model,
        diagnosis_levels,
    )
    for row in slopes:
        row["Numeric covariates"] = ", ".join(numeric_covariates)
        row["Categorical covariates"] = ", ".join(categorical_covariates)
        row["Category references"] = "; ".join(category_references)
    return model_row, slopes


def _failed_candidate(identity, variant, status, data=None):
    return {
        **identity,
        "n": int(len(data)) if data is not None else 0,
        "Person clusters": (
            int(data["Person ID"].nunique()) if data is not None else 0
        ),
        "Diagnosis levels": "",
        "Numeric covariates": ", ".join(variant["numeric_covariates"]),
        "Categorical covariates": ", ".join(
            variant["categorical_covariates"]
        ),
        "Category references": "",
        "Feature beta per SD": np.nan,
        "Feature robust SE": np.nan,
        "Feature 95% CI lower": np.nan,
        "Feature 95% CI upper": np.nan,
        "Feature p": np.nan,
        "Feature FDR p": np.nan,
        "Interaction Wald chi-square": np.nan,
        "Interaction p": np.nan,
        "Interaction FDR p": np.nan,
        "Main model status": status,
        "Interaction model status": status,
    }


def _complete_cases(df, numeric_columns, categorical_columns):
    columns = list(dict.fromkeys([*numeric_columns, *categorical_columns]))
    data = df[columns].copy()
    for column in numeric_columns:
        data[column] = pd.to_numeric(data[column], errors="coerce")
    return data.replace([np.inf, -np.inf], np.nan).dropna()


def _attach_source_variables(analysis_df):
    output = analysis_df.copy()
    output["Ascertainment stratum"] = output["#Subject"].map(
        subject_source_cohort
    )
    output["Clinical collection"] = output["Ascertainment stratum"].map(
        {
            "COBEN": "NINR",
            "HC/HC2": "NU20",
            "pre-LBD/pre-LBD2": "NU20",
        }
    )
    if output["Clinical collection"].isna().any():
        unknown = sorted(
            output.loc[
                output["Clinical collection"].isna(),
                "Ascertainment stratum",
            ].astype(str).unique()
        )
        raise ValueError(f"Unmapped clinical source strata: {unknown}")
    return output


def _ordered_categories(values, column):
    present = set(values.dropna().astype(str))
    preferred = CATEGORY_ORDERS.get(column, ())
    ordered = [value for value in preferred if value in present]
    return ordered + sorted(present.difference(ordered))


def _adjust_model_p_values(frame):
    output = frame.copy()
    grouping = ["Dataset", "Analysis subset", "Model", "Clinical outcome"]
    output["Feature FDR p"] = output.groupby(grouping)["Feature p"].transform(
        _benjamini_hochberg
    )
    output["Interaction FDR p"] = output.groupby(grouping)[
        "Interaction p"
    ].transform(_benjamini_hochberg)
    return output


def _adjust_slope_p_values(frame):
    output = frame.copy()
    if output.empty:
        return output
    grouping = [
        "Dataset",
        "Analysis subset",
        "Model",
        "Clinical outcome",
        "Diagnosis",
    ]
    output["FDR p"] = output.groupby(grouping)["p"].transform(
        _benjamini_hochberg
    )
    return output


def _primary_reproduction_check(models_df, canonical_tables, tolerance=1e-10):
    rows = []
    primary = models_df[
        (models_df["Analysis subset"] == "All collections")
        & (models_df["Model"] == "primary")
        ]
    for dataset, tables in canonical_tables.items():
        canonical_adjusted = tables["adjusted"]
        canonical_interactions = tables["interactions"]
        current = primary[primary["Dataset"] == dataset]
        for row in current.to_dict("records"):
            mask = (
                    canonical_adjusted["Clinical outcome"].eq(
                        row["Clinical outcome"]
                    )
                    & canonical_adjusted["Representative feature"].eq(
                row["Representative feature"]
            )
            )
            interaction_mask = (
                    canonical_interactions["Clinical outcome"].eq(
                        row["Clinical outcome"]
                    )
                    & canonical_interactions["Representative feature"].eq(
                row["Representative feature"]
            )
            )
            old_main = canonical_adjusted.loc[mask]
            old_interaction = canonical_interactions.loc[interaction_mask]
            old_beta = (
                old_main.iloc[0]["Feature beta per SD"]
                if len(old_main) == 1
                else np.nan
            )
            old_interaction_p = (
                old_interaction.iloc[0]["p"]
                if len(old_interaction) == 1
                else np.nan
            )
            beta_delta = _absolute_delta(row["Feature beta per SD"], old_beta)
            interaction_delta = _absolute_delta(
                row["Interaction p"], old_interaction_p
            )
            reproduced = bool(
                np.isfinite(beta_delta)
                and np.isfinite(interaction_delta)
                and beta_delta <= tolerance
                and interaction_delta <= tolerance
            )
            rows.append(
                {
                    "Dataset": dataset,
                    "Clinical outcome": row["Clinical outcome"],
                    "Representative feature": row["Representative feature"],
                    "Canonical beta": old_beta,
                    "Reproduced beta": row["Feature beta per SD"],
                    "Absolute beta delta": beta_delta,
                    "Canonical interaction p": old_interaction_p,
                    "Reproduced interaction p": row["Interaction p"],
                    "Absolute interaction-p delta": interaction_delta,
                    "Reproduced": reproduced,
                }
            )
    return pd.DataFrame(rows)


def _reported_results(frame):
    rows = []
    for finding in REPORTED_FINDINGS:
        matches = frame[
            frame["Dataset"].eq(finding["dataset"])
            & frame["Clinical outcome"].eq(finding["outcome"])
            & frame["Representative feature"].eq(finding["feature"])
            ].copy()
        if matches.empty:
            continue
        matches.insert(0, "Finding role", finding["role"])
        rows.append(matches)
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def _source_counts(analysis_df, dataset):
    counts = (
        analysis_df.groupby(
            ["Clinical collection", "Ascertainment stratum", "Diagnosis"],
            dropna=False,
        )
        .agg(
            Visits=("#Subject", "size"),
            People=("Person ID", "nunique"),
        )
        .reset_index()
    )
    counts.insert(0, "Dataset", dataset)
    return counts


def _settings(output_dir, source_paths):
    rows = [
        {"Setting": "scenario", "Value": SCENARIO["label"]},
        {"Setting": "output_dir", "Value": str(output_dir)},
        {
            "Setting": "candidate_policy",
            "Value": (
                "Frozen representatives from canonical WASO-corrected "
                "follow-up workbooks; no rescreening or reselection"
            ),
        },
        {
            "Setting": "primary_model",
            "Value": (
                "Gaussian GEE; feature standardized within complete cases; "
                "diagnosis + sex + education; independence working "
                "correlation; robust sandwich covariance; person clusters"
            ),
        },
        {
            "Setting": "age_sensitivity",
            "Value": "Primary model plus standardized age",
        },
        {
            "Setting": "collection_sensitivity",
            "Value": "Primary model plus NINR versus NU20 collection",
        },
        {
            "Setting": "ascertainment_sensitivity",
            "Value": (
                "Primary model plus COBEN, HC/HC2, and pre-LBD/pre-LBD2 "
                "recruitment-stratum indicators"
            ),
        },
        {
            "Setting": "within_collection_sensitivity",
            "Value": (
                "Primary and age-adjusted models repeated separately in "
                "NINR and NU20"
            ),
        },
        {
            "Setting": "multiplicity",
            "Value": (
                "Benjamini-Hochberg FDR separately by dataset, analysis "
                "subset, model variant, and clinical outcome"
            ),
        },
        {
            "Setting": "clinical_collection_mapping",
            "Value": "COBEN -> NINR; HC/HC2 and pre-LBD/pre-LBD2 -> NU20",
        },
    ]
    for index, path in enumerate(source_paths, start=1):
        rows.append(
            {"Setting": f"canonical_source_{index}", "Value": str(path)}
        )
    return pd.DataFrame(rows)


def _style_sensitivity_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    significant_fill = PatternFill("solid", fgColor="E2F0D9")
    caution_fill = PatternFill("solid", fgColor="FFF2CC")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        worksheet.row_dimensions[1].height = 36
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        for column_index, column_cells in enumerate(
                worksheet.columns,
                start=1,
        ):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(max(max_length + 2, 12), 48)

        headers = {cell.value: cell.column for cell in worksheet[1]}
        for p_header in (
                "Feature FDR p",
                "Interaction FDR p",
                "FDR p",
        ):
            column = headers.get(p_header)
            if not column:
                continue
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, column)
                try:
                    value = float(cell.value)
                except (TypeError, ValueError):
                    continue
                cell.fill = significant_fill if value < 0.05 else caution_fill

    if "primary_reproduction" in workbook.sheetnames:
        worksheet = workbook["primary_reproduction"]
        reproduced_column = {
            cell.value: cell.column for cell in worksheet[1]
        }.get("Reproduced")
        if reproduced_column:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, reproduced_column)
                cell.fill = significant_fill if cell.value else caution_fill
                cell.font = Font(bold=True)


def _require_paths(*paths):
    missing = [str(path) for path in paths if not Path(path).exists()]
    if missing:
        raise FileNotFoundError(f"Missing canonical sensitivity inputs: {missing}")


def _absolute_delta(left, right):
    left = pd.to_numeric(pd.Series([left]), errors="coerce").iloc[0]
    right = pd.to_numeric(pd.Series([right]), errors="coerce").iloc[0]
    if pd.isna(left) and pd.isna(right):
        return 0.0
    if pd.isna(left) or pd.isna(right):
        return np.nan
    return float(abs(left - right))


def _default_output_dir():
    return RESULTS_ROOT / datetime.now().strftime("%Y%m%d_%H%M%S")
