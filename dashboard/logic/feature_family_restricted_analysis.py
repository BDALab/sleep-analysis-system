import json
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dashboard.logic.feature_correlation_analysis import (
    ANALYSIS_SCENARIOS,
    STATISTIC_PREFIXES,
    analyze_feature_dataset,
)
from dashboard.logic.feature_families import (
    ACTIVITY_EXTENSION_STABLE_FAMILY_IDS,
    feature_family_for_feature,
    feature_family_metadata,
    PRIMARY_SLEEP_STABLE_FAMILY_IDS,
    SECONDARY_LIFESTYLE_FAMILY_IDS,
)
from dashboard.logic.feature_family_followup_analysis import analyze_family_followup
from mysite.settings import MEDIA_ROOT

logger = logging.getLogger(__name__)

SCENARIO_KEY = "hc-vs-predlb"
SCENARIO = next(
    scenario
    for scenario in ANALYSIS_SCENARIOS
    if scenario["key"] == "predlb-vs-hc"
)
SELECTED_COVARIATES = ("gender", "education")
RESULTS_ROOT = Path(MEDIA_ROOT) / "feature-family-restricted-analysis"

PRIMARY_SLEEP_FAMILY_IDS = PRIMARY_SLEEP_STABLE_FAMILY_IDS
ACTIVITY_EXTENSION_FAMILY_IDS = ACTIVITY_EXTENSION_STABLE_FAMILY_IDS


def run_hc_vs_predlb_feature_family_restricted_analysis(output_dir=None):
    """Create stable-family datasets and rerun HC-vs-preDLB association follow-up."""
    output_dir = Path(output_dir) if output_dir else _default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)

    analysis_specs = _analysis_specs()
    result_rows = []
    for spec in analysis_specs:
        result_rows.append(_run_restricted_spec(spec, output_dir))

    summary_df = pd.DataFrame(result_rows)
    settings_df = _settings(output_dir, analysis_specs)
    output_path = output_dir / "feature_family_restricted_analysis_summary.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="analysis_summary", index=False)
        settings_df.to_excel(writer, sheet_name="settings", index=False)
        _style_workbook(writer.book)

    result = {
        "scenario": SCENARIO["label"],
        "run_dir": str(output_dir),
        "summary_path": str(output_path),
        "analysis_count": int(len(summary_df)),
        "analyses": result_rows,
    }
    (output_dir / "feature_family_restricted_analysis_summary.json").write_text(
        json.dumps(result, indent=2),
        encoding="utf-8",
    )
    logger.info(
        "Stable-family restricted HC-vs-preDLB analysis completed: %s",
        output_dir,
    )
    return result


def filter_grouped_stats_by_feature_families(
        source_path,
        output_path,
        allowed_family_ids,
):
    source_path = Path(source_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    grouped_df = pd.read_excel(source_path)
    metadata_columns = [
        column for column in grouped_df.columns if str(column).startswith("#")
    ]
    stat_columns = [
        column
        for column in grouped_df.columns
        if _is_stat_feature(column)
           and feature_family_for_feature(column).family_id in allowed_family_ids
    ]
    filtered_df = grouped_df[metadata_columns + stat_columns].copy()
    filtered_df.to_excel(output_path, index=False)

    manifest_df = pd.DataFrame(
        [feature_family_metadata(column) for column in stat_columns]
    )
    manifest_path = output_path.with_name(f"{output_path.stem}_feature_manifest.xlsx")
    manifest_df.to_excel(manifest_path, index=False)

    return {
        "source_path": str(source_path),
        "output_path": str(output_path),
        "feature_manifest_path": str(manifest_path),
        "row_count": int(len(filtered_df)),
        "metadata_column_count": int(len(metadata_columns)),
        "feature_count": int(len(stat_columns)),
        "family_ids": sorted(set(allowed_family_ids)),
        "family_counts": (
            manifest_df["Feature family ID"].value_counts().to_dict()
            if not manifest_df.empty
            else {}
        ),
    }


def _run_restricted_spec(spec, output_dir):
    spec_dir = output_dir / spec["key"]
    dataset_dir = spec_dir / "datasets"
    association_dir = spec_dir / "association"
    dataset_dir.mkdir(parents=True, exist_ok=True)
    association_dir.mkdir(parents=True, exist_ok=True)

    adjusted_filter = filter_grouped_stats_by_feature_families(
        spec["adjusted_grouped_path"],
        dataset_dir / "grouped_clinical_matrix_with_stats.xlsx",
        spec["family_ids"],
    )
    raw_filter = filter_grouped_stats_by_feature_families(
        spec["raw_grouped_path"],
        dataset_dir / "raw_grouped_clinical_matrix_with_stats.xlsx",
        spec["family_ids"],
    )

    correlation_path = association_dir / "feature_family_restricted_correlation.xlsx"
    correlation = analyze_feature_dataset(
        adjusted_filter["output_path"],
        dataset_name=spec["label"],
        scenario=SCENARIO,
        output_path=correlation_path,
        selected_covariates=SELECTED_COVARIATES,
    )
    followup_path = association_dir / "feature_family_restricted_followup.xlsx"
    followup = analyze_family_followup(
        raw_grouped_path=raw_filter["output_path"],
        family_workbook_path=correlation_path,
        output_path=followup_path,
        dataset_name=spec["label"],
        scenario=SCENARIO,
        selected_covariates=SELECTED_COVARIATES,
    )

    return {
        "analysis_key": spec["key"],
        "analysis_label": spec["label"],
        "dataset_name": spec["dataset_name"],
        "family_ids": ", ".join(sorted(spec["family_ids"])),
        "adjusted_dataset_path": adjusted_filter["output_path"],
        "raw_dataset_path": raw_filter["output_path"],
        "feature_manifest_path": adjusted_filter["feature_manifest_path"],
        "feature_count": adjusted_filter["feature_count"],
        "raw_feature_count": raw_filter["feature_count"],
        "correlation_path": str(correlation_path),
        "followup_path": str(followup_path),
        "significant_group_count": correlation["significant_group_count"],
        "significant_correlation_count": correlation["significant_correlation_count"],
        "group_family_candidate_count": correlation["group_family_candidate_count"],
        "correlation_family_candidate_count": correlation[
            "correlation_family_candidate_count"
        ],
        "followup_candidate_count": followup["followup_candidate_count"],
        "adjusted_result_count": followup["adjusted_result_count"],
        "interaction_result_count": followup["interaction_result_count"],
        "focused_plot_count": followup["focused_plot_count"],
        "family_counts_json": json.dumps(adjusted_filter["family_counts"], sort_keys=True),
    }


def _analysis_specs():
    clinical_run = (
            Path(MEDIA_ROOT)
            / "analysis-preparation"
            / "dataset-clinical"
            / "20260629_132528_560853"
    )
    acc_run = (
            Path(MEDIA_ROOT)
            / "analysis-preparation"
            / "dataset-clinical-acc"
            / "20260629_132533_149540"
    )
    return [
        {
            "key": "primary-sleep",
            "label": "Stable primary sleep families",
            "dataset_name": "dataset-clinical",
            "family_ids": PRIMARY_SLEEP_FAMILY_IDS,
            "adjusted_grouped_path": (
                    clinical_run
                    / "scenarios"
                    / "predlb-vs-hc"
                    / "data"
                    / "grouped_clinical_matrix_with_stats.xlsx"
            ),
            "raw_grouped_path": clinical_run / "raw" / "grouped_clinical_matrix_with_stats.xlsx",
        },
        {
            "key": "primary-sleep-plus-activity",
            "label": "Stable primary sleep + activity variability families",
            "dataset_name": "dataset-clinical-acc",
            "family_ids": PRIMARY_SLEEP_FAMILY_IDS | ACTIVITY_EXTENSION_FAMILY_IDS,
            "adjusted_grouped_path": (
                    acc_run
                    / "scenarios"
                    / "predlb-vs-hc"
                    / "data"
                    / "grouped_clinical_matrix_with_stats.xlsx"
            ),
            "raw_grouped_path": acc_run / "raw" / "grouped_clinical_matrix_with_stats.xlsx",
        },
    ]


def _settings(output_dir, analysis_specs):
    rows = [
        {"Setting": "scenario", "Value": SCENARIO["label"]},
        {"Setting": "output_dir", "Value": str(output_dir)},
        {"Setting": "selected_covariates", "Value": ", ".join(SELECTED_COVARIATES)},
        {
            "Setting": "primary_sleep_family_ids",
            "Value": ", ".join(sorted(PRIMARY_SLEEP_FAMILY_IDS)),
        },
        {
            "Setting": "activity_extension_family_ids",
            "Value": ", ".join(sorted(ACTIVITY_EXTENSION_FAMILY_IDS)),
        },
        {
            "Setting": "secondary_lifestyle_family_ids",
            "Value": ", ".join(sorted(SECONDARY_LIFESTYLE_FAMILY_IDS)),
        },
        {
            "Setting": "method",
            "Value": (
                "Filter canonical WASO-corrected grouped-statistics matrices "
                "to stable feature families, rerun Mann-Whitney/FDR, Spearman/FDR, "
                "and diagnosis-adjusted GEE follow-up for HC vs preDLB."
            ),
        },
    ]
    for spec in analysis_specs:
        rows.extend(
            [
                {
                    "Setting": f"{spec['key']}_adjusted_grouped_path",
                    "Value": str(spec["adjusted_grouped_path"]),
                },
                {
                    "Setting": f"{spec['key']}_raw_grouped_path",
                    "Value": str(spec["raw_grouped_path"]),
                },
            ]
        )
    return pd.DataFrame(rows)


def _is_stat_feature(column):
    return str(column).startswith(STATISTIC_PREFIXES)


def _default_output_dir():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return RESULTS_ROOT / SCENARIO_KEY / timestamp


def _style_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[:200]
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10),
                60,
            )
