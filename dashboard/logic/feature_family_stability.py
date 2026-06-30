import json
import logging
from datetime import datetime
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dashboard.logic.feature_families import (
    feature_family_metadata,
    list_feature_families,
)
from mysite.settings import MEDIA_ROOT

matplotlib.use("Agg")
import matplotlib.pyplot as plt

logger = logging.getLogger(__name__)

ALPHA = 0.05
TOP_N = 30
SCENARIO_KEY = "hc-vs-predlb"
SCENARIO_LABEL = "HC vs preDLB"
RESULTS_ROOT = Path(MEDIA_ROOT) / "feature-family-stability"


def analyze_hc_vs_predlb_feature_family_stability(output_dir=None):
    """Aggregate HC-vs-preDLB evidence by fixed feature family."""
    output_dir = Path(output_dir) if output_dir else _default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)

    source_definitions = _source_definitions()
    classification_frames = []
    association_frames = []
    source_rows = []

    for source in source_definitions:
        source_rows.append(_source_row(source))
        evidence = _load_source_evidence(source)
        if evidence.empty:
            continue
        if source["category"] == "classification":
            classification_frames.append(evidence)
        else:
            association_frames.append(evidence)

    classification_evidence = _concat(classification_frames)
    association_evidence = _concat(association_frames)
    all_evidence = _concat([classification_evidence, association_evidence])

    summary, matrix_counts, matrix_binary = _summarize_evidence(
        all_evidence,
        source_definitions,
    )
    interpretation = _interpretation_summary(summary)
    family_definitions = _family_definitions()
    source_files = pd.DataFrame(source_rows)
    settings = _settings(output_dir, source_definitions)

    output_path = output_dir / "feature_family_stability.xlsx"
    _write_workbook(
        output_path,
        {
            "interpretation_summary": interpretation,
            "evidence_matrix_counts": matrix_counts,
            "evidence_matrix_binary": matrix_binary,
            "feature_family_summary": summary,
            "classification_evidence": classification_evidence,
            "association_evidence": association_evidence,
            "family_definitions": family_definitions,
            "source_files": source_files,
            "settings": settings,
        },
    )

    heatmap_path = output_dir / "feature_family_stability_heatmap.png"
    heatmap_pdf_path = output_dir / "feature_family_stability_heatmap.pdf"
    _save_heatmap(matrix_binary, summary, source_definitions, heatmap_path)
    _save_heatmap(matrix_binary, summary, source_definitions, heatmap_pdf_path)

    result = {
        "scenario": SCENARIO_LABEL,
        "run_dir": str(output_dir),
        "workbook_path": str(output_path),
        "heatmap_path": str(heatmap_path),
        "heatmap_pdf_path": str(heatmap_pdf_path),
        "family_count": int(len(summary)),
        "stable_family_count": int(
            summary["Stability class"].isin(
                (
                    "primary_cross_method_stable",
                    "activity_enhanced_cross_method_stable",
                    "waso_corrected_cross_method",
                    "cross_method_supported",
                )
            ).sum()
        )
        if not summary.empty
        else 0,
    }
    (output_dir / "feature_family_stability_summary.json").write_text(
        json.dumps(result, indent=2),
        encoding="utf-8",
    )
    logger.info(
        "Feature-family stability completed for %s: families=%d output=%s",
        SCENARIO_LABEL,
        result["family_count"],
        output_path,
    )
    return result


def _default_output_dir():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return RESULTS_ROOT / SCENARIO_KEY / timestamp


def _media_path(relative_path):
    return Path(MEDIA_ROOT) / relative_path


def _source_definitions():
    strict_dir = _media_path(
        "classification/grouped-statistics-strict-with-covariates/"
        "dataset-clinical-rfe/20260629_173137/scenario-preDLB_vs_HC"
    )
    non_strict_clinical_dir = _media_path(
        "classification/grouped-statistics-with-covariates/"
        "dataset-clinical/all/20260630_103329/scenario-preDLB_vs_HC"
    )
    non_strict_acc_dir = _media_path(
        "classification/grouped-statistics-with-covariates/"
        "dataset-clinical-acc/all/20260630_095927/scenario-preDLB_vs_HC"
    )
    corr_clinical = _media_path(
        "analysis-preparation/dataset-clinical/20260629_132528_560853/"
        "scenarios/predlb-vs-hc/correlation/feature_clinical_correlation_matrix.xlsx"
    )
    gee_clinical = _media_path(
        "analysis-preparation/dataset-clinical/20260629_132528_560853/"
        "scenarios/predlb-vs-hc/correlation/feature_family_followup_analysis.xlsx"
    )
    corr_acc = _media_path(
        "analysis-preparation/dataset-clinical-acc/20260629_132533_149540/"
        "scenarios/predlb-vs-hc/correlation/feature_clinical_correlation_matrix.xlsx"
    )
    gee_acc = _media_path(
        "analysis-preparation/dataset-clinical-acc/20260629_132533_149540/"
        "scenarios/predlb-vs-hc/correlation/feature_family_followup_analysis.xlsx"
    )

    return [
        _classification_source(
            "strict_rfe_selected",
            "Strict RFE selected",
            strict_dir / "final_model" / "selected_feature_scores.xlsx",
            "strict_rfe_classifier",
            selected_only=True,
        ),
        _classification_source(
            "strict_rfe_importance_top30",
            "Strict RFE importance top 30",
            strict_dir / "final_model" / "feature_importances.xlsx",
            "strict_rfe_classifier",
            score_column="importance",
            top_n=TOP_N,
        ),
        _classification_source(
            "strict_rfe_shap_top30",
            "Strict RFE SHAP top 30",
            strict_dir / "final_model" / "shap_feature_importances.xlsx",
            "strict_rfe_classifier",
            score_column="mean_abs_shap",
            top_n=TOP_N,
        ),
        _classification_source(
            "non_strict_clinical_selected",
            "Non-strict clinical selected",
            non_strict_clinical_dir / "selected_feature_scores.xlsx",
            "non_strict_clinical_classifier",
            selected_only=True,
        ),
        _classification_source(
            "non_strict_clinical_importance_top30",
            "Non-strict clinical importance top 30",
            non_strict_clinical_dir / "feature_importances.xlsx",
            "non_strict_clinical_classifier",
            score_column="importance",
            top_n=TOP_N,
        ),
        _classification_source(
            "non_strict_clinical_shap_top30",
            "Non-strict clinical SHAP top 30",
            non_strict_clinical_dir / "shap_feature_importances.xlsx",
            "non_strict_clinical_classifier",
            score_column="mean_abs_shap",
            top_n=TOP_N,
        ),
        _classification_source(
            "non_strict_acc_selected",
            "Non-strict activity selected",
            non_strict_acc_dir / "selected_feature_scores.xlsx",
            "non_strict_activity_classifier",
            selected_only=True,
        ),
        _classification_source(
            "non_strict_acc_importance_top30",
            "Non-strict activity importance top 30",
            non_strict_acc_dir / "feature_importances.xlsx",
            "non_strict_activity_classifier",
            score_column="importance",
            top_n=TOP_N,
        ),
        _classification_source(
            "non_strict_acc_shap_top30",
            "Non-strict activity SHAP top 30",
            non_strict_acc_dir / "shap_feature_importances.xlsx",
            "non_strict_activity_classifier",
            score_column="mean_abs_shap",
            top_n=TOP_N,
        ),
        _association_source(
            "clinical_group_fdr",
            "Clinical group FDR",
            corr_clinical,
            "significant_group",
            "Features",
            "group_difference",
            "dataset-clinical",
        ),
        _association_source(
            "clinical_correlation_fdr",
            "Clinical outcome FDR",
            corr_clinical,
            "significant_correlations",
            "Features",
            "clinical_correlation",
            "dataset-clinical",
        ),
        _association_source(
            "clinical_followup_candidates",
            "Clinical follow-up candidates",
            gee_clinical,
            "candidate_pairs",
            "Representative feature",
            "followup_candidate",
            "dataset-clinical",
        ),
        _association_source(
            "clinical_gee_adjusted_fdr",
            "Clinical adjusted GEE FDR",
            gee_clinical,
            "adjusted_associations",
            "Representative feature",
            "gee_adjusted",
            "dataset-clinical",
            fdr_column="FDR p",
        ),
        _association_source(
            "clinical_gee_interaction_fdr",
            "Clinical GEE interaction FDR",
            gee_clinical,
            "diagnosis_interactions",
            "Representative feature",
            "gee_interaction",
            "dataset-clinical",
            fdr_column="FDR p",
        ),
        _association_source(
            "acc_group_fdr",
            "Activity group FDR",
            corr_acc,
            "significant_group",
            "Features",
            "group_difference",
            "dataset-clinical-acc",
        ),
        _association_source(
            "acc_correlation_fdr",
            "Activity outcome FDR",
            corr_acc,
            "significant_correlations",
            "Features",
            "clinical_correlation",
            "dataset-clinical-acc",
        ),
        _association_source(
            "acc_followup_candidates",
            "Activity follow-up candidates",
            gee_acc,
            "candidate_pairs",
            "Representative feature",
            "followup_candidate",
            "dataset-clinical-acc",
        ),
        _association_source(
            "acc_gee_adjusted_fdr",
            "Activity adjusted GEE FDR",
            gee_acc,
            "adjusted_associations",
            "Representative feature",
            "gee_adjusted",
            "dataset-clinical-acc",
            fdr_column="FDR p",
        ),
        _association_source(
            "acc_gee_interaction_fdr",
            "Activity GEE interaction FDR",
            gee_acc,
            "diagnosis_interactions",
            "Representative feature",
            "gee_interaction",
            "dataset-clinical-acc",
            fdr_column="FDR p",
        ),
    ]


def _classification_source(
        source_key,
        label,
        path,
        source_group,
        selected_only=False,
        score_column=None,
        top_n=None,
):
    return {
        "source_key": source_key,
        "source_label": label,
        "category": "classification",
        "source_group": source_group,
        "dataset": "dataset-clinical-acc" if "_acc_" in source_key else "dataset-clinical",
        "path": Path(path),
        "sheet_name": "Sheet1",
        "feature_column": "feature",
        "selected_only": selected_only,
        "score_column": score_column,
        "top_n": top_n,
        "fdr_column": None,
    }


def _association_source(
        source_key,
        label,
        path,
        sheet_name,
        feature_column,
        source_group,
        dataset,
        fdr_column=None,
):
    return {
        "source_key": source_key,
        "source_label": label,
        "category": "association",
        "source_group": source_group,
        "dataset": dataset,
        "path": Path(path),
        "sheet_name": sheet_name,
        "feature_column": feature_column,
        "selected_only": False,
        "score_column": None,
        "top_n": None,
        "fdr_column": fdr_column,
    }


def _load_source_evidence(source):
    path = source["path"]
    if not path.exists():
        logger.warning("Feature-family stability source missing: %s", path)
        return pd.DataFrame()

    try:
        df = pd.read_excel(path, sheet_name=source["sheet_name"])
    except ValueError:
        logger.warning(
            "Feature-family stability sheet missing: %s in %s",
            source["sheet_name"],
            path,
        )
        return pd.DataFrame()

    if source["feature_column"] not in df.columns:
        logger.warning(
            "Feature-family stability feature column missing: %s in %s",
            source["feature_column"],
            path,
        )
        return pd.DataFrame()

    if source["selected_only"] and "selected" in df.columns:
        df = df[df["selected"].astype(bool)].copy()

    if source["fdr_column"] and source["fdr_column"] in df.columns:
        df = df[pd.to_numeric(df[source["fdr_column"]], errors="coerce") < ALPHA]
        if "Status" in df.columns:
            df = df[df["Status"].astype(str).str.lower().eq("ok")]
        df = df.copy()

    if source["score_column"] and source["score_column"] in df.columns:
        df[source["score_column"]] = pd.to_numeric(
            df[source["score_column"]],
            errors="coerce",
        )
        df = df.sort_values(source["score_column"], ascending=False)

    if source["top_n"]:
        df = df.head(source["top_n"]).copy()

    if df.empty:
        return pd.DataFrame()

    df = df.rename(columns={source["feature_column"]: "Feature"}).copy()
    df = df.loc[:, ~df.columns.duplicated()].copy()
    metadata_columns = {
        "Feature family ID",
        "Feature family",
        "Feature family domain",
        "Feature family role",
        "Source",
        "Nightly summary",
        "Measurement",
        "Normalized source",
    }
    df = df.drop(
        columns=[column for column in metadata_columns if column in df.columns],
        errors="ignore",
    )
    metadata = pd.DataFrame(
        [feature_family_metadata(feature) for feature in df["Feature"]]
    )
    metadata = metadata.rename(columns={"Features": "Feature"})
    output = pd.concat([df.reset_index(drop=True), metadata.drop(columns=["Feature"]).reset_index(drop=True)], axis=1)

    output.insert(0, "Evidence source", source["source_key"])
    output.insert(1, "Evidence label", source["source_label"])
    output.insert(2, "Evidence category", source["category"])
    output.insert(3, "Evidence group", source["source_group"])
    output.insert(4, "Dataset", source["dataset"])
    output.insert(5, "Source workbook", str(path))
    output.insert(6, "Source sheet", source["sheet_name"])
    return output


def _summarize_evidence(all_evidence, source_definitions):
    source_keys = [source["source_key"] for source in source_definitions]
    family_columns = [
        "Feature family ID",
        "Feature family",
        "Feature family domain",
        "Feature family role",
    ]
    if all_evidence.empty:
        empty = pd.DataFrame(columns=family_columns)
        return empty, empty, empty

    counts = (
        all_evidence.groupby(family_columns + ["Evidence source"])
        .size()
        .reset_index(name="Feature hits")
    )
    matrix_counts = counts.pivot_table(
        index=family_columns,
        columns="Evidence source",
        values="Feature hits",
        fill_value=0,
        aggfunc="sum",
    ).reset_index()
    for source_key in source_keys:
        if source_key not in matrix_counts.columns:
            matrix_counts[source_key] = 0
    matrix_counts = matrix_counts[family_columns + source_keys]
    matrix_binary = matrix_counts.copy()
    matrix_binary[source_keys] = (matrix_binary[source_keys] > 0).astype(int)

    group_stats = []
    for family_values, rows in all_evidence.groupby(family_columns, sort=True):
        row = dict(zip(family_columns, family_values))
        classification = rows[rows["Evidence category"] == "classification"]
        association = rows[rows["Evidence category"] == "association"]
        source_presence = rows["Evidence source"].nunique()
        classification_sources = classification["Evidence source"].nunique()
        association_sources = association["Evidence source"].nunique()
        classification_run_support = classification[
            "Evidence group"
        ].nunique()
        association_method_support = association[
            "Evidence group"
        ].nunique()
        dataset_support = rows["Dataset"].nunique()
        total_hits = len(rows)
        row.update(
            {
                "Evidence source count": int(source_presence),
                "Classification source count": int(classification_sources),
                "Association source count": int(association_sources),
                "Classification run support": int(classification_run_support),
                "Association method support": int(association_method_support),
                "Dataset support": int(dataset_support),
                "Total feature hits": int(total_hits),
                "Unique feature count": int(rows["Feature"].nunique()),
            }
        )
        row["Stability class"] = classify_stability(row)
        row["Interpretation"] = _family_interpretation(row)
        group_stats.append(row)

    summary = pd.DataFrame(group_stats)
    summary = summary.merge(
        matrix_counts,
        on=family_columns,
        how="left",
    )
    numeric_cols = [
                       "Evidence source count",
                       "Classification source count",
                       "Association source count",
                       "Classification run support",
                       "Association method support",
                       "Dataset support",
                       "Total feature hits",
                       "Unique feature count",
                   ] + source_keys
    for col in numeric_cols:
        if col in summary.columns:
            summary[col] = summary[col].fillna(0).astype(int)

    class_order = {
        "primary_cross_method_stable": 0,
        "activity_enhanced_cross_method_stable": 1,
        "waso_corrected_cross_method": 2,
        "cross_method_supported": 3,
        "secondary_lifestyle_signal": 4,
        "classification_supported": 5,
        "association_supported": 6,
        "exploratory": 7,
    }
    summary["_class_order"] = summary["Stability class"].map(class_order).fillna(99)
    summary = summary.sort_values(
        [
            "_class_order",
            "Evidence source count",
            "Association method support",
            "Classification run support",
            "Total feature hits",
            "Feature family",
        ],
        ascending=[True, False, False, False, False, True],
    ).drop(columns=["_class_order"]).reset_index(drop=True)

    matrix_counts = matrix_counts.merge(
        summary[family_columns + ["Stability class", "Evidence source count", "Total feature hits"]],
        on=family_columns,
        how="left",
    )
    matrix_binary = matrix_binary.merge(
        summary[family_columns + ["Stability class", "Evidence source count", "Total feature hits"]],
        on=family_columns,
        how="left",
    )
    matrix_counts = _sort_matrix(matrix_counts)
    matrix_binary = _sort_matrix(matrix_binary)
    return summary, matrix_counts, matrix_binary


def classify_stability(row):
    family_id = row.get("Feature family ID", "")
    domain = row.get("Feature family domain", "")
    role = row.get("Feature family role", "")
    classification_runs = int(row.get("Classification run support", 0))
    association_methods = int(row.get("Association method support", 0))
    classification_sources = int(row.get("Classification source count", 0))
    association_sources = int(row.get("Association source count", 0))

    if domain == "diary_lifestyle":
        return "secondary_lifestyle_signal"
    if family_id == "waso" and classification_runs >= 1 and association_methods >= 1:
        return "waso_corrected_cross_method"
    if (
            role == "primary_activity_enhanced"
            and classification_runs >= 1
            and association_methods >= 2
    ):
        return "activity_enhanced_cross_method_stable"
    if (
            role.startswith("primary")
            and classification_runs >= 2
            and association_methods >= 2
    ):
        return "primary_cross_method_stable"
    if classification_runs >= 1 and association_methods >= 1:
        return "cross_method_supported"
    if classification_sources >= 2:
        return "classification_supported"
    if association_sources >= 2:
        return "association_supported"
    return "exploratory"


def _family_interpretation(row):
    family_id = row.get("Feature family ID")
    family = row.get("Feature family")
    stability = row.get("Stability class")
    if family_id == "long_awakenings":
        return (
            "Main sleep-continuity candidate. Recurrent evidence supports the "
            "interpretation that longer nocturnal awakenings differ between HC "
            "and preDLB and relate to clinical scales."
        )
    if family_id == "sleep_onset_latency":
        return (
            "Stable sleep-timing candidate. Interpret cautiously where evidence "
            "comes from normalized or derived variants, but the family-level "
            "signal is recurrent."
        )
    if family_id == "waso":
        return (
            "WASO-corrected family remains visible across methods, but the old "
            "normalized-WASO/MFS focused candidate should remain downgraded."
        )
    if family_id == "sleep_efficiency":
        return (
            "Supportive sleep-continuity candidate. Useful as part of the broader "
            "sleep-disruption pattern rather than as a standalone biomarker."
        )
    if family_id == "wake_bouts":
        return (
            "Supportive awakening-frequency family. More sensitive to sleep/wake "
            "classification noise than long-awakening duration."
        )
    if family_id == "activity_variability":
        return (
            "Activity-enhanced candidate. Evidence is strongest when activity "
            "features are included, so it should be reported separately from "
            "sleep/diary-only models."
        )
    if row.get("Feature family domain") == "diary_lifestyle":
        return (
            "Secondary and confounding-sensitive. Useful for prediction "
            "sensitivity, but not a primary disease-physiology claim."
        )
    return f"{family} is classified as {stability}; treat as exploratory unless clinically prespecified."


def _interpretation_summary(summary):
    if summary.empty:
        return summary
    columns = [
        "Feature family",
        "Feature family ID",
        "Feature family domain",
        "Feature family role",
        "Stability class",
        "Evidence source count",
        "Classification run support",
        "Association method support",
        "Dataset support",
        "Total feature hits",
        "Unique feature count",
        "Interpretation",
    ]
    return summary[columns].copy()


def _sort_matrix(matrix):
    if matrix.empty:
        return matrix
    return matrix.sort_values(
        ["Evidence source count", "Total feature hits", "Feature family"],
        ascending=[False, False, True],
    ).reset_index(drop=True)


def _family_definitions():
    return pd.DataFrame(
        [
            {
                "Feature family ID": family.family_id,
                "Feature family": family.label,
                "Domain": family.domain,
                "Role": family.role,
                "Description": family.description,
            }
            for family in list_feature_families()
        ]
    )


def _source_row(source):
    return {
        "Evidence source": source["source_key"],
        "Evidence label": source["source_label"],
        "Evidence category": source["category"],
        "Evidence group": source["source_group"],
        "Dataset": source["dataset"],
        "Workbook": str(source["path"]),
        "Sheet": source["sheet_name"],
        "Feature column": source["feature_column"],
        "Selected only": source["selected_only"],
        "Score column": source["score_column"] or "",
        "Top n": source["top_n"] or "",
        "FDR column": source["fdr_column"] or "",
    }


def _settings(output_dir, source_definitions):
    return pd.DataFrame(
        [
            {"Setting": "scenario", "Value": SCENARIO_LABEL},
            {"Setting": "output_dir", "Value": str(output_dir)},
            {"Setting": "alpha", "Value": ALPHA},
            {"Setting": "classification_top_n", "Value": TOP_N},
            {
                "Setting": "source_count",
                "Value": len(source_definitions),
            },
            {
                "Setting": "important_limitation",
                "Value": (
                    "Strict RFE evidence is based on final full-data model "
                    "selection/importances, not per-outer-fold selection frequency."
                ),
            },
            {
                "Setting": "interpretation_scope",
                "Value": (
                    "Cross-analysis family consensus for HC vs preDLB using "
                    "canonical WASO-corrected outputs."
                ),
            },
        ]
    )


def _write_workbook(output_path, sheets):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
    _style_workbook(output_path)


def _style_workbook(output_path):
    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[:200]
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10),
                55,
            )
    workbook.save(output_path)


def _save_heatmap(matrix_binary, summary, source_definitions, output_path):
    if matrix_binary.empty:
        return
    source_keys = [source["source_key"] for source in source_definitions]
    present_source_keys = [
        key for key in source_keys if key in matrix_binary.columns and matrix_binary[key].sum() > 0
    ]
    plot_df = matrix_binary[
        ["Feature family", "Stability class"] + present_source_keys
        ].copy()
    plot_df = plot_df[plot_df[present_source_keys].sum(axis=1) > 0]
    plot_df = plot_df.head(18)
    data = plot_df[present_source_keys].to_numpy(dtype=float)

    labels = {
        source["source_key"]: source["source_label"].replace(" ", "\n")
        for source in source_definitions
    }
    x_labels = [labels.get(key, key) for key in present_source_keys]
    y_labels = [
        f"{row['Feature family']}\n{row['Stability class']}"
        for _, row in plot_df.iterrows()
    ]

    width = max(10, len(present_source_keys) * 0.85)
    height = max(6, len(plot_df) * 0.55)
    fig, ax = plt.subplots(figsize=(width, height))
    im = ax.imshow(data, aspect="auto", cmap="YlGnBu", vmin=0, vmax=1)
    ax.set_xticks(np.arange(len(present_source_keys)))
    ax.set_xticklabels(x_labels, rotation=45, ha="right", fontsize=8)
    ax.set_yticks(np.arange(len(plot_df)))
    ax.set_yticklabels(y_labels, fontsize=9)
    ax.set_title("HC vs preDLB feature-family stability across analysis sources")
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            ax.text(
                j,
                i,
                "1" if data[i, j] else "",
                ha="center",
                va="center",
                fontsize=8,
                color="white" if data[i, j] else "#102A43",
            )
    ax.set_xlabel("Evidence source")
    ax.set_ylabel("Feature family")
    fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02, label="Evidence present")
    fig.tight_layout()
    fig.savefig(output_path, dpi=220 if output_path.suffix.lower() == ".png" else None, bbox_inches="tight")
    plt.close(fig)


def _concat(frames):
    frames = [frame for frame in frames if frame is not None and not frame.empty]
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)
