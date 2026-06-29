import logging
import re
import textwrap
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import chi2, norm, rankdata
from statsmodels.genmod.cov_struct import Independence
from statsmodels.genmod.families import Gaussian
from statsmodels.genmod.generalized_estimating_equations import GEE

from dashboard.models import Subject

matplotlib.use("Agg")
from matplotlib import pyplot as plt

logger = logging.getLogger(__name__)

ALPHA = 0.05
MIN_GROUP_N = 5
DIAGNOSIS_LABELS = {
    0: "HC",
    2: "MCI-AD",
    3: "preDLB",
}
CLINICAL_OUTCOMES = {
    "RBDq": "rbdq",
    "UPDRS": "updrs",
    "MFS": "mfs",
    "Visuospatial": "visuospatial",
    "Attention": "attention",
    "Executive": "executive",
}
COVARIATE_COLUMNS = {
    "age": "#Age",
    "gender": "#Gender",
    "education": "#Education",
}
DIAGNOSIS_COLORS = {
    "HC": "#286983",
    "MCI-AD": "#D28B26",
    "preDLB": "#B84A3C",
}
FOCUSED_PLOT_SPECS = (
    {
        "dataset": "dataset-clinical-acc",
        "scenario": "predlb-mci-vs-hc",
        "outcome": "UPDRS",
        "feature": "IQR.activity.Median Absolute Deviation",
        "title": "Activity variability and motor impairment",
        "x_label": "IQR of nightly activity median absolute deviation",
    },
    {
        "dataset": "dataset-clinical-acc",
        "scenario": "predlb-mci-vs-hc",
        "outcome": "Attention",
        "feature": "IQR.activity.Relative Interdencile Range",
        "title": "Activity variability and attention",
        "x_label": "IQR of nightly relative interdecile range",
    },
    {
        "dataset": "dataset-clinical-acc",
        "scenario": "predlb-mci-vs-hc",
        "outcome": "Executive",
        "feature": "SD.actigraphy_norm.Awakening > 5 minutes",
        "title": "Nocturnal awakening variability and executive performance",
        "x_label": "SD of normalized awakenings longer than 5 minutes",
    },
    {
        "dataset": "dataset-clinical",
        "scenario": "predlb-mci-vs-hc",
        "outcome": "MFS",
        "feature": "Range.actigraphy_norm.Wake after sleep onset",
        "title": (
            "Wake-after-sleep-onset variability and cognitive fluctuations"
        ),
        "x_label": "Range of normalized actigraphy wake after sleep onset",
    },
    {
        "dataset": "dataset-clinical",
        "scenario": "predlb-vs-hc",
        "outcome": "RBDq",
        "feature": "Max.actigraphy.Wake bouts",
        "title": "Wake bouts and RBD symptoms",
        "x_label": "Maximum nightly actigraphy wake bouts",
    },
    {
        "dataset": "dataset-clinical",
        "scenario": "predlb-vs-hc",
        "outcome": "RBDq",
        "feature": "Median.actigraphy_norm.Sleep efficiency",
        "title": "Sleep efficiency and RBD symptoms",
        "x_label": "Median normalized actigraphy sleep efficiency",
    },
    {
        "dataset": "dataset-clinical",
        "scenario": "predlb-vs-hc",
        "outcome": "RBDq",
        "feature": "Median.actigraphy.Wake after sleep onset",
        "title": "Wake after sleep onset and RBD symptoms",
        "x_label": "Median actigraphy wake after sleep onset",
    },
    {
        "dataset": "dataset-clinical-acc",
        "scenario": "predlb-vs-hc",
        "outcome": "UPDRS",
        "feature": "IQR.activity.Median Absolute Deviation",
        "title": "Activity variability and motor impairment",
        "x_label": "IQR of nightly activity median absolute deviation",
    },
    {
        "dataset": "dataset-clinical-acc",
        "scenario": "predlb-vs-hc",
        "outcome": "Attention",
        "feature": "SD.activity.Relative Interdencile Range",
        "title": "Activity variability and attention",
        "x_label": "SD of nightly relative interdecile range",
    },
    {
        "dataset": "dataset-clinical-acc",
        "scenario": "predlb-vs-hc",
        "outcome": "Visuospatial",
        "feature": "Min.activity.80th Percentile",
        "title": "Activity level and visuospatial performance",
        "x_label": "Minimum nightly activity 80th percentile",
    },
    {
        "dataset": "dataset-clinical",
        "scenario": "predlb-vs-hc",
        "outcome": "Executive",
        "feature": "MAD.actigraphy_norm.Awakening > 5 minutes",
        "title": (
            "Nocturnal awakening variability and executive performance"
        ),
        "x_label": "MAD of normalized awakenings longer than 5 minutes",
    },
    {
        "dataset": "dataset-clinical",
        "scenario": "predlb-vs-hc",
        "outcome": "MFS",
        "feature": "Range.actigraphy_norm.Wake after sleep onset",
        "title": (
            "Wake-after-sleep-onset variability and cognitive fluctuations"
        ),
        "x_label": "Range of normalized actigraphy wake after sleep onset",
    },
)


def analyze_family_followup(
        raw_grouped_path,
        family_workbook_path,
        output_path,
        dataset_name,
        scenario,
        selected_covariates=(),
):
    raw_grouped_path = Path(raw_grouped_path)
    family_workbook_path = Path(family_workbook_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    grouped_df = pd.read_excel(raw_grouped_path)
    candidate_df = pd.read_excel(
        family_workbook_path,
        sheet_name="corr_family_candidates",
    )
    analysis_df = _prepare_analysis_data(grouped_df, scenario)
    candidate_pairs = _candidate_pairs(candidate_df)

    stratified_df = _stratified_correlations(analysis_df, candidate_pairs)
    adjusted_df, interactions_df, slopes_df = _adjusted_models(
        analysis_df,
        candidate_pairs,
        selected_covariates,
    )
    summary_df = _interpret_results(
        candidate_pairs,
        stratified_df,
        adjusted_df,
        interactions_df,
    )
    plot_index_df = _create_focused_plots(
        analysis_df=analysis_df,
        candidate_pairs=candidate_pairs,
        adjusted_df=adjusted_df,
        interactions_df=interactions_df,
        selected_covariates=selected_covariates,
        dataset_name=dataset_name,
        scenario=scenario,
        output_dir=output_path.parent / "focused_plots",
    )
    settings_df = _settings(
        raw_grouped_path,
        family_workbook_path,
        dataset_name,
        scenario,
        selected_covariates,
        analysis_df,
        candidate_pairs,
        summary_df,
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(
            writer,
            sheet_name="interpretation_summary",
            index=False,
        )
        stratified_df.to_excel(
            writer,
            sheet_name="stratified_correlations",
            index=False,
        )
        adjusted_df.to_excel(
            writer,
            sheet_name="adjusted_associations",
            index=False,
        )
        interactions_df.to_excel(
            writer,
            sheet_name="diagnosis_interactions",
            index=False,
        )
        slopes_df.to_excel(
            writer,
            sheet_name="diagnosis_specific_slopes",
            index=False,
        )
        candidate_pairs.to_excel(
            writer,
            sheet_name="candidate_pairs",
            index=False,
        )
        plot_index_df.to_excel(
            writer,
            sheet_name="focused_plot_index",
            index=False,
        )
        settings_df.to_excel(writer, sheet_name="settings", index=False)
        _style_workbook(writer.book)

    interpretation_counts = (
        summary_df["Interpretation"].value_counts().to_dict()
        if not summary_df.empty
        else {}
    )
    logger.info(
        "Feature-family follow-up completed for %s / %s: candidates=%d, "
        "stratified=%d, adjusted=%d, interactions=%d, output=%s",
        dataset_name,
        scenario["label"],
        len(candidate_pairs),
        len(stratified_df),
        len(adjusted_df),
        len(interactions_df),
        output_path,
    )
    return {
        "followup_output_path": str(output_path),
        "followup_candidate_count": int(len(candidate_pairs)),
        "stratified_result_count": int(len(stratified_df)),
        "adjusted_result_count": int(len(adjusted_df)),
        "interaction_result_count": int(len(interactions_df)),
        "interpretation_counts": interpretation_counts,
        "focused_plot_count": int(len(plot_index_df)),
        "focused_plot_paths": plot_index_df.get(
            "PNG path",
            pd.Series(dtype=str),
        ).tolist(),
    }


def _prepare_analysis_data(grouped_df, scenario):
    prepared = grouped_df.copy()
    if "#Subject" not in prepared.columns:
        raise KeyError("Raw grouped dataset is missing #Subject")
    prepared["#Subject"] = prepared["#Subject"].astype(str).str.strip()
    codes = prepared["#Subject"].dropna().unique().tolist()
    subject_rows = list(
        Subject.objects.filter(code__in=codes).values(
            "code",
            "diagnosis_code",
            *CLINICAL_OUTCOMES.values(),
        )
    )
    subject_df = pd.DataFrame(subject_rows).rename(columns={"code": "#Subject"})
    prepared = prepared.merge(subject_df, on="#Subject", how="inner")
    allowed_codes = set(scenario["reference_codes"]) | set(scenario["case_codes"])
    prepared = prepared[
        prepared["diagnosis_code"].isin(allowed_codes)
    ].copy()
    prepared["Diagnosis"] = prepared["diagnosis_code"].map(DIAGNOSIS_LABELS)
    prepared["Person ID"] = prepared["#Subject"].map(_person_id)
    for column in (
            *CLINICAL_OUTCOMES.values(),
            *COVARIATE_COLUMNS.values(),
    ):
        if column in prepared.columns:
            prepared[column] = pd.to_numeric(prepared[column], errors="coerce")
    return prepared


def _candidate_pairs(candidate_df):
    columns = [
        "Clinical outcome",
        "Feature family",
        "Redundancy cluster",
        "Representative feature",
        "Pooled n",
        "Pooled Spearman rho",
        "Pooled FDR p",
    ]
    if candidate_df.empty:
        return pd.DataFrame(columns=columns)
    pairs = candidate_df.rename(
        columns={
            "n": "Pooled n",
            "Spearman rho": "Pooled Spearman rho",
            "FDR p": "Pooled FDR p",
        }
    )
    return pairs[columns].drop_duplicates(
        ["Clinical outcome", "Representative feature"]
    ).reset_index(drop=True)


def _stratified_correlations(analysis_df, candidate_pairs):
    columns = [
        "Clinical outcome",
        "Feature family",
        "Redundancy cluster",
        "Representative feature",
        "Diagnosis",
        "n",
        "Person clusters",
        "Spearman rho",
        "Robust SE",
        "95% CI lower",
        "95% CI upper",
        "p",
        "FDR p",
        "Status",
    ]
    rows = []
    diagnoses = [
        label
        for code, label in DIAGNOSIS_LABELS.items()
        if code in set(analysis_df["diagnosis_code"].dropna())
    ]
    for candidate in candidate_pairs.to_dict("records"):
        outcome_field = CLINICAL_OUTCOMES[candidate["Clinical outcome"]]
        feature = candidate["Representative feature"]
        if feature not in analysis_df.columns:
            rows.append(
                _empty_stratified_row(
                    candidate,
                    "",
                    "feature missing from raw grouped dataset",
                )
            )
            continue
        for diagnosis in diagnoses:
            data = _complete_cases(
                analysis_df[analysis_df["Diagnosis"] == diagnosis],
                [feature, outcome_field],
            )
            if (
                    len(data) < MIN_GROUP_N
                    or data[feature].nunique() < 2
                    or data[outcome_field].nunique() < 2
            ):
                rows.append(
                    {
                        **_candidate_identity(candidate),
                        "Diagnosis": diagnosis,
                        "n": int(len(data)),
                        "Person clusters": int(data["Person ID"].nunique()),
                        "Spearman rho": np.nan,
                        "Robust SE": np.nan,
                        "95% CI lower": np.nan,
                        "95% CI upper": np.nan,
                        "p": np.nan,
                        "FDR p": np.nan,
                        "Status": "insufficient or constant data",
                    }
                )
                continue

            x_rank = _standardize(rankdata(data[feature], method="average"))
            y_rank = _standardize(
                rankdata(data[outcome_field], method="average")
            )
            model = _fit_gee(
                y_rank,
                pd.DataFrame({"Spearman rho": x_rank}),
                data["Person ID"],
            )
            estimate = model["coefficients"].get("Spearman rho", np.nan)
            rows.append(
                {
                    **_candidate_identity(candidate),
                    "Diagnosis": diagnosis,
                    "n": model["n"],
                    "Person clusters": model["cluster_count"],
                    "Spearman rho": estimate,
                    "Robust SE": model["standard_errors"].get(
                        "Spearman rho",
                        np.nan,
                    ),
                    "95% CI lower": model["ci_lower"].get(
                        "Spearman rho",
                        np.nan,
                    ),
                    "95% CI upper": model["ci_upper"].get(
                        "Spearman rho",
                        np.nan,
                    ),
                    "p": model["p_values"].get("Spearman rho", np.nan),
                    "FDR p": np.nan,
                    "Status": model["status"],
                }
            )
    output = pd.DataFrame(rows, columns=columns)
    if not output.empty:
        output["FDR p"] = output.groupby("Clinical outcome")["p"].transform(
            _benjamini_hochberg
        )
    return output


def _adjusted_models(analysis_df, candidate_pairs, selected_covariates):
    adjusted_rows = []
    interaction_rows = []
    slope_rows = []
    diagnosis_levels = _diagnosis_levels(analysis_df)
    reference_diagnosis = diagnosis_levels[0] if diagnosis_levels else "HC"

    for candidate in candidate_pairs.to_dict("records"):
        outcome_name = candidate["Clinical outcome"]
        outcome_field = CLINICAL_OUTCOMES[outcome_name]
        feature = candidate["Representative feature"]
        covariate_columns = [
            COVARIATE_COLUMNS[covariate]
            for covariate in selected_covariates
            if COVARIATE_COLUMNS.get(covariate) in analysis_df.columns
        ]
        required = [
            feature,
            outcome_field,
            "Diagnosis",
            "Person ID",
            *covariate_columns,
        ]
        if feature not in analysis_df.columns:
            adjusted_rows.append(
                _failed_model_row(
                    candidate,
                    selected_covariates,
                    "feature missing from raw grouped dataset",
                )
            )
            interaction_rows.append(
                _failed_interaction_row(
                    candidate,
                    selected_covariates,
                    "feature missing from raw grouped dataset",
                )
            )
            continue

        data = _complete_cases(analysis_df, required)
        if (
                len(data) < MIN_GROUP_N
                or data[feature].nunique() < 2
                or data[outcome_field].nunique() < 2
        ):
            status = "insufficient or constant data"
            adjusted_rows.append(
                _failed_model_row(candidate, selected_covariates, status, data)
            )
            interaction_rows.append(
                _failed_interaction_row(
                    candidate,
                    selected_covariates,
                    status,
                    data,
                )
            )
            continue

        feature_values = _standardize(data[feature].to_numpy(dtype=float))
        design = pd.DataFrame(
            {"Feature (per SD)": feature_values},
            index=data.index,
        )
        diagnosis_columns = []
        for diagnosis in diagnosis_levels[1:]:
            column = f"Diagnosis: {diagnosis}"
            design[column] = (data["Diagnosis"] == diagnosis).astype(float)
            diagnosis_columns.append(column)
        for covariate, column in zip(
                selected_covariates,
                covariate_columns,
        ):
            values = data[column].to_numpy(dtype=float)
            if covariate == "gender":
                design["Gender"] = values - np.nanmin(values)
            else:
                design[covariate.title()] = _standardize(values)

        main_model = _fit_gee(
            data[outcome_field].to_numpy(dtype=float),
            design,
            data["Person ID"],
        )
        adjusted_rows.append(
            {
                **_candidate_identity(candidate),
                "n": main_model["n"],
                "Person clusters": main_model["cluster_count"],
                "Diagnosis levels": ", ".join(diagnosis_levels),
                "Controlled covariates": (
                        ", ".join(selected_covariates) or "none"
                ),
                "Feature beta per SD": main_model["coefficients"].get(
                    "Feature (per SD)",
                    np.nan,
                ),
                "Robust SE": main_model["standard_errors"].get(
                    "Feature (per SD)",
                    np.nan,
                ),
                "95% CI lower": main_model["ci_lower"].get(
                    "Feature (per SD)",
                    np.nan,
                ),
                "95% CI upper": main_model["ci_upper"].get(
                    "Feature (per SD)",
                    np.nan,
                ),
                "p": main_model["p_values"].get(
                    "Feature (per SD)",
                    np.nan,
                ),
                "FDR p": np.nan,
                "Status": main_model["status"],
            }
        )

        interaction_design = design.copy()
        interaction_columns = []
        for diagnosis_column in diagnosis_columns:
            diagnosis = diagnosis_column.replace("Diagnosis: ", "", 1)
            interaction_column = f"Feature x {diagnosis}"
            interaction_design[interaction_column] = (
                    interaction_design["Feature (per SD)"]
                    * interaction_design[diagnosis_column]
            )
            interaction_columns.append(interaction_column)

        interaction_model = _fit_gee(
            data[outcome_field].to_numpy(dtype=float),
            interaction_design,
            data["Person ID"],
        )
        wald_statistic, interaction_p = _wald_test(
            interaction_model,
            interaction_columns,
        )
        interaction_rows.append(
            {
                **_candidate_identity(candidate),
                "n": interaction_model["n"],
                "Person clusters": interaction_model["cluster_count"],
                "Reference diagnosis": reference_diagnosis,
                "Compared diagnoses": ", ".join(diagnosis_levels[1:]),
                "Controlled covariates": (
                        ", ".join(selected_covariates) or "none"
                ),
                "Interaction terms": len(interaction_columns),
                "Wald chi-square": wald_statistic,
                "p": interaction_p,
                "FDR p": np.nan,
                "Status": interaction_model["status"],
            }
        )
        slope_rows.extend(
            _diagnosis_specific_slopes(
                candidate,
                interaction_model,
                diagnosis_levels,
            )
        )

    adjusted_df = pd.DataFrame(
        adjusted_rows,
        columns=[
            "Clinical outcome",
            "Feature family",
            "Redundancy cluster",
            "Representative feature",
            "n",
            "Person clusters",
            "Diagnosis levels",
            "Controlled covariates",
            "Feature beta per SD",
            "Robust SE",
            "95% CI lower",
            "95% CI upper",
            "p",
            "FDR p",
            "Status",
        ],
    )
    interactions_df = pd.DataFrame(
        interaction_rows,
        columns=[
            "Clinical outcome",
            "Feature family",
            "Redundancy cluster",
            "Representative feature",
            "n",
            "Person clusters",
            "Reference diagnosis",
            "Compared diagnoses",
            "Controlled covariates",
            "Interaction terms",
            "Wald chi-square",
            "p",
            "FDR p",
            "Status",
        ],
    )
    slopes_df = pd.DataFrame(
        slope_rows,
        columns=[
            "Clinical outcome",
            "Feature family",
            "Redundancy cluster",
            "Representative feature",
            "Diagnosis",
            "n",
            "Person clusters",
            "Feature beta per SD",
            "Robust SE",
            "95% CI lower",
            "95% CI upper",
            "p",
            "FDR p",
            "Status",
        ],
    )
    for frame in (adjusted_df, interactions_df):
        if not frame.empty:
            frame["FDR p"] = frame.groupby("Clinical outcome")["p"].transform(
                _benjamini_hochberg
            )
    if not slopes_df.empty:
        slopes_df["FDR p"] = slopes_df.groupby(
            ["Clinical outcome", "Diagnosis"]
        )["p"].transform(_benjamini_hochberg)
    return adjusted_df, interactions_df, slopes_df


def _fit_gee(y, design, clusters):
    design = design.copy()
    design.insert(0, "Intercept", 1.0)
    names = design.columns.tolist()
    x = design.to_numpy(dtype=float)
    y = np.asarray(y, dtype=float)
    clusters = pd.Series(clusters).astype(str).to_numpy()
    n, parameter_count = x.shape
    cluster_values = np.unique(clusters)
    cluster_count = len(cluster_values)
    if n <= parameter_count or cluster_count < 2:
        return _empty_model(
            names,
            n,
            cluster_count,
            "insufficient degrees of freedom",
        )
    if np.linalg.matrix_rank(x) < parameter_count:
        return _empty_model(
            names,
            n,
            cluster_count,
            "rank-deficient design",
        )

    try:
        result = GEE(
            endog=y,
            exog=design,
            groups=clusters,
            family=Gaussian(),
            cov_struct=Independence(),
        ).fit()
    except Exception as exc:
        logger.warning("GEE fit failed", exc_info=True)
        return _empty_model(
            names,
            n,
            cluster_count,
            f"GEE failed: {exc}",
        )

    coefficients = np.asarray(result.params, dtype=float)
    standard_errors = np.asarray(result.bse, dtype=float)
    p_values = np.asarray(result.pvalues, dtype=float)
    confidence_intervals = np.asarray(result.conf_int(alpha=ALPHA), dtype=float)
    covariance = np.asarray(result.cov_params(), dtype=float)
    return {
        "names": names,
        "coefficients": dict(zip(names, coefficients)),
        "standard_errors": dict(zip(names, standard_errors)),
        "ci_lower": dict(zip(names, confidence_intervals[:, 0])),
        "ci_upper": dict(zip(names, confidence_intervals[:, 1])),
        "p_values": dict(zip(names, p_values)),
        "covariance": covariance,
        "n": int(n),
        "cluster_count": int(cluster_count),
        "status": "ok" if result.converged else "GEE did not converge",
    }


def _wald_test(model, terms):
    if not terms or model["covariance"] is None:
        return np.nan, np.nan
    indices = [model["names"].index(term) for term in terms]
    coefficients = np.array(
        [model["coefficients"][term] for term in terms],
        dtype=float,
    )
    covariance = model["covariance"][np.ix_(indices, indices)]
    statistic = float(
        coefficients.T @ np.linalg.pinv(covariance) @ coefficients
    )
    return statistic, float(chi2.sf(statistic, len(indices)))


def _diagnosis_specific_slopes(candidate, model, diagnosis_levels):
    rows = []
    if model["covariance"] is None:
        return rows
    feature_index = model["names"].index("Feature (per SD)")
    for diagnosis in diagnosis_levels:
        contrast = np.zeros(len(model["names"]), dtype=float)
        contrast[feature_index] = 1.0
        if diagnosis != diagnosis_levels[0]:
            interaction = f"Feature x {diagnosis}"
            if interaction in model["names"]:
                contrast[model["names"].index(interaction)] = 1.0
        coefficients = np.array(
            [model["coefficients"][name] for name in model["names"]]
        )
        estimate = float(contrast @ coefficients)
        variance = float(contrast @ model["covariance"] @ contrast)
        standard_error = np.sqrt(max(variance, 0.0))
        statistic = (
            estimate / standard_error if standard_error > 0 else np.nan
        )
        p_value = (
            float(2.0 * norm.sf(abs(statistic)))
            if np.isfinite(statistic)
            else np.nan
        )
        critical = norm.ppf(0.975)
        rows.append(
            {
                **_candidate_identity(candidate),
                "Diagnosis": diagnosis,
                "n": model["n"],
                "Person clusters": model["cluster_count"],
                "Feature beta per SD": estimate,
                "Robust SE": standard_error,
                "95% CI lower": estimate - critical * standard_error,
                "95% CI upper": estimate + critical * standard_error,
                "p": p_value,
                "FDR p": np.nan,
                "Status": model["status"],
            }
        )
    return rows


def _interpret_results(
        candidate_pairs,
        stratified_df,
        adjusted_df,
        interactions_df,
):
    columns = [
        "Clinical outcome",
        "Feature family",
        "Redundancy cluster",
        "Representative feature",
        "Pooled n",
        "Pooled Spearman rho",
        "Pooled FDR p",
        "Within-diagnosis support",
        "Supported diagnoses",
        "Adjusted beta per SD",
        "Adjusted FDR p",
        "Interaction FDR p",
        "Interpretation",
        "Interpretation note",
    ]
    rows = []
    for candidate in candidate_pairs.to_dict("records"):
        mask = (
                (stratified_df["Clinical outcome"] == candidate["Clinical outcome"])
                & (
                        stratified_df["Representative feature"]
                        == candidate["Representative feature"]
                )
        )
        stratified = stratified_df[mask]
        supported = stratified[
            pd.to_numeric(stratified["FDR p"], errors="coerce") < ALPHA
            ]
        adjusted = _matching_row(adjusted_df, candidate)
        interaction = _matching_row(interactions_df, candidate)
        adjusted_q = adjusted.get("FDR p", np.nan)
        interaction_q = interaction.get("FDR p", np.nan)
        adjusted_status = adjusted.get("Status", "missing")
        within_supported = not supported.empty
        adjusted_supported = pd.notna(adjusted_q) and adjusted_q < ALPHA
        interaction_supported = (
                pd.notna(interaction_q) and interaction_q < ALPHA
        )

        if adjusted_status != "ok":
            interpretation = "Not estimable in raw-feature model"
            note = (
                "The adjusted pooled feature was variable after nightly "
                "residualization and aggregation, but the corresponding raw "
                "grouped feature is constant, missing, or otherwise not "
                "estimable in the direct GEE model."
            )
        elif interaction_supported:
            interpretation = "Diagnosis-specific interaction"
            note = (
                "The feature-outcome slope differs by diagnosis after "
                "covariate adjustment."
            )
        elif adjusted_supported and within_supported:
            interpretation = "Within-group supported and diagnosis-adjusted"
            note = (
                "The association remains after diagnosis adjustment and is "
                "also detected within at least one diagnosis."
            )
        elif adjusted_supported:
            interpretation = "Diagnosis-adjusted"
            note = (
                "The feature remains associated with the outcome after "
                "adjusting for diagnosis and selected covariates."
            )
        elif within_supported:
            interpretation = "Within-group supported"
            note = (
                "At least one diagnosis shows an association, but the common "
                "adjusted feature effect is not retained."
            )
        else:
            interpretation = "Not retained after diagnosis adjustment"
            note = (
                "The pooled discovery is not significant within diagnoses or "
                "after diagnosis adjustment; group separation may contribute."
            )

        rows.append(
            {
                **candidate,
                "Within-diagnosis support": within_supported,
                "Supported diagnoses": ", ".join(
                    supported["Diagnosis"].dropna().astype(str).tolist()
                ) or "none",
                "Adjusted beta per SD": adjusted.get(
                    "Feature beta per SD",
                    np.nan,
                ),
                "Adjusted FDR p": adjusted_q,
                "Interaction FDR p": interaction_q,
                "Interpretation": interpretation,
                "Interpretation note": note,
            }
        )
    return pd.DataFrame(rows, columns=columns).sort_values(
        [
            "Interpretation",
            "Clinical outcome",
            "Pooled FDR p",
        ]
    ).reset_index(drop=True)


def _create_focused_plots(
        analysis_df,
        candidate_pairs,
        adjusted_df,
        interactions_df,
        selected_covariates,
        dataset_name,
        scenario,
        output_dir,
):
    columns = [
        "Plot",
        "Clinical outcome",
        "Representative feature",
        "Scenario",
        "Model shown",
        "n",
        "Person clusters",
        "Adjusted beta per SD",
        "Adjusted FDR p",
        "Interaction FDR p",
        "PNG path",
        "PDF path",
        "Interpretation note",
    ]
    specs = _focused_plot_specs(dataset_name, scenario["key"])
    if not specs:
        return pd.DataFrame(columns=columns)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    for position, spec in specs:
        candidate = _focused_candidate(candidate_pairs, spec)
        if not candidate:
            logger.warning(
                "Focused plot candidate not found for %s / %s: %s ~ %s",
                dataset_name,
                scenario["label"],
                spec["outcome"],
                spec["feature"],
            )
            continue
        adjusted = _matching_row(adjusted_df, candidate)
        interaction = _matching_row(interactions_df, candidate)
        plot_result = _plot_focused_association(
            analysis_df=analysis_df,
            candidate=candidate,
            adjusted=adjusted,
            interaction=interaction,
            selected_covariates=selected_covariates,
            spec=spec,
            output_dir=output_dir,
            position=position,
            scenario_label=scenario["label"],
        )
        if not plot_result:
            continue
        rows.append(
            {
                "Plot": spec["title"],
                "Clinical outcome": spec["outcome"],
                "Representative feature": spec["feature"],
                "Scenario": scenario["label"],
                "Model shown": plot_result["model_shown"],
                "n": plot_result["n"],
                "Person clusters": plot_result["cluster_count"],
                "Adjusted beta per SD": adjusted.get(
                    "Feature beta per SD",
                    np.nan,
                ),
                "Adjusted FDR p": adjusted.get("FDR p", np.nan),
                "Interaction FDR p": interaction.get("FDR p", np.nan),
                "PNG path": str(plot_result["png_path"]),
                "PDF path": str(plot_result["pdf_path"]),
                "Interpretation note": plot_result["note"],
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _focused_plot_specs(dataset_name, scenario_key):
    return [
        (position, spec)
        for position, spec in enumerate(FOCUSED_PLOT_SPECS, start=1)
        if (
                spec["dataset"] == dataset_name
                and spec["scenario"] == scenario_key
        )
    ]


def _focused_candidate(candidate_pairs, spec):
    if candidate_pairs.empty:
        return {}
    rows = candidate_pairs[
        (candidate_pairs["Clinical outcome"] == spec["outcome"])
        & (candidate_pairs["Representative feature"] == spec["feature"])
        ]
    return rows.iloc[0].to_dict() if not rows.empty else {}


def _plot_focused_association(
        analysis_df,
        candidate,
        adjusted,
        interaction,
        selected_covariates,
        spec,
        output_dir,
        position,
        scenario_label,
):
    feature = spec["feature"]
    outcome_field = CLINICAL_OUTCOMES[spec["outcome"]]
    covariate_columns = [
        COVARIATE_COLUMNS[covariate]
        for covariate in selected_covariates
        if COVARIATE_COLUMNS.get(covariate) in analysis_df.columns
    ]
    required = [
        feature,
        outcome_field,
        "Diagnosis",
        "Person ID",
        *covariate_columns,
    ]
    if feature not in analysis_df.columns:
        return {}
    data = _complete_cases(analysis_df, required)
    if (
            len(data) < MIN_GROUP_N
            or data[feature].nunique() < 2
            or data[outcome_field].nunique() < 2
    ):
        return {}

    diagnosis_levels = _diagnosis_levels(data)
    raw_feature = data[feature].to_numpy(dtype=float)
    feature_mean = float(np.mean(raw_feature))
    feature_sd = float(np.std(raw_feature, ddof=0))
    if not np.isfinite(feature_sd) or feature_sd == 0:
        return {}

    design = pd.DataFrame(
        {
            "Feature (per SD)": (
                    (raw_feature - feature_mean) / feature_sd
            ),
        },
        index=data.index,
    )
    diagnosis_columns = []
    for diagnosis in diagnosis_levels[1:]:
        column = f"Diagnosis: {diagnosis}"
        design[column] = (data["Diagnosis"] == diagnosis).astype(float)
        diagnosis_columns.append(column)
    for covariate, column in zip(selected_covariates, covariate_columns):
        values = data[column].to_numpy(dtype=float)
        if covariate == "gender":
            design["Gender"] = values - np.nanmin(values)
        else:
            design[covariate.title()] = _standardize(values)

    y = data[outcome_field].to_numpy(dtype=float)
    main_model = _fit_gee(y, design, data["Person ID"])
    interaction_design = design.copy()
    for diagnosis_column in diagnosis_columns:
        diagnosis = diagnosis_column.replace("Diagnosis: ", "", 1)
        interaction_design[f"Feature x {diagnosis}"] = (
                interaction_design["Feature (per SD)"]
                * interaction_design[diagnosis_column]
        )
    interaction_model = _fit_gee(y, interaction_design, data["Person ID"])
    interaction_q = pd.to_numeric(
        interaction.get("FDR p", np.nan),
        errors="coerce",
    )
    interaction_supported = (
            pd.notna(interaction_q) and interaction_q < ALPHA
    )
    if (
            interaction_supported
            and
            interaction_model["status"] == "ok"
            and interaction_model["covariance"] is not None
    ):
        plot_model = interaction_model
        model_shown = "Diagnosis-specific adjusted GEE slopes"
        note = (
            "Lines and 95% confidence bands come from the adjusted GEE "
            "interaction model."
        )
    elif main_model["status"] == "ok" and main_model["covariance"] is not None:
        plot_model = main_model
        model_shown = "Common adjusted GEE slope"
        if pd.notna(interaction_q):
            note = (
                "The interaction was not FDR-significant; parallel lines and "
                "95% confidence bands come from the diagnosis-adjusted main "
                "model."
            )
        else:
            note = (
                "The interaction model was not estimable; parallel lines and "
                "95% confidence bands come from the diagnosis-adjusted main "
                "model."
            )
    else:
        return {}

    fig, ax = plt.subplots(figsize=(10.8, 7.2))
    jittered = data[feature].nunique() <= 10
    feature_range = float(data[feature].max() - data[feature].min())
    jitter_step = (feature_range if feature_range > 0 else 1.0) * 0.012
    for diagnosis_index, diagnosis in enumerate(diagnosis_levels):
        diagnosis_data = data[data["Diagnosis"] == diagnosis].copy()
        color = DIAGNOSIS_COLORS.get(diagnosis, "#555555")
        diagnosis_offset = (
            (diagnosis_index - (len(diagnosis_levels) - 1) / 2)
            * jitter_step
            if jittered
            else 0.0
        )
        diagnosis_data["_display_feature"] = (
                diagnosis_data[feature] + diagnosis_offset
        )
        for _, person_data in diagnosis_data.groupby("Person ID"):
            if len(person_data) > 1:
                person_data = person_data.sort_values(feature)
                ax.plot(
                    person_data["_display_feature"],
                    person_data[outcome_field],
                    color=color,
                    alpha=0.16,
                    linewidth=0.8,
                    zorder=1,
                )
        ax.scatter(
            diagnosis_data["_display_feature"],
            diagnosis_data[outcome_field],
            s=38,
            color=color,
            edgecolor="white",
            linewidth=0.55,
            alpha=0.72,
            label=f"{diagnosis} (n={len(diagnosis_data)})",
            zorder=2,
        )
        grid = _plot_grid(diagnosis_data[feature])
        prediction, lower, upper = _prediction_band(
            plot_model,
            grid,
            diagnosis,
            diagnosis_levels,
            feature_mean,
            feature_sd,
        )
        ax.plot(
            grid,
            prediction,
            color=color,
            linewidth=2.5,
            zorder=3,
        )
        ax.fill_between(
            grid,
            lower,
            upper,
            color=color,
            alpha=0.13,
            linewidth=0,
            zorder=1,
        )

    adjusted_note = _plot_significance_note(
        candidate,
        adjusted,
        interaction,
        selected_covariates,
        data["Person ID"].nunique(),
    )
    fig.suptitle(
        spec["title"],
        x=0.10,
        y=0.97,
        ha="left",
        fontsize=17,
        fontweight="bold",
    )
    ax.set_title(
        f"{scenario_label} | {model_shown}",
        loc="left",
        pad=14,
        fontsize=10.5,
        color="#4D5963",
    )
    ax.set_xlabel(spec["x_label"], fontsize=11.5)
    ax.set_ylabel(spec["outcome"], fontsize=11.5)
    ax.grid(axis="y", color="#D7DEE2", linewidth=0.8, alpha=0.75)
    ax.grid(axis="x", visible=False)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#AAB5BB")
    ax.spines["bottom"].set_color("#AAB5BB")
    ax.legend(frameon=False, loc="best")
    covariate_position_note = (
        "Continuous covariates are held at their means."
    )
    if "gender" in selected_covariates:
        covariate_position_note += " Gender is held at the reference level."
    if jittered:
        covariate_position_note += (
            " Points are separated slightly on the x-axis to reveal "
            "overlapping diagnoses; model lines use the original values."
        )
    fig.text(
        0.07,
        0.015,
        textwrap.fill(
            (
                f"{adjusted_note} {covariate_position_note} Thin lines connect "
                "repeated visits from the same person."
            ),
            width=145,
        ),
        fontsize=8.5,
        color="#4D5963",
    )
    fig.subplots_adjust(left=0.10, right=0.97, top=0.86, bottom=0.16)

    stem = (
        f"{position:02d}_"
        f"{_slugify(spec['outcome'])}_"
        f"{_slugify(spec['feature'])}"
    )
    png_path = output_dir / f"{stem}.png"
    pdf_path = output_dir / f"{stem}.pdf"
    fig.savefig(png_path, dpi=300, facecolor="white")
    fig.savefig(pdf_path, facecolor="white")
    plt.close(fig)
    return {
        "png_path": png_path,
        "pdf_path": pdf_path,
        "n": int(len(data)),
        "cluster_count": int(data["Person ID"].nunique()),
        "model_shown": model_shown,
        "note": note,
    }


def _plot_grid(values):
    numeric = pd.to_numeric(values, errors="coerce").dropna().to_numpy()
    lower, upper = np.quantile(numeric, [0.025, 0.975])
    if lower == upper:
        lower, upper = float(np.min(numeric)), float(np.max(numeric))
    return np.linspace(lower, upper, 120)


def _prediction_band(
        model,
        raw_grid,
        diagnosis,
        diagnosis_levels,
        feature_mean,
        feature_sd,
):
    rows = []
    for raw_value in raw_grid:
        values = {name: 0.0 for name in model["names"]}
        values["Intercept"] = 1.0
        standardized = (raw_value - feature_mean) / feature_sd
        values["Feature (per SD)"] = standardized
        if diagnosis != diagnosis_levels[0]:
            diagnosis_column = f"Diagnosis: {diagnosis}"
            if diagnosis_column in values:
                values[diagnosis_column] = 1.0
            interaction_column = f"Feature x {diagnosis}"
            if interaction_column in values:
                values[interaction_column] = standardized
        rows.append([values[name] for name in model["names"]])
    matrix = np.asarray(rows, dtype=float)
    coefficients = np.asarray(
        [model["coefficients"][name] for name in model["names"]],
        dtype=float,
    )
    prediction = matrix @ coefficients
    variance = np.einsum(
        "ij,jk,ik->i",
        matrix,
        model["covariance"],
        matrix,
    )
    standard_error = np.sqrt(np.maximum(variance, 0.0))
    critical = norm.ppf(0.975)
    return (
        prediction,
        prediction - critical * standard_error,
        prediction + critical * standard_error,
    )


def _plot_significance_note(
        candidate,
        adjusted,
        interaction,
        selected_covariates,
        person_count,
):
    covariates = ", ".join(selected_covariates) or "none"
    adjusted_q = adjusted.get("FDR p", np.nan)
    interaction_q = interaction.get("FDR p", np.nan)
    adjusted_text = (
        f"{adjusted_q:.3g}" if pd.notna(adjusted_q) else "not estimable"
    )
    interaction_text = (
        f"{interaction_q:.3g}" if pd.notna(interaction_q) else "not estimable"
    )
    return (
        f"Pooled discovery: rho={candidate['Pooled Spearman rho']:.3f}, "
        f"FDR p={candidate['Pooled FDR p']:.3g}. "
        f"Adjusted GEE FDR p={adjusted_text}; interaction FDR "
        f"p={interaction_text}. Covariates: {covariates}; "
        f"{person_count} people."
    )


def _settings(
        raw_grouped_path,
        family_workbook_path,
        dataset_name,
        scenario,
        selected_covariates,
        analysis_df,
        candidate_pairs,
        summary_df,
):
    counts = (
        summary_df["Interpretation"].value_counts().to_dict()
        if not summary_df.empty
        else {}
    )
    rows = [
        {"setting": "dataset", "value": dataset_name},
        {"setting": "scenario", "value": scenario["label"]},
        {"setting": "raw_grouped_source", "value": str(raw_grouped_path)},
        {
            "setting": "family_candidate_source",
            "value": str(family_workbook_path),
        },
        {
            "setting": "candidate_definition",
            "value": (
                "FDR-significant pooled feature-outcome associations reduced "
                "to representatives of |Spearman rho| >= 0.85 clusters"
            ),
        },
        {
            "setting": "controlled_covariates",
            "value": ", ".join(selected_covariates) or "none",
        },
        {
            "setting": "within_diagnosis_method",
            "value": (
                "Spearman rho estimated as standardized rank-regression slope "
                "using Gaussian GEE with independence working correlation; "
                "robust SE, CI, and p clustered by underlying person"
            ),
        },
        {
            "setting": "adjusted_model",
            "value": (
                "Gaussian GEE with independence working correlation and robust "
                "person-clustered sandwich covariance; raw outcome, feature "
                "standardized within complete cases, diagnosis and "
                "scenario-specific covariates entered directly"
            ),
        },
        {
            "setting": "interaction_model",
            "value": (
                "Adjusted model plus feature x diagnosis terms; omnibus "
                "cluster-robust Wald chi-square"
            ),
        },
        {
            "setting": "diagnosis_specific_slopes",
            "value": (
                "Linear contrasts from the interaction model, expressed as "
                "outcome units per one-SD increase in the feature"
            ),
        },
        {
            "setting": "person_identity",
            "value": (
                "HC2-/HC3- collapsed to HC-; pre-LBD2-/pre-LBD3- collapsed "
                "to pre-LBD-; other subject codes unchanged"
            ),
        },
        {
            "setting": "multiple_testing",
            "value": (
                "Benjamini-Hochberg FDR separately by clinical outcome for "
                "stratified correlations, adjusted associations, and "
                "interactions; diagnosis-specific slopes additionally "
                "separated by diagnosis"
            ),
        },
        {"setting": "alpha", "value": ALPHA},
        {"setting": "candidate_pair_count", "value": len(candidate_pairs)},
        {"setting": "subject_visit_count", "value": len(analysis_df)},
        {
            "setting": "underlying_person_count",
            "value": analysis_df["Person ID"].nunique(),
        },
        *[
            {
                "setting": f"interpretation_{_slugify(label)}",
                "value": count,
            }
            for label, count in sorted(counts.items())
        ],
        {
            "setting": "exploratory_warning",
            "value": (
                "Candidates were selected in the same data; these follow-up "
                "tests are explanatory sensitivity analyses, not independent "
                "confirmation."
            ),
        },
    ]
    return pd.DataFrame(rows)


def _diagnosis_levels(analysis_df):
    present = set(analysis_df["Diagnosis"].dropna())
    return [
        label
        for label in ("HC", "MCI-AD", "preDLB")
        if label in present
    ]


def _complete_cases(df, columns):
    available = list(dict.fromkeys(["Person ID", *columns]))
    data = df[available].copy()
    for column in columns:
        if column not in ("Diagnosis", "Person ID"):
            data[column] = pd.to_numeric(data[column], errors="coerce")
    return data.replace([np.inf, -np.inf], np.nan).dropna()


def _person_id(subject_code):
    code = str(subject_code).strip()
    code = re.sub(r"^HC[23]-", "HC-", code, flags=re.IGNORECASE)
    code = re.sub(
        r"^pre[-_]?LBD[23]-",
        "pre-LBD-",
        code,
        flags=re.IGNORECASE,
    )
    return code


def _standardize(values):
    values = np.asarray(values, dtype=float)
    standard_deviation = np.std(values, ddof=0)
    if not np.isfinite(standard_deviation) or standard_deviation == 0:
        return np.zeros_like(values)
    return (values - np.mean(values)) / standard_deviation


def _benjamini_hochberg(p_values):
    numeric = pd.to_numeric(p_values, errors="coerce")
    adjusted = pd.Series(np.nan, index=numeric.index, dtype=float)
    valid = numeric.dropna()
    if valid.empty:
        return adjusted
    ordered = valid.sort_values()
    count = len(ordered)
    raw = ordered.to_numpy(dtype=float) * count / np.arange(1, count + 1)
    monotonic = np.minimum.accumulate(raw[::-1])[::-1]
    adjusted.loc[ordered.index] = np.clip(monotonic, 0.0, 1.0)
    return adjusted


def _candidate_identity(candidate):
    return {
        "Clinical outcome": candidate["Clinical outcome"],
        "Feature family": candidate["Feature family"],
        "Redundancy cluster": candidate["Redundancy cluster"],
        "Representative feature": candidate["Representative feature"],
    }


def _matching_row(frame, candidate):
    if frame.empty:
        return {}
    rows = frame[
        (frame["Clinical outcome"] == candidate["Clinical outcome"])
        & (
                frame["Representative feature"]
                == candidate["Representative feature"]
        )
        ]
    return rows.iloc[0].to_dict() if not rows.empty else {}


def _empty_model(names, n, cluster_count, status):
    empty = {name: np.nan for name in names}
    return {
        "names": names,
        "coefficients": empty.copy(),
        "standard_errors": empty.copy(),
        "ci_lower": empty.copy(),
        "ci_upper": empty.copy(),
        "p_values": empty.copy(),
        "covariance": None,
        "n": int(n),
        "cluster_count": int(cluster_count),
        "status": status,
    }


def _empty_stratified_row(candidate, diagnosis, status):
    return {
        **_candidate_identity(candidate),
        "Diagnosis": diagnosis,
        "n": 0,
        "Person clusters": 0,
        "Spearman rho": np.nan,
        "Robust SE": np.nan,
        "95% CI lower": np.nan,
        "95% CI upper": np.nan,
        "p": np.nan,
        "FDR p": np.nan,
        "Status": status,
    }


def _failed_model_row(candidate, selected_covariates, status, data=None):
    return {
        **_candidate_identity(candidate),
        "n": int(len(data)) if data is not None else 0,
        "Person clusters": (
            int(data["Person ID"].nunique())
            if data is not None and "Person ID" in data
            else 0
        ),
        "Diagnosis levels": "",
        "Controlled covariates": ", ".join(selected_covariates) or "none",
        "Feature beta per SD": np.nan,
        "Robust SE": np.nan,
        "95% CI lower": np.nan,
        "95% CI upper": np.nan,
        "p": np.nan,
        "FDR p": np.nan,
        "Status": status,
    }


def _failed_interaction_row(candidate, selected_covariates, status, data=None):
    return {
        **_candidate_identity(candidate),
        "n": int(len(data)) if data is not None else 0,
        "Person clusters": (
            int(data["Person ID"].nunique())
            if data is not None and "Person ID" in data
            else 0
        ),
        "Reference diagnosis": "HC",
        "Compared diagnoses": "",
        "Controlled covariates": ", ".join(selected_covariates) or "none",
        "Interaction terms": 0,
        "Wald chi-square": np.nan,
        "p": np.nan,
        "FDR p": np.nan,
        "Status": status,
    }


def _slugify(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_")


def _style_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    significant_fill = PatternFill("solid", fgColor="E2F0D9")
    caution_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        worksheet.row_dimensions[1].height = 36
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        for column_index, column_cells in enumerate(
                worksheet.columns,
                start=1,
        ):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(max(max_length + 2, 12), 48)

    summary = workbook["interpretation_summary"]
    summary.column_dimensions["B"].width = 36
    summary.column_dimensions["D"].width = 48
    summary.column_dimensions["M"].width = 42
    summary.column_dimensions["N"].width = 64
    for row in range(2, summary.max_row + 1):
        summary.row_dimensions[row].height = 48
        summary.cell(row, 14).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        interpretation = summary.cell(row, 13).value
        if interpretation == "Not retained after diagnosis adjustment":
            summary.cell(row, 13).fill = caution_fill
        else:
            summary.cell(row, 13).fill = significant_fill
            summary.cell(row, 13).font = Font(bold=True)

    for sheet_name in (
            "stratified_correlations",
            "adjusted_associations",
            "diagnosis_interactions",
            "diagnosis_specific_slopes",
    ):
        worksheet = workbook[sheet_name]
        worksheet.column_dimensions["B"].width = 36
        worksheet.column_dimensions["D"].width = 48
        header_by_name = {
            cell.value: cell.column for cell in worksheet[1]
        }
        for p_header in ("p", "FDR p"):
            column = header_by_name.get(p_header)
            if not column:
                continue
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, column)
                cell.number_format = "0.000E+00"
                if (
                        p_header == "FDR p"
                        and isinstance(cell.value, (int, float))
                        and cell.value < ALPHA
                ):
                    cell.fill = significant_fill
                    cell.font = Font(bold=True)

    candidates = workbook["candidate_pairs"]
    candidates.column_dimensions["B"].width = 36
    candidates.column_dimensions["D"].width = 48
    plot_index = workbook["focused_plot_index"]
    plot_index.column_dimensions["A"].width = 48
    plot_index.column_dimensions["C"].width = 48
    plot_index.column_dimensions["E"].width = 38
    plot_index.column_dimensions["K"].width = 80
    plot_index.column_dimensions["L"].width = 80
    plot_index.column_dimensions["M"].width = 64
    for row in range(2, plot_index.max_row + 1):
        plot_index.row_dimensions[row].height = 42
        plot_index.cell(row, 13).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
    settings = workbook["settings"]
    settings.column_dimensions["A"].width = 44
    settings.column_dimensions["B"].width = 100
    for row in range(2, settings.max_row + 1):
        settings.cell(row, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        if len(str(settings.cell(row, 2).value or "")) > 80:
            settings.row_dimensions[row].height = 42
