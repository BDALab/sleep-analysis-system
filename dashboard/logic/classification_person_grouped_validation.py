import json
import logging
import os
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.base import clone
from sklearn.compose import ColumnTransformer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    auc,
    average_precision_score,
    balanced_accuracy_score,
    confusion_matrix,
    matthews_corrcoef,
    precision_score,
    recall_score,
    roc_curve,
)
from sklearn.model_selection import RandomizedSearchCV, StratifiedGroupKFold, cross_val_predict
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder

from dashboard.logic.analysis_preparation import prepare_analysis_dataset
from dashboard.logic.classification_covariates import resolve_adjustment_columns
from dashboard.logic.classification_grouped_statistics import (
    ADJUSTMENT_COVARIATE_COLUMNS,
    DIARY_COVARIATE_COLUMNS,
    FEATURE_BLOCK_ALL,
    FEATURE_SELECTOR_MODE_KBEST,
    FEATURE_SELECTOR_MODE_RFE,
    SEED,
    STATS_PREFIXES,
    TARGET_COLUMN,
    TARGET_LABEL_COLUMN,
    _base_summary_row,
    _binarize_proba,
    _build_pipeline_with_selector,
    _classification_report_dataframe,
    _codes_to_label,
    _confusion_matrix_dataframe,
    _is_gpu_error,
    _json_ready_dict,
    _prepare_dataset,
    _prepare_scenario_features,
    _save_curve_points,
    _save_json,
    _save_metrics,
    _save_roc_pr_figure,
    _scenario_covariate_mapping,
    _scenario_label,
    _tune_threshold,
)
from dashboard.logic.classification_grouped_statistics_strict import (
    _feature_selection_metadata,
    _filter_columns_by_feature_family,
    _search_settings_for_selector_mode,
)
from dashboard.logic.classification_validity_checks import (
    _cohort_chi_square,
    subject_person_id,
    subject_source_cohort,
    subject_visit_index,
)
from dashboard.logic.feature_families import (
    ACTIVITY_EXTENSION_STABLE_FAMILY_IDS,
    PRIMARY_SLEEP_STABLE_FAMILY_IDS,
)
from dashboard.logic.xgboost_runtime import xgboost_runtime_metadata
from mysite.settings import MEDIA_ROOT

logger = logging.getLogger(__name__)

HC_VS_PREDLB_SCENARIO = ((3,), (0,))
PERSON_GROUPED_RESULTS_ROOT = (
        Path(MEDIA_ROOT) / "classification" / "person-grouped-thesis" / "hc-vs-predlb"
)
PERSON_GROUPED_OUTER_CV_SPLITS = max(
    2,
    int(os.environ.get("GENEACTIV_PERSON_GROUPED_OUTER_CV_SPLITS", "5")),
)
PERSON_GROUPED_INNER_CV_SPLITS = max(
    2,
    int(os.environ.get("GENEACTIV_PERSON_GROUPED_INNER_CV_SPLITS", "5")),
)

PERSON_GROUPED_RUN_SPECS = (
    {
        "run_key": "broad_strict_rfe",
        "run_label": "Broad strict RFE, person-grouped",
        "dataset_name": "dataset-clinical",
        "output_name": "broad-strict-rfe",
        "include_diary_covariates": True,
        "feature_selector_mode": FEATURE_SELECTOR_MODE_RFE,
        "allowed_feature_family_ids": frozenset(),
        "notes": "All grouped sleep/diary features plus diary predictors, XGBoost RFE.",
    },
    {
        "run_key": "stable_sleep",
        "run_label": "Stable primary sleep families, person-grouped",
        "dataset_name": "dataset-clinical",
        "output_name": "stable-primary-sleep",
        "include_diary_covariates": False,
        "feature_selector_mode": FEATURE_SELECTOR_MODE_KBEST,
        "allowed_feature_family_ids": PRIMARY_SLEEP_STABLE_FAMILY_IDS,
        "notes": "HC vs preDLB, fixed primary sleep families only.",
    },
    {
        "run_key": "stable_sleep_activity",
        "run_label": "Stable sleep + activity variability, person-grouped",
        "dataset_name": "dataset-clinical-acc",
        "output_name": "stable-sleep-activity",
        "include_diary_covariates": False,
        "feature_selector_mode": FEATURE_SELECTOR_MODE_KBEST,
        "allowed_feature_family_ids": (
                PRIMARY_SLEEP_STABLE_FAMILY_IDS | ACTIVITY_EXTENSION_STABLE_FAMILY_IDS
        ),
        "notes": "HC vs preDLB, primary sleep families plus activity variability.",
    },
)


def run_hc_vs_predlb_person_grouped_classification(output_dir=None, run_specs=None):
    output_dir = Path(output_dir) if output_dir else _default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)
    run_specs = tuple(run_specs or PERSON_GROUPED_RUN_SPECS)

    summary_rows = []
    first_visit_rows = []
    cohort_distribution_frames = []
    cohort_performance_frames = []
    source_only_rows = []
    leave_one_cohort_frames = []
    within_cohort_frames = []

    for spec in run_specs:
        result = _run_person_grouped_spec(spec, output_dir)
        summary_rows.append(result["summary"])
        first_visit_rows.append(result["first_visit_summary"])
        cohort_distribution_frames.append(result["cohort_distribution"])
        cohort_performance_frames.append(result["cohort_performance"])
        source_only_rows.append(result["source_only_summary"])
        leave_one_cohort_frames.append(result["leave_one_cohort_summary"])
        within_cohort_frames.append(result["within_cohort_summary"])

    summary_df = pd.DataFrame(summary_rows)
    first_visit_df = pd.DataFrame(first_visit_rows)
    cohort_distribution_df = pd.concat(cohort_distribution_frames, ignore_index=True)
    cohort_performance_df = pd.concat(cohort_performance_frames, ignore_index=True)
    source_only_df = pd.DataFrame(source_only_rows)
    leave_one_cohort_df = pd.concat(leave_one_cohort_frames, ignore_index=True)
    within_cohort_df = pd.concat(within_cohort_frames, ignore_index=True)
    settings_df = _settings(output_dir)

    summary_path = output_dir / "hc_vs_predlb_person_grouped_classification_summary.xlsx"
    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        first_visit_df.to_excel(writer, sheet_name="first_visit_summary", index=False)
        cohort_distribution_df.to_excel(writer, sheet_name="cohort_distribution", index=False)
        cohort_performance_df.to_excel(writer, sheet_name="cohort_performance", index=False)
        source_only_df.to_excel(writer, sheet_name="source_only_negative_control", index=False)
        leave_one_cohort_df.to_excel(writer, sheet_name="leave_one_cohort_out", index=False)
        within_cohort_df.to_excel(writer, sheet_name="within_cohort_nested_cv", index=False)
        settings_df.to_excel(writer, sheet_name="settings", index=False)
        _style_workbook(writer.book)

    result = {
        "run_dir": str(output_dir),
        "summary_path": str(summary_path),
        "summary": summary_df.replace({np.nan: None}).to_dict("records"),
    }
    (output_dir / "hc_vs_predlb_person_grouped_classification_summary.json").write_text(
        json.dumps(result, indent=2),
        encoding="utf-8",
    )
    logger.info("HC-vs-preDLB person-grouped classification completed: %s", output_dir)
    return result


def _run_person_grouped_spec(spec, output_root):
    spec_dir = output_root / spec["output_name"]
    spec_dir.mkdir(parents=True, exist_ok=True)

    prepared = _prepare_spec_data(spec, spec_dir)
    predictions, fold_details = _run_grouped_nested_cv(
        X=prepared["X"],
        y=prepared["y"],
        groups=prepared["groups"],
        subjects=prepared["subjects"],
        source_cohorts=prepared["source_cohorts"],
        visit_indices=prepared["visit_indices"],
        feature_selector_mode=spec["feature_selector_mode"],
        n_covariates=len(prepared["adjustment_columns"]),
        outer_splits=PERSON_GROUPED_OUTER_CV_SPLITS,
        inner_splits=PERSON_GROUPED_INNER_CV_SPLITS,
    )
    predictions.to_excel(spec_dir / "subject_predictions.xlsx", index=False)
    fold_details.to_excel(spec_dir / "outer_fold_details.xlsx", index=False)

    y_true = predictions["y_true"].astype(int).to_numpy()
    y_pred_default = predictions["y_pred_default"].astype(int).to_numpy()
    y_pred_tuned = predictions["y_pred_tuned"].astype(int).to_numpy()
    y_prob = predictions["pred_probability_positive"].astype(float).to_numpy()

    default_metrics = _binary_metrics(y_true, y_pred_default, y_prob)
    tuned_metrics = _binary_metrics(y_true, y_pred_tuned, y_prob)
    first_visit_predictions = predictions[predictions["is_first_visit"]].copy()
    first_visit_metrics = _binary_metrics(
        first_visit_predictions["y_true"],
        first_visit_predictions["y_pred_default"],
        first_visit_predictions["pred_probability_positive"],
    )

    _save_prediction_outputs(
        spec_dir=spec_dir,
        predictions=predictions,
        y_true=y_true,
        y_prob=y_prob,
        y_pred_default=y_pred_default,
        y_pred_tuned=y_pred_tuned,
        target_names=prepared["target_names"],
        default_metrics=default_metrics,
        tuned_metrics=tuned_metrics,
    )

    cohort_distribution = _cohort_distribution(predictions, spec)
    cohort_performance = _cohort_performance(predictions, spec)
    cohort_chi2 = _cohort_chi_square(
        pd.crosstab(predictions["source_cohort"], predictions["diagnosis_label"])
    )
    source_only_summary, source_only_predictions = _source_only_negative_control(
        predictions=predictions,
        outer_splits=PERSON_GROUPED_OUTER_CV_SPLITS,
    )
    source_only_predictions.to_excel(
        spec_dir / "source_only_negative_control_predictions.xlsx",
        index=False,
    )

    leave_one_cohort_summary, leave_one_cohort_predictions = _leave_one_cohort_out_validation(
        prepared=prepared,
        spec=spec,
    )
    leave_one_cohort_summary.to_excel(spec_dir / "leave_one_cohort_out_summary.xlsx", index=False)
    leave_one_cohort_predictions.to_excel(
        spec_dir / "leave_one_cohort_out_predictions.xlsx",
        index=False,
    )

    within_cohort_summary, within_cohort_predictions = _within_cohort_nested_cv(
        prepared=prepared,
        spec=spec,
    )
    within_cohort_summary.to_excel(spec_dir / "within_cohort_nested_cv_summary.xlsx", index=False)
    within_cohort_predictions.to_excel(
        spec_dir / "within_cohort_nested_cv_predictions.xlsx",
        index=False,
    )

    repeated_person_count = int(
        (predictions.groupby("person_group")["#Subject"].nunique() > 1).sum()
    )
    repeated_subject_count = int(predictions["person_group"].duplicated(keep=False).sum())
    base_summary = _base_summary_row(
        scenario_label=_scenario_label(*HC_VS_PREDLB_SCENARIO),
        feature_block_key=FEATURE_BLOCK_ALL,
        positive_codes=HC_VS_PREDLB_SCENARIO[0],
        negative_codes=HC_VS_PREDLB_SCENARIO[1],
        subject_count=len(predictions),
        positive_count=int((y_true == 1).sum()),
        negative_count=int((y_true == 0).sum()),
    )
    summary = {
        **base_summary,
        "run_key": spec["run_key"],
        "run_label": spec["run_label"],
        "dataset_name": spec["dataset_name"],
        "run_dir": str(spec_dir),
        "mode": "person_grouped_nested_cv",
        "status": "completed",
        "feature_selector_mode": spec["feature_selector_mode"],
        "include_diary_covariates": bool(spec["include_diary_covariates"]),
        "allowed_feature_family_ids": ", ".join(
            sorted(spec.get("allowed_feature_family_ids") or [])
        ),
        "selected_adjustment_covariates": ", ".join(prepared["selected_covariates"]) or "none",
        "subject_count": int(len(predictions)),
        "person_group_count": int(predictions["person_group"].nunique()),
        "repeated_person_count": repeated_person_count,
        "subjects_from_repeated_people": repeated_subject_count,
        "outer_cv": f"StratifiedGroupKFold({fold_details['fold'].nunique()})",
        "inner_cv_max_splits": PERSON_GROUPED_INNER_CV_SPLITS,
        "roc_auc": default_metrics["AUC"],
        "pr_auc": default_metrics["PR_AUC"],
        **_metric_columns(default_metrics, prefix="default"),
        **_metric_columns(tuned_metrics, prefix="tuned"),
        "first_visit_auc": first_visit_metrics["AUC"],
        "first_visit_bacc": first_visit_metrics["BACC"],
        "cohort_diagnosis_chi2": cohort_chi2["cohort_diagnosis_chi2"],
        "cohort_diagnosis_p": cohort_chi2["cohort_diagnosis_p"],
        "cohort_diagnosis_cramers_v": cohort_chi2["cohort_diagnosis_cramers_v"],
        "cohort_diagnosis_note": cohort_chi2["cohort_diagnosis_note"],
        "source_only_auc": source_only_summary["AUC"],
        "source_only_bacc": source_only_summary["BACC"],
    }
    first_visit_summary = {
        "run_key": spec["run_key"],
        "run_label": spec["run_label"],
        "subject_count": int(len(first_visit_predictions)),
        "person_group_count": int(first_visit_predictions["person_group"].nunique()),
        **first_visit_metrics,
    }

    summary_df = pd.DataFrame([summary])
    first_visit_df = pd.DataFrame([first_visit_summary])
    with pd.ExcelWriter(spec_dir / "classification_summary.xlsx", engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        first_visit_df.to_excel(writer, sheet_name="first_visit_summary", index=False)
        fold_details.to_excel(writer, sheet_name="outer_fold_details", index=False)
        cohort_distribution.to_excel(writer, sheet_name="cohort_distribution", index=False)
        cohort_performance.to_excel(writer, sheet_name="cohort_performance", index=False)
        pd.DataFrame([source_only_summary]).to_excel(
            writer,
            sheet_name="source_only_negative_control",
            index=False,
        )
        leave_one_cohort_summary.to_excel(writer, sheet_name="leave_one_cohort_out", index=False)
        within_cohort_summary.to_excel(writer, sheet_name="within_cohort_nested_cv", index=False)
        _style_workbook(writer.book)

    return {
        "summary": summary,
        "first_visit_summary": first_visit_summary,
        "cohort_distribution": cohort_distribution,
        "cohort_performance": cohort_performance,
        "source_only_summary": {
            "run_key": spec["run_key"],
            "run_label": spec["run_label"],
            **source_only_summary,
        },
        "leave_one_cohort_summary": leave_one_cohort_summary,
        "within_cohort_summary": within_cohort_summary,
    }


def _prepare_spec_data(spec, spec_dir):
    preparation = prepare_analysis_dataset(spec["dataset_name"])
    scenario_covariates = _scenario_covariate_mapping(preparation)
    selected_covariates = tuple(scenario_covariates.get(HC_VS_PREDLB_SCENARIO, ()))

    base_df = pd.read_excel(preparation["raw_grouped_stats_path"])
    prepared_df, excluded_labels_df, dataset_overview_df, covariate_info = _prepare_dataset(
        base_df,
        include_diary_covariates=spec["include_diary_covariates"],
    )
    prepared_df.to_excel(spec_dir / "prepared_dataset.xlsx", index=False)
    dataset_overview_df.to_excel(spec_dir / "dataset_overview.xlsx", index=False)
    if not excluded_labels_df.empty:
        excluded_labels_df.to_excel(spec_dir / "excluded_subjects_missing_labels.xlsx", index=False)

    positive_codes, negative_codes = HC_VS_PREDLB_SCENARIO
    scenario_codes = set(positive_codes) | set(negative_codes)
    scenario_df = prepared_df[prepared_df[TARGET_COLUMN].isin(scenario_codes)].copy()
    scenario_df["binary_target"] = scenario_df[TARGET_COLUMN].apply(
        lambda code: 1 if code in positive_codes else 0
    )
    scenario_df["binary_target_label"] = scenario_df["binary_target"].map(
        {0: _codes_to_label(negative_codes), 1: _codes_to_label(positive_codes)}
    )
    scenario_df["person_group"] = scenario_df["#Subject"].map(subject_person_id)
    scenario_df["source_cohort"] = scenario_df["#Subject"].map(subject_source_cohort)
    scenario_df["visit_index"] = scenario_df["#Subject"].map(subject_visit_index)
    scenario_df.to_excel(spec_dir / "df_original.xlsx", index=False)

    stats_columns = [column for column in scenario_df.columns if str(column).startswith(STATS_PREFIXES)]
    diary_predictor_columns = [
        column for column in DIARY_COVARIATE_COLUMNS if column in scenario_df.columns
    ]
    allowed_family_ids = frozenset(spec.get("allowed_feature_family_ids") or ())
    excluded_family_feature_columns = []
    if allowed_family_ids:
        candidate_feature_columns = list(dict.fromkeys(stats_columns + diary_predictor_columns))
        stats_columns, diary_predictor_columns, family_filter_df = _filter_columns_by_feature_family(
            stats_columns=stats_columns,
            additional_feature_columns=diary_predictor_columns,
            allowed_feature_family_ids=allowed_family_ids,
        )
        kept_columns = set(stats_columns + diary_predictor_columns)
        excluded_family_feature_columns = [
            column for column in candidate_feature_columns if column not in kept_columns
        ]
        scenario_df = scenario_df.drop(columns=excluded_family_feature_columns)
        family_filter_df.to_excel(spec_dir / "feature_family_filter.xlsx", index=False)

    filtered_df, feature_mapping_df, feature_coverage_df = _prepare_scenario_features(
        scenario_df,
        stats_columns=stats_columns,
        additional_feature_columns=diary_predictor_columns,
    )
    filtered_df["person_group"] = filtered_df["#Subject"].map(subject_person_id)
    filtered_df["source_cohort"] = filtered_df["#Subject"].map(subject_source_cohort)
    filtered_df["visit_index"] = filtered_df["#Subject"].map(subject_visit_index)
    filtered_df.to_excel(spec_dir / "df_preprocessed.xlsx", index=False)
    feature_mapping_df.to_excel(spec_dir / "feature_name_mapping.xlsx", index=False)
    feature_coverage_df.to_excel(spec_dir / "feature_coverage.xlsx", index=False)

    feature_columns = [
        column
        for column in filtered_df.columns
        if column not in {
            "#Subject",
            "#Age",
            "#Gender",
            "#Education",
            "#Disease",
            TARGET_COLUMN,
            TARGET_LABEL_COLUMN,
            "binary_target",
            "binary_target_label",
            "person_group",
            "source_cohort",
            "visit_index",
        }
    ]
    if not feature_columns:
        raise ValueError(f"No features left for {spec['run_key']} after filtering")

    adjustment_columns = resolve_adjustment_columns(
        selected_covariates=selected_covariates,
        available_columns=filtered_df.columns,
        covariate_columns=ADJUSTMENT_COVARIATE_COLUMNS,
    )
    X = filtered_df[feature_columns + adjustment_columns].apply(
        pd.to_numeric,
        errors="coerce",
    ).to_numpy()
    y = filtered_df["binary_target"].astype(int).to_numpy()
    groups = filtered_df["person_group"].astype(str).to_numpy()

    _validate_person_groups(groups, y)
    _save_json(
        {
            "dataset_name": spec["dataset_name"],
            "source_path": str(preparation["raw_grouped_stats_path"]),
            "run_key": spec["run_key"],
            "run_label": spec["run_label"],
            "mode": "person_grouped_nested_cv_thesis",
            "scenario": "HC vs preDLB",
            "seed": SEED,
            "outer_cv": f"StratifiedGroupKFold(max_splits={PERSON_GROUPED_OUTER_CV_SPLITS})",
            "inner_cv": f"StratifiedGroupKFold(max_splits={PERSON_GROUPED_INNER_CV_SPLITS})",
            "feature_selector_mode": spec["feature_selector_mode"],
            "feature_selection": _feature_selection_metadata(spec["feature_selector_mode"]),
            "include_diary_covariates": bool(spec["include_diary_covariates"]),
            "feature_family_filter": {
                "enabled": bool(allowed_family_ids),
                "allowed_family_ids": sorted(allowed_family_ids),
                "excluded_feature_count": len(excluded_family_feature_columns),
            },
            "feature_count": len(feature_columns),
            "adjustment_covariates": list(selected_covariates),
            "adjustment_columns": adjustment_columns,
            "person_group_rule": (
                "HC2/HC3 and preDLB/preLBD second/third visit codes are grouped "
                "with the same base subject before all outer and inner splits."
            ),
            "cohort_sensitivity": [
                "cohort-stratified performance from person-grouped predictions",
                "source-only negative-control classifier using the same person-grouped CV",
                "leave-one-source-cohort-out validation",
                "within-source-cohort nested CV where sample size permits",
            ],
            "preparation_manifest": _json_ready_dict(preparation),
            "diary_covariates": _json_ready_dict(covariate_info),
            "xgboost_runtime": xgboost_runtime_metadata(),
            "notes": spec["notes"],
        },
        spec_dir / "analysis_metadata.json",
    )
    _save_json(
        {
            "feature_labels": feature_columns,
            "foldwise_adjustment_covariates": list(selected_covariates),
            "foldwise_adjustment_columns": adjustment_columns,
            "adjustment_fit_scope": "inner/outer training fold only",
            "excluded_by_feature_family_filter": excluded_family_feature_columns,
        },
        spec_dir / "feature_labels.json",
    )
    scenario_subjects_df = filtered_df[
        [
            column
            for column in (
            "#Subject",
            "person_group",
            "source_cohort",
            "visit_index",
            "#Age",
            "#Gender",
            "#Education",
            "#Disease",
            TARGET_COLUMN,
            TARGET_LABEL_COLUMN,
            "binary_target",
            "binary_target_label",
        )
            if column in filtered_df.columns
        ]
    ].copy()
    scenario_subjects_df.to_excel(spec_dir / "scenario_subjects.xlsx", index=False)

    return {
        "X": X,
        "y": y,
        "groups": groups,
        "subjects": filtered_df["#Subject"].astype(str).to_numpy(),
        "source_cohorts": filtered_df["source_cohort"].astype(str).to_numpy(),
        "visit_indices": filtered_df["visit_index"].astype(int).to_numpy(),
        "feature_columns": feature_columns,
        "adjustment_columns": adjustment_columns,
        "selected_covariates": selected_covariates,
        "target_names": [_codes_to_label(negative_codes), _codes_to_label(positive_codes)],
    }


def _run_grouped_nested_cv(
        X,
        y,
        groups,
        subjects,
        source_cohorts,
        visit_indices,
        feature_selector_mode,
        n_covariates,
        outer_splits,
        inner_splits,
):
    outer_cv = _build_stratified_group_cv(y, groups, outer_splits)
    prediction_rows = []
    fold_rows = []

    for fold_index, (train_index, test_index) in enumerate(
            outer_cv.split(X, y, groups=groups),
            start=1,
    ):
        X_train, X_test = X[train_index], X[test_index]
        y_train, y_test = y[train_index], y[test_index]
        groups_train = groups[train_index]
        inner_cv = _build_stratified_group_cv(y_train, groups_train, inner_splits)

        best_estimator, search = _run_grouped_search(
            X=X_train,
            y=y_train,
            groups=groups_train,
            cv=inner_cv,
            feature_selector_mode=feature_selector_mode,
            n_covariates=n_covariates,
        )
        tuned_threshold = _estimate_grouped_threshold_from_training(
            estimator=best_estimator,
            X_train=X_train,
            y_train=y_train,
            groups_train=groups_train,
            cv=inner_cv,
        )
        fold_model = clone(best_estimator)
        fold_model = _fit_with_device_fallback(fold_model, X_train, y_train)
        y_prob = _predict_positive_probability(fold_model, X_test)
        y_pred_default = fold_model.predict(X_test).astype(int)
        y_pred_tuned = _binarize_proba(y_prob, tuned_threshold).astype(int)

        for local_pos, row_index in enumerate(test_index):
            prediction_rows.append(
                {
                    "fold": int(fold_index),
                    "#Subject": str(subjects[row_index]),
                    "person_group": str(groups[row_index]),
                    "source_cohort": str(source_cohorts[row_index]),
                    "visit_index": int(visit_indices[row_index]),
                    "is_first_visit": int(visit_indices[row_index]) == 1,
                    "diagnosis_label": "preDLB" if int(y_test[local_pos]) == 1 else "HC",
                    "y_true": int(y_test[local_pos]),
                    "y_pred_default": int(y_pred_default[local_pos]),
                    "y_pred_tuned": int(y_pred_tuned[local_pos]),
                    "pred_probability_positive": float(y_prob[local_pos]),
                    "tuned_threshold": float(tuned_threshold),
                }
            )

        fold_rows.append(
            {
                "fold": int(fold_index),
                "train_subjects": int(len(train_index)),
                "test_subjects": int(len(test_index)),
                "train_person_groups": int(len(np.unique(groups_train))),
                "test_person_groups": int(len(np.unique(groups[test_index]))),
                "test_subjects_list": ", ".join(map(str, subjects[test_index])),
                "test_person_groups_list": ", ".join(sorted(set(map(str, groups[test_index])))),
                "test_source_cohorts": ", ".join(sorted(set(map(str, source_cohorts[test_index])))),
                "inner_cv_splits": int(inner_cv.n_splits),
                "best_inner_score": float(search.best_score_),
                "tuned_threshold": float(tuned_threshold),
                "best_params_json": json.dumps(
                    _json_ready_dict(search.best_params_),
                    ensure_ascii=True,
                    sort_keys=True,
                ),
            }
        )

    return pd.DataFrame(prediction_rows), pd.DataFrame(fold_rows)


def _run_grouped_search(X, y, groups, cv, feature_selector_mode, n_covariates):
    search_settings = _search_settings_for_selector_mode(feature_selector_mode)
    search = RandomizedSearchCV(
        estimator=_build_pipeline_with_selector(
            feature_selector_mode,
            n_covariates=n_covariates,
        ),
        cv=cv,
        **search_settings,
    )
    try:
        search.fit(X, y, groups=groups)
    except Exception as exc:
        if _is_gpu_error(exc) and search.estimator.get_params().get("clf__device") == "cuda":
            logger.warning("CUDA grouped classification search failed; retrying on CPU.", exc_info=True)
            search.estimator.set_params(clf__device="cpu")
            search.fit(X, y, groups=groups)
        else:
            raise
    return search.best_estimator_, search


def _estimate_grouped_threshold_from_training(estimator, X_train, y_train, groups_train, cv):
    try:
        probabilities = _cross_val_predict_grouped_probabilities(
            estimator=estimator,
            X=X_train,
            y=y_train,
            groups=groups_train,
            cv=cv,
        )
        threshold, _ = _tune_threshold(y_train, probabilities)
        return float(threshold)
    except Exception:
        logger.warning(
            "Grouped inner-CV threshold tuning failed; using default threshold 0.5.",
            exc_info=True,
        )
        return 0.5


def _cross_val_predict_grouped_probabilities(estimator, X, y, groups, cv):
    estimator_for_cv = clone(estimator)
    try:
        return cross_val_predict(
            estimator_for_cv,
            X,
            y,
            groups=groups,
            cv=cv,
            method="predict_proba",
            n_jobs=1,
        )[:, 1]
    except Exception as exc:
        if _is_gpu_error(exc) and estimator_for_cv.get_params().get("clf__device") == "cuda":
            logger.warning("CUDA grouped cross_val_predict failed; retrying on CPU.", exc_info=True)
            estimator_for_cv.set_params(clf__device="cpu")
            return cross_val_predict(
                estimator_for_cv,
                X,
                y,
                groups=groups,
                cv=cv,
                method="predict_proba",
                n_jobs=1,
            )[:, 1]
        raise


def _fit_with_device_fallback(estimator, X, y):
    try:
        estimator.fit(X, y)
    except Exception as exc:
        if _is_gpu_error(exc) and estimator.get_params().get("clf__device") == "cuda":
            logger.warning("CUDA grouped classification fit failed; retrying on CPU.", exc_info=True)
            estimator.set_params(clf__device="cpu")
            estimator.fit(X, y)
        else:
            raise
    return estimator


def _predict_positive_probability(estimator, X):
    try:
        return estimator.predict_proba(X)[:, 1].astype(float)
    except Exception as exc:
        if _is_gpu_error(exc) and estimator.get_params().get("clf__device") == "cuda":
            logger.warning("CUDA grouped predict_proba failed; retrying on CPU.", exc_info=True)
            estimator.set_params(clf__device="cpu")
            return estimator.predict_proba(X)[:, 1].astype(float)
        raise


def _leave_one_cohort_out_validation(prepared, spec):
    rows = []
    prediction_frames = []
    cohorts = np.asarray(prepared["source_cohorts"], dtype=str)

    for cohort in sorted(set(cohorts)):
        test_mask = cohorts == cohort
        train_mask = ~test_mask
        skip_reason = _holdout_skip_reason(
            prepared["y"][train_mask],
            prepared["y"][test_mask],
            prepared["groups"][train_mask],
        )
        if skip_reason:
            rows.append(_skipped_sensitivity_row(spec, "leave_one_cohort_out", cohort, skip_reason))
            continue

        X_train = prepared["X"][train_mask]
        y_train = prepared["y"][train_mask]
        groups_train = prepared["groups"][train_mask]
        X_test = prepared["X"][test_mask]
        y_test = prepared["y"][test_mask]
        inner_cv = _build_stratified_group_cv(
            y_train,
            groups_train,
            PERSON_GROUPED_INNER_CV_SPLITS,
        )
        estimator, search = _run_grouped_search(
            X=X_train,
            y=y_train,
            groups=groups_train,
            cv=inner_cv,
            feature_selector_mode=spec["feature_selector_mode"],
            n_covariates=len(prepared["adjustment_columns"]),
        )
        threshold = _estimate_grouped_threshold_from_training(
            estimator=estimator,
            X_train=X_train,
            y_train=y_train,
            groups_train=groups_train,
            cv=inner_cv,
        )
        model = clone(estimator)
        model = _fit_with_device_fallback(model, X_train, y_train)
        y_prob = _predict_positive_probability(model, X_test)
        y_pred_default = model.predict(X_test).astype(int)
        y_pred_tuned = _binarize_proba(y_prob, threshold).astype(int)
        metrics = _binary_metrics(y_test, y_pred_default, y_prob)
        rows.append(
            {
                "run_key": spec["run_key"],
                "run_label": spec["run_label"],
                "validation_type": "leave_one_cohort_out",
                "held_out_source_cohort": cohort,
                "status": "completed",
                "train_subjects": int(train_mask.sum()),
                "test_subjects": int(test_mask.sum()),
                "train_person_groups": int(len(np.unique(groups_train))),
                "test_person_groups": int(len(np.unique(prepared["groups"][test_mask]))),
                "test_hc_count": int((y_test == 0).sum()),
                "test_predlb_count": int((y_test == 1).sum()),
                "best_inner_score": float(search.best_score_),
                "tuned_threshold": float(threshold),
                **metrics,
            }
        )
        prediction_frames.append(
            _sensitivity_predictions_frame(
                prepared=prepared,
                mask=test_mask,
                validation_type="leave_one_cohort_out",
                cohort=cohort,
                y_pred_default=y_pred_default,
                y_pred_tuned=y_pred_tuned,
                y_prob=y_prob,
                threshold=threshold,
            )
        )

    predictions = (
        pd.concat(prediction_frames, ignore_index=True)
        if prediction_frames
        else pd.DataFrame()
    )
    return pd.DataFrame(rows), predictions


def _within_cohort_nested_cv(prepared, spec):
    rows = []
    prediction_frames = []
    cohorts = np.asarray(prepared["source_cohorts"], dtype=str)

    for cohort in sorted(set(cohorts)):
        cohort_mask = cohorts == cohort
        y = prepared["y"][cohort_mask]
        groups = prepared["groups"][cohort_mask]
        skip_reason = _grouped_cv_skip_reason(y, groups)
        if skip_reason:
            rows.append(_skipped_sensitivity_row(spec, "within_cohort_nested_cv", cohort, skip_reason))
            continue

        predictions, _ = _run_grouped_nested_cv(
            X=prepared["X"][cohort_mask],
            y=y,
            groups=groups,
            subjects=prepared["subjects"][cohort_mask],
            source_cohorts=prepared["source_cohorts"][cohort_mask],
            visit_indices=prepared["visit_indices"][cohort_mask],
            feature_selector_mode=spec["feature_selector_mode"],
            n_covariates=len(prepared["adjustment_columns"]),
            outer_splits=PERSON_GROUPED_OUTER_CV_SPLITS,
            inner_splits=PERSON_GROUPED_INNER_CV_SPLITS,
        )
        metrics = _binary_metrics(
            predictions["y_true"],
            predictions["y_pred_default"],
            predictions["pred_probability_positive"],
        )
        rows.append(
            {
                "run_key": spec["run_key"],
                "run_label": spec["run_label"],
                "validation_type": "within_cohort_nested_cv",
                "source_cohort": cohort,
                "status": "completed",
                "subject_count": int(len(predictions)),
                "person_group_count": int(predictions["person_group"].nunique()),
                "hc_count": int((predictions["y_true"] == 0).sum()),
                "predlb_count": int((predictions["y_true"] == 1).sum()),
                **metrics,
            }
        )
        predictions.insert(0, "validation_type", "within_cohort_nested_cv")
        prediction_frames.append(predictions)

    predictions = (
        pd.concat(prediction_frames, ignore_index=True)
        if prediction_frames
        else pd.DataFrame()
    )
    return pd.DataFrame(rows), predictions


def _source_only_negative_control(predictions, outer_splits):
    y = predictions["y_true"].astype(int).to_numpy()
    groups = predictions["person_group"].astype(str).to_numpy()
    X = predictions[["source_cohort"]].copy()
    cv = _build_stratified_group_cv(y, groups, outer_splits)
    model = Pipeline(
        [
            (
                "encode",
                ColumnTransformer(
                    [
                        (
                            "source",
                            OneHotEncoder(handle_unknown="ignore"),
                            ["source_cohort"],
                        )
                    ]
                ),
            ),
            (
                "clf",
                LogisticRegression(
                    solver="liblinear",
                    random_state=SEED,
                    class_weight="balanced",
                ),
            ),
        ]
    )
    probabilities = cross_val_predict(
        model,
        X,
        y,
        groups=groups,
        cv=cv,
        method="predict_proba",
    )[:, 1]
    predicted = (probabilities >= 0.5).astype(int)
    metrics = _binary_metrics(y, predicted, probabilities)
    output = predictions[
        ["#Subject", "person_group", "source_cohort", "visit_index", "diagnosis_label", "y_true"]
    ].copy()
    output["source_only_probability_positive"] = probabilities
    output["source_only_predicted"] = predicted
    return metrics, output


def _cohort_distribution(predictions, spec):
    output = (
        predictions.groupby(["source_cohort", "diagnosis_label"])
        .size()
        .reset_index(name="subject_count")
    )
    output.insert(0, "run_label", spec["run_label"])
    output.insert(0, "run_key", spec["run_key"])
    return output


def _cohort_performance(predictions, spec):
    rows = []
    for cohort, cohort_df in predictions.groupby("source_cohort"):
        metrics = _binary_metrics(
            cohort_df["y_true"],
            cohort_df["y_pred_default"],
            cohort_df["pred_probability_positive"],
        )
        rows.append(
            {
                "run_key": spec["run_key"],
                "run_label": spec["run_label"],
                "source_cohort": cohort,
                "subject_count": int(len(cohort_df)),
                "person_group_count": int(cohort_df["person_group"].nunique()),
                "hc_count": int((cohort_df["y_true"] == 0).sum()),
                "predlb_count": int((cohort_df["y_true"] == 1).sum()),
                **metrics,
            }
        )
    return pd.DataFrame(rows)


def _save_prediction_outputs(
        spec_dir,
        predictions,
        y_true,
        y_prob,
        y_pred_default,
        y_pred_tuned,
        target_names,
        default_metrics,
        tuned_metrics,
):
    fpr, tpr, roc_thresholds = roc_curve(y_true, y_prob)
    precision_curve, recall_curve, pr_thresholds = _precision_recall_points(y_true, y_prob)
    _save_curve_points(
        roc_path=spec_dir / "roc_curve_points.xlsx",
        pr_path=spec_dir / "pr_curve_points.xlsx",
        fpr=fpr,
        tpr=tpr,
        roc_thresholds=roc_thresholds,
        precision_curve=precision_curve,
        recall_curve=recall_curve,
        pr_thresholds=pr_thresholds,
    )
    _classification_report_dataframe(y_true, y_pred_default, target_names).to_excel(
        spec_dir / "classification_report_default.xlsx"
    )
    _classification_report_dataframe(y_true, y_pred_tuned, target_names).to_excel(
        spec_dir / "classification_report_tuned.xlsx"
    )
    _confusion_matrix_dataframe(y_true, y_pred_default, target_names).to_excel(
        spec_dir / "confusion_matrix_default.xlsx"
    )
    _confusion_matrix_dataframe(y_true, y_pred_tuned, target_names).to_excel(
        spec_dir / "confusion_matrix_tuned.xlsx"
    )
    _save_metrics(default_metrics, spec_dir / "cls_results_original.xlsx")
    _save_metrics(tuned_metrics, spec_dir / "cls_results_tuned_nested.xlsx")
    _save_roc_pr_figure(
        y_true=y_true,
        y_prob=y_prob,
        y_pred_tuned=y_pred_tuned,
        target_names=target_names,
        tuned_threshold=float(predictions["tuned_threshold"].median()),
        output_path=spec_dir / "cls_roc.pdf",
    )


def _precision_recall_points(y_true, y_prob):
    from sklearn.metrics import precision_recall_curve

    return precision_recall_curve(y_true, y_prob)


def _build_stratified_group_cv(y, groups, max_splits):
    y = np.asarray(y, dtype=int)
    groups = np.asarray(groups, dtype=str)
    _validate_person_groups(groups, y)
    group_labels = pd.DataFrame({"group": groups, "y": y}).drop_duplicates()
    class_group_counts = group_labels.groupby("y")["group"].nunique()
    if len(class_group_counts) < 2:
        raise ValueError("Grouped CV requires at least two classes")
    n_splits = min(int(max_splits), int(class_group_counts.min()))
    if n_splits < 2:
        raise ValueError(
            "Grouped CV requires at least two person groups per class; "
            f"got {class_group_counts.to_dict()}"
        )
    return StratifiedGroupKFold(n_splits=n_splits, shuffle=True, random_state=SEED)


def _validate_person_groups(groups, y):
    group_labels = pd.DataFrame({"group": groups, "y": y}).drop_duplicates()
    mixed = group_labels.groupby("group")["y"].nunique()
    mixed = mixed[mixed > 1]
    if not mixed.empty:
        raise ValueError(
            "Person-grouped classification requires one diagnosis per person group. "
            f"Mixed groups: {mixed.index.tolist()}"
        )


def _binary_metrics(y_true, y_pred, y_prob):
    y_true = pd.Series(y_true).astype(int)
    y_pred = pd.Series(y_pred).astype(int)
    y_prob = pd.Series(y_prob).astype(float)
    if y_true.nunique() < 2:
        return {
            "AUC": np.nan,
            "PR_AUC": np.nan,
            "BACC": np.nan,
            "MCC": np.nan,
            "SEN": np.nan,
            "SPE": np.nan,
            "PRE": np.nan,
            "TN": np.nan,
            "FP": np.nan,
            "FN": np.nan,
            "TP": np.nan,
        }
    fpr, tpr, _ = roc_curve(y_true, y_prob)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return {
        "AUC": float(auc(fpr, tpr)),
        "PR_AUC": float(average_precision_score(y_true, y_prob)),
        "BACC": float(balanced_accuracy_score(y_true, y_pred)),
        "MCC": float(matthews_corrcoef(y_true, y_pred)),
        "SEN": float(recall_score(y_true, y_pred, zero_division=0)),
        "SPE": float(tn / (tn + fp)) if (tn + fp) else np.nan,
        "PRE": float(precision_score(y_true, y_pred, zero_division=0)),
        "TN": int(tn),
        "FP": int(fp),
        "FN": int(fn),
        "TP": int(tp),
    }


def _metric_columns(metrics, prefix):
    return {f"{prefix}_{key.lower()}": value for key, value in metrics.items()}


def _holdout_skip_reason(y_train, y_test, groups_train):
    if pd.Series(y_train).nunique() < 2:
        return "training set does not contain both classes"
    if pd.Series(y_test).nunique() < 2:
        return "held-out cohort does not contain both classes"
    try:
        _build_stratified_group_cv(y_train, groups_train, PERSON_GROUPED_INNER_CV_SPLITS)
    except ValueError as exc:
        return f"inner grouped CV not possible: {exc}"
    return ""


def _grouped_cv_skip_reason(y, groups):
    try:
        _build_stratified_group_cv(y, groups, PERSON_GROUPED_OUTER_CV_SPLITS)
    except ValueError as exc:
        return str(exc)
    return ""


def _skipped_sensitivity_row(spec, validation_type, cohort, reason):
    return {
        "run_key": spec["run_key"],
        "run_label": spec["run_label"],
        "validation_type": validation_type,
        "source_cohort": cohort,
        "held_out_source_cohort": cohort,
        "status": "skipped",
        "skip_reason": reason,
    }


def _sensitivity_predictions_frame(
        prepared,
        mask,
        validation_type,
        cohort,
        y_pred_default,
        y_pred_tuned,
        y_prob,
        threshold,
):
    rows = []
    indices = np.where(mask)[0]
    for local_pos, row_index in enumerate(indices):
        y_true = int(prepared["y"][row_index])
        rows.append(
            {
                "validation_type": validation_type,
                "cohort": cohort,
                "#Subject": str(prepared["subjects"][row_index]),
                "person_group": str(prepared["groups"][row_index]),
                "source_cohort": str(prepared["source_cohorts"][row_index]),
                "visit_index": int(prepared["visit_indices"][row_index]),
                "diagnosis_label": "preDLB" if y_true == 1 else "HC",
                "y_true": y_true,
                "y_pred_default": int(y_pred_default[local_pos]),
                "y_pred_tuned": int(y_pred_tuned[local_pos]),
                "pred_probability_positive": float(y_prob[local_pos]),
                "tuned_threshold": float(threshold),
            }
        )
    return pd.DataFrame(rows)


def _settings(output_dir):
    return pd.DataFrame(
        [
            {"setting": "output_dir", "value": str(output_dir)},
            {"setting": "scenario", "value": "HC vs preDLB"},
            {"setting": "outer_cv", "value": f"StratifiedGroupKFold({PERSON_GROUPED_OUTER_CV_SPLITS})"},
            {"setting": "inner_cv", "value": f"StratifiedGroupKFold({PERSON_GROUPED_INNER_CV_SPLITS})"},
            {
                "setting": "purpose",
                "value": (
                    "Thesis-oriented classifier validation with true person-grouped "
                    "outer/inner CV and cohort/source sensitivity analyses."
                ),
            },
            {
                "setting": "important_boundary",
                "value": (
                    "This is not a replacement for external validation. Source-cohort "
                    "confounding is explicitly measured and should be reported."
                ),
            },
        ]
    )


def _default_output_dir():
    return PERSON_GROUPED_RESULTS_ROOT / datetime.now().strftime("%Y%m%d_%H%M%S")


def _style_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[:200]
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10),
                60,
            )
