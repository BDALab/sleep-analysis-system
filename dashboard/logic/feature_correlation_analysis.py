import logging
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.cluster.hierarchy import fcluster, linkage
from scipy.spatial.distance import squareform
from scipy.stats import mannwhitneyu, shapiro, spearmanr

from dashboard.logic.analysis_preparation import prepare_all_analysis_datasets
from dashboard.models import Subject
from mysite.settings import MEDIA_ROOT

logger = logging.getLogger(__name__)

GROUPED_DATASET_CLINICAL_PATH = (
        Path(MEDIA_ROOT)
        / "covariates"
        / "dataset-clinical"
        / "data"
        / "grouped_clinical_matrix_with_stats.xlsx"
)
GROUPED_DATASET_CLINICAL_ACC_PATH = (
        Path(MEDIA_ROOT)
        / "covariates"
        / "dataset-clinical-acc"
        / "data"
        / "grouped_clinical_matrix_with_stats.xlsx"
)
RESULTS_ROOT = Path(MEDIA_ROOT) / "feature-correlation-analysis"

HC_CODE = 0
MCI_CODE = 2
PRE_DLB_CODE = 3
ALPHA = 0.05
FAMILY_CORRELATION_THRESHOLD = 0.85
ANALYSIS_SCENARIOS = (
    {
        "key": "predlb-vs-hc",
        "label": "preDLB vs HC",
        "reference_label": "HC",
        "reference_codes": (HC_CODE,),
        "case_label": "preDLB",
        "case_codes": (PRE_DLB_CODE,),
    },
    {
        "key": "predlb-mci-vs-hc",
        "label": "preDLB + MCI-AD vs HC",
        "reference_label": "HC",
        "reference_codes": (HC_CODE,),
        "case_label": "preDLB + MCI-AD",
        "case_codes": (PRE_DLB_CODE, MCI_CODE),
    },
    {
        "key": "mci-vs-hc",
        "label": "MCI-AD vs HC",
        "reference_label": "HC",
        "reference_codes": (HC_CODE,),
        "case_label": "MCI-AD",
        "case_codes": (MCI_CODE,),
    },
)
STATISTIC_PREFIXES = (
    "Mean.",
    "Median.",
    "Min.",
    "Max.",
    "Slope.",
    "SD.",
    "MAD.",
    "Range.",
    "IQR.",
    "CV.",
)
CLINICAL_OUTCOMES = {
    "RBDq": "rbdq",
    "UPDRS": "updrs",
    "MFS": "mfs",
    "Visuospatial": "visuospatial",
    "Attention": "attention",
    "Executive": "executive",
}
CORRELATION_COLUMNS = (
    "Features",
    "Clinical outcome",
    "n",
    "Spearman rho",
    "p",
    "adj p",
)
AGGREGATION_PREFERENCE = {
    "Median": 0,
    "IQR": 1,
    "MAD": 1,
    "Mean": 2,
    "SD": 3,
    "CV": 3,
    "Min": 4,
    "Max": 4,
    "Range": 4,
    "Slope": 5,
}


def analyze_all_feature_datasets():
    results = []
    preparations = prepare_all_analysis_datasets()
    scenarios_by_key = {
        scenario["key"]: scenario for scenario in ANALYSIS_SCENARIOS
    }
    for preparation in preparations:
        for prepared_scenario in preparation["scenarios"]:
            scenario = scenarios_by_key[prepared_scenario["scenario_key"]]
            output_path = (
                    Path(prepared_scenario["grouped_stats_path"]).parent.parent
                    / "correlation"
                    / "feature_clinical_correlation_matrix.xlsx"
            )
            results.append(
                analyze_feature_dataset(
                    prepared_scenario["grouped_stats_path"],
                    dataset_name=preparation["dataset_name"],
                    scenario=scenario,
                    output_path=output_path,
                    selected_covariates=prepared_scenario["selected_covariates"],
                )
            )
    return results


def analyze_feature_dataset(
        source_path,
        dataset_name=None,
        scenario=None,
        output_path=None,
        selected_covariates=(),
):
    source_path = Path(source_path)
    if not source_path.exists():
        raise FileNotFoundError(f"Grouped feature dataset not found: {source_path}")

    dataset_name = dataset_name or source_path.stem
    scenario = scenario or ANALYSIS_SCENARIOS[-1]
    output_path = (
        Path(output_path)
        if output_path
        else (
                RESULTS_ROOT
                / dataset_name
                / scenario["key"]
                / "feature_clinical_correlation_matrix.xlsx"
        )
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    grouped_df = pd.read_excel(source_path)
    if "#Subject" not in grouped_df.columns:
        raise KeyError(f"Dataset {source_path.name} is missing #Subject")

    analysis_df = _prepare_subject_level_data(grouped_df, scenario)
    feature_columns = [
        column
        for column in analysis_df.columns
        if str(column).startswith(STATISTIC_PREFIXES)
    ]
    if not feature_columns:
        raise ValueError(f"No feature columns found in {source_path}")

    result_rows = []
    normality_rows = []
    for feature in feature_columns:
        row, normality = _analyze_feature(analysis_df, feature, scenario)
        result_rows.append(row)
        normality_rows.extend(normality)

    results_df = pd.DataFrame(result_rows)
    output_columns = _group_output_columns(scenario)
    results_df["adj p"] = _benjamini_hochberg(
        results_df[output_columns[-2]]
    )
    results_df = results_df[output_columns]

    correlations_df = _analyze_correlations(analysis_df, feature_columns)
    correlation_matrices = {
        "correlation_rho": _correlation_matrix(correlations_df, "Spearman rho"),
        "correlation_p": _correlation_matrix(correlations_df, "p"),
        "correlation_adj_p": _correlation_matrix(correlations_df, "adj p"),
        "correlation_n": _correlation_matrix(correlations_df, "n"),
    }
    normality_df = pd.DataFrame(normality_rows)
    significant_group_df = results_df[
        pd.to_numeric(results_df["adj p"], errors="coerce") < ALPHA
        ].copy()
    significant_correlations_df = correlations_df[
        pd.to_numeric(correlations_df["adj p"], errors="coerce") < ALPHA
        ].copy()
    family_outputs = _build_interpretable_feature_families(
        analysis_df,
        feature_columns,
        results_df,
        correlations_df,
        scenario,
    )
    significant_by_outcome = {
        outcome: int(
            (
                    (correlations_df["Clinical outcome"] == outcome)
                    & (pd.to_numeric(correlations_df["adj p"], errors="coerce") < ALPHA)
            ).sum()
        )
        for outcome in CLINICAL_OUTCOMES
    }

    settings_df = pd.DataFrame(
        [
            {"setting": "source", "value": str(source_path)},
            {"setting": "scenario", "value": scenario["label"]},
            {
                "setting": "controlled_covariates",
                "value": ", ".join(selected_covariates) or "none",
            },
            {
                "setting": "cohort",
                "value": (
                    f"{scenario['reference_label']} "
                    f"(diagnosis_code={list(scenario['reference_codes'])}) vs "
                    f"{scenario['case_label']} "
                    f"(diagnosis_code={list(scenario['case_codes'])})"
                ),
            },
            {"setting": "unit_of_analysis", "value": "one row per database Subject"},
            {
                "setting": "normality_test",
                "value": (
                    "Shapiro-Wilk, separately in "
                    f"{scenario['reference_label']} and {scenario['case_label']}"
                ),
            },
            {"setting": "group_test", "value": "two-sided Mann-Whitney U"},
            {
                "setting": "correlation",
                "value": (
                    "Spearman rho for each calculated feature x clinical outcome "
                    f"pair in the {scenario['label']} cohort"
                ),
            },
            {
                "setting": "clinical_outcomes",
                "value": ", ".join(CLINICAL_OUTCOMES),
            },
            {
                "setting": "multiple_testing",
                "value": "Benjamini-Hochberg FDR separately across features for each clinical outcome",
            },
            {
                "setting": "feature_family_method",
                "value": (
                    "Project-specific conceptual families followed by complete-linkage "
                    "clustering of absolute within-cohort Spearman correlations"
                ),
            },
            {
                "setting": "feature_family_redundancy_threshold",
                "value": f"|Spearman rho| >= {FAMILY_CORRELATION_THRESHOLD:.2f}",
            },
            {
                "setting": "family_representative_selection",
                "value": (
                    "largest absolute effect, then completeness, robust aggregation "
                    "preference, non-normalized source, and FDR-adjusted p-value"
                ),
            },
            {"setting": "alpha", "value": ALPHA},
            {"setting": "feature_count", "value": len(feature_columns)},
            {"setting": "clinical_outcome_count", "value": len(CLINICAL_OUTCOMES)},
            {
                "setting": "correlation_pair_count",
                "value": int(len(correlations_df)),
            },
            {
                "setting": f"{scenario['reference_label']}_subject_count",
                "value": int((analysis_df["analysis_group"] == "reference").sum()),
            },
            {
                "setting": f"{scenario['case_label']}_subject_count",
                "value": int((analysis_df["analysis_group"] == "case").sum()),
            },
            {
                "setting": "significant_group_features_FDR",
                "value": int(len(significant_group_df)),
            },
            {
                "setting": "significant_clinical_correlations_FDR",
                "value": int(len(significant_correlations_df)),
            },
            {
                "setting": "group_family_candidate_count",
                "value": int(len(family_outputs["group_candidates"])),
            },
            {
                "setting": "correlation_family_candidate_count",
                "value": int(len(family_outputs["correlation_candidates"])),
            },
            *[
                {
                    "setting": f"significant_{outcome}_correlations_FDR",
                    "value": count,
                }
                for outcome, count in significant_by_outcome.items()
            ],
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="analysis", index=False)
        significant_group_df.to_excel(writer, sheet_name="significant_group", index=False)
        for sheet_name, matrix in correlation_matrices.items():
            matrix.to_excel(writer, sheet_name=sheet_name, index=False)
        significant_correlations_df.to_excel(
            writer,
            sheet_name="significant_correlations",
            index=False,
        )
        correlations_df.to_excel(writer, sheet_name="correlations_long", index=False)
        normality_df.to_excel(writer, sheet_name="normality_details", index=False)
        family_outputs["family_map"].to_excel(
            writer,
            sheet_name="family_map",
            index=False,
        )
        family_outputs["family_summary"].to_excel(
            writer,
            sheet_name="family_summary",
            index=False,
        )
        family_outputs["group_candidates"].to_excel(
            writer,
            sheet_name="group_family_candidates",
            index=False,
        )
        family_outputs["correlation_candidates"].to_excel(
            writer,
            sheet_name="corr_family_candidates",
            index=False,
        )
        settings_df.to_excel(writer, sheet_name="settings", index=False)
        _style_workbook(writer.book)

    logger.info(
        "Feature correlation analysis completed for %s / %s: features=%d, "
        "outcomes=%d, significant_group=%d, significant_correlations=%d, output=%s",
        dataset_name,
        scenario["label"],
        len(feature_columns),
        len(CLINICAL_OUTCOMES),
        len(significant_group_df),
        len(significant_correlations_df),
        output_path,
    )
    return {
        "dataset_name": dataset_name,
        "scenario_key": scenario["key"],
        "scenario_label": scenario["label"],
        "reference_label": scenario["reference_label"],
        "case_label": scenario["case_label"],
        "source_path": str(source_path),
        "output_path": str(output_path),
        "feature_count": int(len(feature_columns)),
        "reference_subject_count": int(
            (analysis_df["analysis_group"] == "reference").sum()
        ),
        "case_subject_count": int((analysis_df["analysis_group"] == "case").sum()),
        "significant_group_count": int(len(significant_group_df)),
        "clinical_outcome_count": int(len(CLINICAL_OUTCOMES)),
        "correlation_pair_count": int(len(correlations_df)),
        "significant_correlation_count": int(len(significant_correlations_df)),
        "group_family_candidate_count": int(
            len(family_outputs["group_candidates"])
        ),
        "correlation_family_candidate_count": int(
            len(family_outputs["correlation_candidates"])
        ),
        "significant_by_outcome": significant_by_outcome,
        "preview": results_df.head(20).replace({np.nan: None}).to_dict("records"),
    }


def _prepare_subject_level_data(grouped_df, scenario):
    prepared = grouped_df.copy()
    prepared["#Subject"] = prepared["#Subject"].astype(str).str.strip()
    subject_codes = prepared["#Subject"].dropna().unique().tolist()
    subject_rows = list(
        Subject.objects.filter(code__in=subject_codes).values(
            "code",
            "diagnosis_code",
            *CLINICAL_OUTCOMES.values(),
        )
    )
    subject_df = pd.DataFrame(subject_rows).rename(columns={"code": "#Subject"})
    if subject_df.empty:
        raise ValueError("No Subject rows matched the grouped dataset")

    prepared = prepared.merge(subject_df, on="#Subject", how="inner")
    reference_codes = set(scenario["reference_codes"])
    case_codes = set(scenario["case_codes"])
    allowed_codes = reference_codes | case_codes
    prepared = prepared[prepared["diagnosis_code"].isin(allowed_codes)].copy()
    prepared["analysis_group"] = np.where(
        prepared["diagnosis_code"].isin(reference_codes),
        "reference",
        "case",
    )
    for model_field in CLINICAL_OUTCOMES.values():
        prepared[model_field] = pd.to_numeric(prepared[model_field], errors="coerce")
    return prepared


def _group_output_columns(scenario):
    reference_label = scenario["reference_label"]
    case_label = scenario["case_label"]
    return [
        "Features",
        "Test normality",
        f"median ({reference_label})",
        f"median ({case_label})",
        f"MAD ({reference_label})",
        f"MAD ({case_label})",
        f"Mann-whitney U test ({reference_label} vs. {case_label})",
        "adj p",
    ]


def _analyze_feature(df, feature, scenario):
    reference = _numeric_values(
        df.loc[df["analysis_group"] == "reference", feature]
    )
    case = _numeric_values(df.loc[df["analysis_group"] == "case", feature])

    reference_normality = _shapiro_result(reference)
    case_normality = _shapiro_result(case)
    normality_summary = _normality_summary(
        reference_normality,
        case_normality,
        scenario["reference_label"],
        scenario["case_label"],
    )

    u_statistic = np.nan
    u_p_value = np.nan
    if len(reference) >= 1 and len(case) >= 1:
        try:
            u_statistic, u_p_value = mannwhitneyu(
                reference,
                case,
                alternative="two-sided",
                method="auto",
            )
        except ValueError:
            logger.warning("Mann-Whitney U failed for %s", feature, exc_info=True)

    output_columns = _group_output_columns(scenario)
    return {
        "Features": feature,
        "Test normality": normality_summary,
        output_columns[2]: _median(reference),
        output_columns[3]: _median(case),
        output_columns[4]: _mad(reference),
        output_columns[5]: _mad(case),
        output_columns[6]: _safe_float(u_p_value),
        "adj p": np.nan,
    }, [
        _normality_detail(
            feature,
            scenario["reference_label"],
            reference_normality,
        ),
        _normality_detail(feature, scenario["case_label"], case_normality),
    ]


def _analyze_correlations(df, feature_columns):
    rows = []
    for outcome_name, model_field in CLINICAL_OUTCOMES.items():
        outcome_rows = []
        for feature in feature_columns:
            pair_data = df[[feature, model_field]].copy()
            pair_data[feature] = pd.to_numeric(pair_data[feature], errors="coerce")
            pair_data[model_field] = pd.to_numeric(
                pair_data[model_field],
                errors="coerce",
            )
            pair_data = pair_data.replace([np.inf, -np.inf], np.nan).dropna()

            rho = np.nan
            p_value = np.nan
            if (
                    len(pair_data) >= 3
                    and pair_data[feature].nunique() > 1
                    and pair_data[model_field].nunique() > 1
            ):
                rho, p_value = spearmanr(
                    pair_data[feature],
                    pair_data[model_field],
                )

            outcome_rows.append(
                {
                    "Features": feature,
                    "Clinical outcome": outcome_name,
                    "n": int(len(pair_data)),
                    "Spearman rho": _safe_float(rho),
                    "p": _safe_float(p_value),
                    "adj p": np.nan,
                }
            )

        outcome_df = pd.DataFrame(outcome_rows)
        outcome_df["adj p"] = _benjamini_hochberg(outcome_df["p"])
        rows.append(outcome_df)

    return pd.concat(rows, ignore_index=True)[list(CORRELATION_COLUMNS)]


def _correlation_matrix(correlations_df, value_column):
    matrix = correlations_df.pivot(
        index="Features",
        columns="Clinical outcome",
        values=value_column,
    )
    matrix = matrix.reindex(columns=list(CLINICAL_OUTCOMES))
    return matrix.reset_index()


def _build_interpretable_feature_families(
        analysis_df,
        feature_columns,
        group_results_df,
        correlations_df,
        scenario,
):
    family_map = _cluster_feature_families(analysis_df, feature_columns)
    group_candidates = _select_group_family_candidates(
        analysis_df,
        group_results_df,
        family_map,
        scenario,
    )
    correlation_candidates = _select_correlation_family_candidates(
        correlations_df,
        family_map,
    )
    family_summary = _summarize_family_candidates(
        group_results_df,
        correlations_df,
        family_map,
        group_candidates,
        correlation_candidates,
    )
    return {
        "family_map": family_map,
        "family_summary": family_summary,
        "group_candidates": group_candidates,
        "correlation_candidates": correlation_candidates,
    }


def _cluster_feature_families(
        analysis_df,
        feature_columns,
        threshold=FAMILY_CORRELATION_THRESHOLD,
):
    metadata = pd.DataFrame(
        [_feature_metadata(feature) for feature in feature_columns]
    )
    cluster_labels = {}
    for family, family_rows in metadata.groupby("Feature family", sort=True):
        features = family_rows["Features"].sort_values().tolist()
        if len(features) == 1:
            labels = np.array([1], dtype=int)
        else:
            numeric = analysis_df[features].apply(
                pd.to_numeric,
                errors="coerce",
            )
            correlations = numeric.corr(
                method="spearman",
                min_periods=3,
            ).abs()
            distances = 1.0 - correlations
            distances = distances.fillna(1.0).clip(lower=0.0, upper=1.0)
            np.fill_diagonal(distances.values, 0.0)
            tree = linkage(
                squareform(distances.values, checks=False),
                method="complete",
            )
            labels = fcluster(
                tree,
                t=1.0 - threshold,
                criterion="distance",
            )

        ordered_clusters = {
            old_label: new_label
            for new_label, old_label in enumerate(
                sorted(
                    set(labels),
                    key=lambda label: min(
                        feature
                        for feature, current_label in zip(features, labels)
                        if current_label == label
                    ),
                ),
                start=1,
            )
        }
        family_slug = _slugify(family)
        for feature, label in zip(features, labels):
            cluster_labels[feature] = (
                f"{family_slug}-C{ordered_clusters[label]:02d}"
            )

    metadata["Redundancy cluster"] = metadata["Features"].map(cluster_labels)
    cluster_sizes = metadata.groupby("Redundancy cluster")["Features"].size()
    metadata["Cluster size"] = metadata["Redundancy cluster"].map(cluster_sizes)
    completeness = analysis_df[feature_columns].apply(
        lambda column: pd.to_numeric(
            column,
            errors="coerce",
        ).replace([np.inf, -np.inf], np.nan).notna().sum()
    )
    metadata["Available n"] = metadata["Features"].map(completeness).astype(int)
    metadata["Redundancy threshold"] = (
        f"|Spearman rho| >= {threshold:.2f}"
    )
    return metadata[
        [
            "Features",
            "Feature family",
            "Redundancy cluster",
            "Cluster size",
            "Source",
            "Nightly summary",
            "Measurement",
            "Normalized source",
            "Available n",
            "Redundancy threshold",
        ]
    ].sort_values(
        ["Feature family", "Redundancy cluster", "Features"]
    ).reset_index(drop=True)


def _feature_metadata(feature):
    parts = str(feature).split(".", 2)
    aggregation = parts[0] if parts else ""
    source = parts[1] if len(parts) > 1 else "unknown"
    measurement = parts[2] if len(parts) > 2 else str(feature)
    return {
        "Features": feature,
        "Feature family": _feature_family(source, measurement),
        "Source": source,
        "Nightly summary": aggregation,
        "Measurement": measurement,
        "Normalized source": source.endswith("_norm"),
    }


def _feature_family(source, measurement):
    source_lower = source.lower()
    measurement_lower = measurement.lower()
    if source_lower.startswith(("actigraphy", "diary")):
        rules = (
            ("sleep onset latency", "Sleep timing: onset latency"),
            ("wake after sleep onset", "Nocturnal wakefulness"),
            ("wake after sleep offset", "Post-sleep wakefulness"),
            ("awakening > 5 minutes", "Awakenings"),
            ("wake bouts", "Awakenings"),
            ("sleep efficiency", "Sleep efficiency"),
            ("sleep fragmentation", "Sleep fragmentation"),
            ("total sleep time", "Sleep duration"),
            ("time in bed", "Time in bed"),
        )
        for keyword, family in rules:
            if keyword in measurement_lower:
                return family
        return "Other sleep measures"

    if source_lower == "activity":
        if re.search(r"\b(1st|5th|10th|20th) percentile\b", measurement_lower):
            return "Activity level: lower distribution"
        if re.search(r"\b(80th|90th|95th|99th) percentile\b", measurement_lower):
            return "Activity level: upper distribution"
        if any(
                keyword in measurement_lower
                for keyword in (
                        "standard deviation",
                        "variance",
                        "median absolute deviation",
                        "index of dispersion",
                        "range",
                        "interquartile",
                        "interpercentile",
                        "interdencile",
                        "modulation",
                )
        ):
            return "Activity variability and dispersion"
        if any(
                keyword in measurement_lower
                for keyword in (
                        "skewness",
                        "kurtosis",
                )
        ):
            return "Activity distribution shape"
        if any(
                keyword in measurement_lower
                for keyword in (
                        "entropy",
                        "teager kaiser",
                )
        ):
            return "Activity complexity and energy"
        if any(
                keyword in measurement_lower
                for keyword in (
                        "relative position",
                        "max",
                        "min",
                )
        ):
            return "Activity extrema and timing"
        if any(
                keyword in measurement_lower
                for keyword in (
                        "mean",
                        "median",
                        "mode",
                        "harmonic mean",
                )
        ):
            return "Activity level: central tendency"
        return "Other activity measures"

    return "Other measures"


def _select_group_family_candidates(
        analysis_df,
        group_results_df,
        family_map,
        scenario,
):
    significant = group_results_df[
        pd.to_numeric(group_results_df["adj p"], errors="coerce") < ALPHA
        ].copy()
    columns = [
        "Feature family",
        "Redundancy cluster",
        "Representative feature",
        "Cluster members",
        "Member count",
        "Significant member count",
        f"median ({scenario['reference_label']})",
        f"median ({scenario['case_label']})",
        "Rank-biserial effect",
        "Absolute effect",
        "Effect direction",
        "Available n",
        "FDR p",
        "Selection basis",
    ]
    if significant.empty:
        return pd.DataFrame(columns=columns)

    candidates = significant.merge(family_map, on="Features", how="left")
    effect_rows = []
    for feature in candidates["Features"]:
        reference = _numeric_values(
            analysis_df.loc[
                analysis_df["analysis_group"] == "reference",
                feature,
            ]
        )
        case = _numeric_values(
            analysis_df.loc[
                analysis_df["analysis_group"] == "case",
                feature,
            ]
        )
        effect = _rank_biserial_effect(reference, case)
        effect_rows.append(
            {
                "Features": feature,
                "Rank-biserial effect": effect,
                "Absolute effect": abs(effect) if effect is not None else np.nan,
                "Available n": int(len(reference) + len(case)),
            }
        )
    effects = pd.DataFrame(effect_rows)
    candidates = candidates.drop(columns=["Available n"]).merge(
        effects,
        on="Features",
        how="left",
    )
    candidates = _add_representative_sort_columns(candidates)

    output_rows = []
    for cluster, rows in candidates.groupby("Redundancy cluster", sort=True):
        representative = rows.sort_values(
            [
                "Absolute effect",
                "Available n",
                "_aggregation_preference",
                "_normalized_preference",
                "adj p",
                "Features",
            ],
            ascending=[False, False, True, True, True, True],
        ).iloc[0]
        cluster_members = family_map.loc[
            family_map["Redundancy cluster"] == cluster,
            "Features",
        ].sort_values().tolist()
        effect = representative["Rank-biserial effect"]
        direction = (
            f"{scenario['reference_label']} higher"
            if effect > 0
            else f"{scenario['case_label']} higher"
            if effect < 0
            else "no direction"
        )
        output_rows.append(
            {
                "Feature family": representative["Feature family"],
                "Redundancy cluster": cluster,
                "Representative feature": representative["Features"],
                "Cluster members": "; ".join(cluster_members),
                "Member count": len(cluster_members),
                "Significant member count": int(len(rows)),
                f"median ({scenario['reference_label']})": representative[
                    f"median ({scenario['reference_label']})"
                ],
                f"median ({scenario['case_label']})": representative[
                    f"median ({scenario['case_label']})"
                ],
                "Rank-biserial effect": effect,
                "Absolute effect": representative["Absolute effect"],
                "Effect direction": direction,
                "Available n": int(representative["Available n"]),
                "FDR p": representative["adj p"],
                "Selection basis": (
                    "Largest absolute group effect in the redundancy cluster; "
                    "ties favor completeness, robust summary, raw source, then FDR p"
                ),
            }
        )
    return pd.DataFrame(output_rows, columns=columns).sort_values(
        ["Absolute effect", "FDR p"],
        ascending=[False, True],
    ).reset_index(drop=True)


def _select_correlation_family_candidates(correlations_df, family_map):
    significant = correlations_df[
        pd.to_numeric(correlations_df["adj p"], errors="coerce") < ALPHA
        ].copy()
    columns = [
        "Clinical outcome",
        "Feature family",
        "Redundancy cluster",
        "Representative feature",
        "Cluster members",
        "Member count",
        "Significant member count",
        "n",
        "Spearman rho",
        "Absolute rho",
        "Direction",
        "p",
        "FDR p",
        "Selection basis",
    ]
    if significant.empty:
        return pd.DataFrame(columns=columns)

    candidates = significant.merge(family_map, on="Features", how="left")
    candidates["Absolute rho"] = candidates["Spearman rho"].abs()
    candidates = _add_representative_sort_columns(candidates)
    output_rows = []
    for (outcome, cluster), rows in candidates.groupby(
            ["Clinical outcome", "Redundancy cluster"],
            sort=True,
    ):
        representative = rows.sort_values(
            [
                "Absolute rho",
                "n",
                "_aggregation_preference",
                "_normalized_preference",
                "adj p",
                "Features",
            ],
            ascending=[False, False, True, True, True, True],
        ).iloc[0]
        cluster_members = family_map.loc[
            family_map["Redundancy cluster"] == cluster,
            "Features",
        ].sort_values().tolist()
        rho = representative["Spearman rho"]
        output_rows.append(
            {
                "Clinical outcome": outcome,
                "Feature family": representative["Feature family"],
                "Redundancy cluster": cluster,
                "Representative feature": representative["Features"],
                "Cluster members": "; ".join(cluster_members),
                "Member count": len(cluster_members),
                "Significant member count": int(len(rows)),
                "n": int(representative["n"]),
                "Spearman rho": rho,
                "Absolute rho": abs(rho),
                "Direction": "positive" if rho > 0 else "negative",
                "p": representative["p"],
                "FDR p": representative["adj p"],
                "Selection basis": (
                    "Largest absolute rho in the outcome-specific redundancy "
                    "cluster; ties favor completeness, robust summary, raw source, "
                    "then FDR p"
                ),
            }
        )
    return pd.DataFrame(output_rows, columns=columns).sort_values(
        ["Clinical outcome", "Absolute rho", "FDR p"],
        ascending=[True, False, True],
    ).reset_index(drop=True)


def _summarize_family_candidates(
        group_results_df,
        correlations_df,
        family_map,
        group_candidates,
        correlation_candidates,
):
    rows = []
    significant_group = group_results_df[
        pd.to_numeric(group_results_df["adj p"], errors="coerce") < ALPHA
        ].merge(family_map, on="Features", how="left")
    for family, family_rows in significant_group.groupby(
            "Feature family",
            sort=True,
    ):
        selected = group_candidates[
            group_candidates["Feature family"] == family
            ]
        rows.append(
            {
                "Analysis": "Group difference",
                "Clinical outcome": "",
                "Feature family": family,
                "Significant variants": int(len(family_rows)),
                "Redundancy clusters": int(
                    family_rows["Redundancy cluster"].nunique()
                ),
                "Candidate representatives": int(len(selected)),
            }
        )

    significant_correlations = correlations_df[
        pd.to_numeric(correlations_df["adj p"], errors="coerce") < ALPHA
        ].merge(family_map, on="Features", how="left")
    for (outcome, family), family_rows in significant_correlations.groupby(
            ["Clinical outcome", "Feature family"],
            sort=True,
    ):
        selected = correlation_candidates[
            (correlation_candidates["Clinical outcome"] == outcome)
            & (correlation_candidates["Feature family"] == family)
            ]
        rows.append(
            {
                "Analysis": "Clinical correlation",
                "Clinical outcome": outcome,
                "Feature family": family,
                "Significant variants": int(len(family_rows)),
                "Redundancy clusters": int(
                    family_rows["Redundancy cluster"].nunique()
                ),
                "Candidate representatives": int(len(selected)),
            }
        )
    return pd.DataFrame(
        rows,
        columns=[
            "Analysis",
            "Clinical outcome",
            "Feature family",
            "Significant variants",
            "Redundancy clusters",
            "Candidate representatives",
        ],
    )


def _add_representative_sort_columns(df):
    ranked = df.copy()
    ranked["_aggregation_preference"] = ranked["Nightly summary"].map(
        AGGREGATION_PREFERENCE
    ).fillna(9)
    ranked["_normalized_preference"] = ranked["Normalized source"].astype(int)
    return ranked


def _rank_biserial_effect(reference, case):
    if not len(reference) or not len(case):
        return None
    try:
        statistic, _ = mannwhitneyu(
            reference,
            case,
            alternative="two-sided",
            method="auto",
        )
    except ValueError:
        return None
    return _safe_float(
        (2.0 * statistic / (len(reference) * len(case))) - 1.0
    )


def _slugify(value):
    slug = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return slug or "family"


def _numeric_values(series):
    return pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()


def _shapiro_result(values):
    if len(values) < 3:
        return {
            "n": int(len(values)),
            "statistic": None,
            "p_value": None,
            "normal": None,
            "status": "insufficient",
        }
    if values.nunique() < 2:
        return {
            "n": int(len(values)),
            "statistic": None,
            "p_value": None,
            "normal": None,
            "status": "constant",
        }
    statistic, p_value = shapiro(values)
    return {
        "n": int(len(values)),
        "statistic": _safe_float(statistic),
        "p_value": _safe_float(p_value),
        "normal": bool(p_value >= ALPHA),
        "status": "tested",
    }


def _normality_summary(
        reference_result,
        case_result,
        reference_label,
        case_label,
):
    def describe(label, result):
        if result["status"] == "constant":
            return f"{label}: constant (n={result['n']})"
        if result["p_value"] is None:
            return f"{label}: insufficient n={result['n']}"
        distribution = "normal" if result["normal"] else "non-normal"
        return f"{label}: {distribution} (p={result['p_value']:.4g})"

    return (
        f"{describe(reference_label, reference_result)}; "
        f"{describe(case_label, case_result)}"
    )


def _normality_detail(feature, group, result):
    return {
        "Features": feature,
        "group": group,
        "n": result["n"],
        "Shapiro-Wilk W": result["statistic"],
        "p": result["p_value"],
        "normal_at_alpha_0.05": result["normal"],
        "status": result["status"],
    }


def _median(values):
    return _safe_float(values.median()) if len(values) else None


def _mad(values):
    if not len(values):
        return None
    median = values.median()
    return _safe_float((values - median).abs().median())


def _benjamini_hochberg(p_values):
    numeric = pd.to_numeric(p_values, errors="coerce")
    adjusted = pd.Series(np.nan, index=numeric.index, dtype=float)
    valid = numeric.dropna()
    if valid.empty:
        return adjusted

    ordered = valid.sort_values()
    count = len(ordered)
    raw_adjusted = ordered.to_numpy(dtype=float) * count / np.arange(1, count + 1)
    monotonic = np.minimum.accumulate(raw_adjusted[::-1])[::-1]
    adjusted.loc[ordered.index] = np.clip(monotonic, 0.0, 1.0)
    return adjusted


def _style_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    significant_fill = PatternFill("solid", fgColor="E2F0D9")
    header_font = Font(color="FFFFFF", bold=True)
    analysis_widths = {
        1: 42,
        2: 48,
        3: 14,
        4: 14,
        5: 14,
        6: 14,
        7: 20,
        8: 14,
    }

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 12),
                48,
            )

    for sheet_name in ("analysis", "significant_group"):
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "B2"
        worksheet.row_dimensions[1].height = 38
        for column, width in analysis_widths.items():
            worksheet.column_dimensions[get_column_letter(column)].width = width

        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
            for column in range(3, 9):
                worksheet.cell(row, column).alignment = Alignment(
                    horizontal="right",
                    vertical="top",
                )
            for column in (3, 4, 5, 6):
                worksheet.cell(row, column).number_format = "0.0000"
            for column in (7, 8):
                worksheet.cell(row, column).number_format = "0.000E+00"
            value = worksheet.cell(row, 8).value
            if isinstance(value, (int, float)) and value < ALPHA:
                worksheet.cell(row, 8).fill = significant_fill
                worksheet.cell(row, 8).font = Font(bold=True)

    for sheet_name in (
            "correlation_rho",
            "correlation_p",
            "correlation_adj_p",
            "correlation_n",
    ):
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "B2"
        worksheet.column_dimensions["A"].width = 48
        for column in range(2, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column)].width = 16
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, column)
                cell.alignment = Alignment(horizontal="right")
                if sheet_name == "correlation_n":
                    cell.number_format = "0"
                elif sheet_name == "correlation_rho":
                    cell.number_format = "0.0000"
                else:
                    cell.number_format = "0.000E+00"
                if (
                        sheet_name == "correlation_adj_p"
                        and isinstance(cell.value, (int, float))
                        and cell.value < ALPHA
                ):
                    cell.fill = significant_fill
                    cell.font = Font(bold=True)

    for sheet_name in ("significant_correlations", "correlations_long"):
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        widths = (48, 18, 10, 16, 16, 16)
        for column, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(column)].width = width
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, 3).number_format = "0"
            worksheet.cell(row, 4).number_format = "0.0000"
            worksheet.cell(row, 5).number_format = "0.000E+00"
            worksheet.cell(row, 6).number_format = "0.000E+00"
            value = worksheet.cell(row, 6).value
            if isinstance(value, (int, float)) and value < ALPHA:
                worksheet.cell(row, 6).fill = significant_fill
                worksheet.cell(row, 6).font = Font(bold=True)

    family_map = workbook["family_map"]
    family_map.freeze_panes = "A2"
    family_widths = (48, 38, 34, 12, 18, 18, 38, 18, 12, 24)
    for column, width in enumerate(family_widths, start=1):
        family_map.column_dimensions[get_column_letter(column)].width = width

    family_summary = workbook["family_summary"]
    summary_widths = (22, 18, 38, 20, 20, 24)
    for column, width in enumerate(summary_widths, start=1):
        family_summary.column_dimensions[get_column_letter(column)].width = width

    group_candidates = workbook["group_family_candidates"]
    group_candidates.freeze_panes = "A2"
    group_candidate_widths = (
        38, 34, 48, 70, 14, 18, 16, 16, 16, 16, 16, 12, 14, 58,
    )
    for column, width in enumerate(group_candidate_widths, start=1):
        group_candidates.column_dimensions[
            get_column_letter(column)
        ].width = width
    for row in range(2, group_candidates.max_row + 1):
        for column in range(1, group_candidates.max_column + 1):
            group_candidates.cell(row, column).alignment = Alignment(
                wrap_text=column in (3, 4, 14),
                vertical="top",
            )
        group_candidates.row_dimensions[row].height = 58

    correlation_candidates = workbook["corr_family_candidates"]
    correlation_candidates.freeze_panes = "A2"
    correlation_candidate_widths = (
        18, 38, 34, 48, 70, 14, 18, 10, 16, 16, 12, 14, 14, 58,
    )
    for column, width in enumerate(correlation_candidate_widths, start=1):
        correlation_candidates.column_dimensions[
            get_column_letter(column)
        ].width = width
    for row in range(2, correlation_candidates.max_row + 1):
        for column in range(1, correlation_candidates.max_column + 1):
            correlation_candidates.cell(row, column).alignment = Alignment(
                wrap_text=column in (4, 5, 14),
                vertical="top",
            )
        correlation_candidates.row_dimensions[row].height = 72

    settings = workbook["settings"]
    settings.column_dimensions["A"].width = 42
    settings.column_dimensions["B"].width = 85
    for row in range(2, settings.max_row + 1):
        settings.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        if len(str(settings.cell(row, 2).value or "")) > 70:
            settings.row_dimensions[row].height = 32


def _safe_float(value):
    return float(value) if value is not None and pd.notna(value) else None
