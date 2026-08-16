import json
import logging
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import chi2_contingency
from sklearn.compose import ColumnTransformer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    auc,
    average_precision_score,
    balanced_accuracy_score,
    confusion_matrix,
    matthews_corrcoef,
    precision_score,
    recall_score,
    roc_curve,
)
from sklearn.model_selection import LeaveOneOut, cross_val_predict
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder

from mysite.settings import MEDIA_ROOT

logger = logging.getLogger(__name__)

VALIDITY_RESULTS_ROOT = Path(MEDIA_ROOT) / "classification" / "validity-checks" / "hc-vs-predlb"

HC_VS_PREDLB_STRICT_RUNS = (
    {
        "run_key": "strict_rfe",
        "label": "Broad strict RFE",
        "run_dir": (
                Path(MEDIA_ROOT)
                / "classification"
                / "grouped-statistics-strict-with-covariates"
                / "dataset-clinical-rfe"
                / "20260629_173137"
        ),
        "interpretation": "best broad strict classifier",
    },
    {
        "run_key": "stable_sleep",
        "label": "Stable primary sleep families",
        "run_dir": (
                Path(MEDIA_ROOT)
                / "classification"
                / "grouped-statistics-strict-with-covariates"
                / "dataset-clinical-stable-primary-sleep-hc-predlb"
                / "20260701_120433"
        ),
        "interpretation": "interpretable sleep-family sensitivity classifier",
    },
    {
        "run_key": "stable_sleep_activity",
        "label": "Stable primary sleep + activity variability",
        "run_dir": (
                Path(MEDIA_ROOT)
                / "classification"
                / "grouped-statistics-strict-with-covariates"
                / "dataset-clinical-acc-stable-primary-sleep-activity-hc-predlb"
                / "20260701_125206"
        ),
        "interpretation": "activity-enhanced stable-family sensitivity classifier",
    },
)


def run_hc_vs_predlb_classification_validity_checks(output_dir=None):
    output_dir = Path(output_dir) if output_dir else _default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)

    summary_rows = []
    subject_frames = []
    repeat_frames = []
    cohort_frames = []
    cohort_metric_frames = []
    source_only_rows = []

    for spec in HC_VS_PREDLB_STRICT_RUNS:
        result = _analyze_run(spec)
        summary_rows.append(result["summary"])
        subject_frames.append(result["subjects"])
        if not result["person_repeats"].empty:
            repeat_frames.append(result["person_repeats"])
        cohort_frames.append(result["cohort_distribution"])
        cohort_metric_frames.append(result["cohort_metrics"])
        source_only_rows.append(result["source_only"])

    summary_df = pd.DataFrame(summary_rows)
    subjects_df = pd.concat(subject_frames, ignore_index=True)
    repeats_df = (
        pd.concat(repeat_frames, ignore_index=True)
        if repeat_frames
        else pd.DataFrame()
    )
    cohort_distribution_df = pd.concat(cohort_frames, ignore_index=True)
    cohort_metrics_df = pd.concat(cohort_metric_frames, ignore_index=True)
    source_only_df = pd.DataFrame(source_only_rows)
    settings_df = _settings(output_dir)

    output_path = output_dir / "hc_vs_predlb_classification_validity_checks.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        subjects_df.to_excel(writer, sheet_name="subjects_annotated", index=False)
        repeats_df.to_excel(writer, sheet_name="person_repeats", index=False)
        cohort_distribution_df.to_excel(writer, sheet_name="cohort_distribution", index=False)
        cohort_metrics_df.to_excel(writer, sheet_name="cohort_performance", index=False)
        source_only_df.to_excel(writer, sheet_name="source_ascertainment", index=False)
        settings_df.to_excel(writer, sheet_name="settings", index=False)
        _style_workbook(writer.book)

    result = {
        "run_dir": str(output_dir),
        "output_path": str(output_path),
        "summary": summary_df.replace({np.nan: None}).to_dict("records"),
    }
    (output_dir / "hc_vs_predlb_classification_validity_checks.json").write_text(
        json.dumps(result, indent=2),
        encoding="utf-8",
    )
    logger.info("HC-vs-preDLB classification validity checks completed: %s", output_path)
    return result


def subject_person_id(subject_code):
    code = str(subject_code).strip()
    hc_match = re.match(
        r"^HC(?P<visit>[23])?(?P<sep>[-_])(?P<person>.+)$",
        code,
        flags=re.IGNORECASE,
    )
    if hc_match:
        return f"HC{hc_match.group('sep')}{hc_match.group('person')}"

    predlb_match = re.match(
        r"^(?P<base>pre[-_]?(?:LBD|DLB))(?P<visit>[23])?(?P<sep>[-_])(?P<person>.+)$",
        code,
        flags=re.IGNORECASE,
    )
    if predlb_match:
        return (
            f"{predlb_match.group('base')}"
            f"{predlb_match.group('sep')}"
            f"{predlb_match.group('person')}"
        )

    return code


def subject_visit_index(subject_code):
    code = str(subject_code).strip()
    visit_match = re.match(r"^HC(?P<visit>[23])[-_]", code, flags=re.IGNORECASE)
    if visit_match:
        return int(visit_match.group("visit"))

    visit_match = re.match(
        r"^pre[-_]?(?:LBD|DLB)(?P<visit>[23])[-_]",
        code,
        flags=re.IGNORECASE,
    )
    if visit_match:
        return int(visit_match.group("visit"))
    return 1


def subject_source_cohort(subject_code):
    code = str(subject_code).strip()
    if re.match(r"^COBEN-", code, flags=re.IGNORECASE):
        return "COBEN"
    if re.match(r"^HC[23]?[-_]", code, flags=re.IGNORECASE):
        return "HC/HC2"
    if re.match(r"^pre[-_]?(?:LBD|DLB)[23]?[-_]", code, flags=re.IGNORECASE):
        return "pre-LBD/pre-LBD2"
    if re.match(r"^MY-HC-", code, flags=re.IGNORECASE):
        return "MY-HC"
    if re.match(r"^DREAMT", code, flags=re.IGNORECASE):
        return "DREAMT"
    if re.match(r"^MECSLEEP", code, flags=re.IGNORECASE):
        return "MECSLEEP"
    if re.match(r"^FNOL", code, flags=re.IGNORECASE):
        return "FNOL"
    prefix = re.match(r"^[A-Za-z]+(?:-[A-Za-z]+)?", code)
    return prefix.group(0) if prefix else "unknown"


def _analyze_run(spec):
    run_dir = Path(spec["run_dir"])
    scenario_dir = run_dir / "scenario-preDLB_vs_HC"
    predictions_path = scenario_dir / "subject_predictions.xlsx"
    subjects_path = scenario_dir / "scenario_subjects.xlsx"
    summary_path = run_dir / "classification_summary.xlsx"

    for path in (predictions_path, subjects_path, summary_path):
        if not path.exists():
            raise FileNotFoundError(f"Missing validity-check input: {path}")

    predictions = pd.read_excel(predictions_path)
    subjects = pd.read_excel(subjects_path)
    data = predictions.merge(subjects, on="#Subject", how="left", suffixes=("", "_subject"))
    data["run_key"] = spec["run_key"]
    data["run_label"] = spec["label"]
    data["person_id"] = data["#Subject"].map(subject_person_id)
    data["visit_index"] = data["#Subject"].map(subject_visit_index)
    data["is_first_visit"] = data["visit_index"].eq(1)
    data["source_cohort"] = data["#Subject"].map(subject_source_cohort)
    data["diagnosis_label"] = data["y_true"].map({0: "HC", 1: "preDLB"})

    default_summary = pd.read_excel(summary_path, sheet_name="nested_default_metrics").iloc[0]
    tuned_summary = pd.read_excel(summary_path, sheet_name="nested_tuned_metrics").iloc[0]
    first_visit_default = _metrics(
        data.loc[data["is_first_visit"], "y_true"],
        data.loc[data["is_first_visit"], "y_pred_default"],
        data.loc[data["is_first_visit"], "pred_probability_positive"],
    )
    first_visit_tuned = _metrics(
        data.loc[data["is_first_visit"], "y_true"],
        data.loc[data["is_first_visit"], "y_pred_tuned"],
        data.loc[data["is_first_visit"], "pred_probability_positive"],
    )
    repeated_subject_count = int(data["person_id"].duplicated(keep=False).sum())
    repeated_person_count = int((data.groupby("person_id")["#Subject"].nunique() > 1).sum())
    cohort_table = pd.crosstab(data["source_cohort"], data["diagnosis_label"])
    cohort_chi2 = _cohort_chi_square(cohort_table)
    source_only = _source_only_ascertainment_benchmark(data)

    summary = {
        "run_key": spec["run_key"],
        "run_label": spec["label"],
        "interpretation": spec["interpretation"],
        "run_dir": str(run_dir),
        "subject_count": int(len(data)),
        "person_count": int(data["person_id"].nunique()),
        "repeated_person_count": repeated_person_count,
        "subjects_from_repeated_people": repeated_subject_count,
        "first_visit_subject_count": int(data["is_first_visit"].sum()),
        "all_visits_auc": float(default_summary["roc_auc"]),
        "all_visits_pr_auc": float(default_summary["pr_auc"]),
        "all_visits_default_bacc": float(default_summary["BACC"]),
        "all_visits_default_sensitivity": float(default_summary["SEN"]),
        "all_visits_default_specificity": float(default_summary["SPE"]),
        "all_visits_tuned_bacc": float(tuned_summary["BACC"]),
        "first_visit_default_auc": first_visit_default["AUC"],
        "first_visit_default_bacc": first_visit_default["BACC"],
        "first_visit_default_sensitivity": first_visit_default["SEN"],
        "first_visit_default_specificity": first_visit_default["SPE"],
        "first_visit_tuned_bacc": first_visit_tuned["BACC"],
        "source_cohort_count": int(data["source_cohort"].nunique()),
        **cohort_chi2,
        "source_only_auc": source_only["source_only_auc"],
        "source_only_bacc": source_only["source_only_bacc"],
        "source_only_note": source_only["source_only_note"],
    }

    repeats = _person_repeats(data, spec)
    cohort_distribution = _cohort_distribution(data, spec)
    cohort_metrics = _cohort_metrics(data, spec)
    source_only["run_key"] = spec["run_key"]
    source_only["run_label"] = spec["label"]
    return {
        "summary": summary,
        "subjects": data,
        "person_repeats": repeats,
        "cohort_distribution": cohort_distribution,
        "cohort_metrics": cohort_metrics,
        "source_only": source_only,
    }


def _metrics(y_true, y_pred, y_prob):
    y_true = pd.Series(y_true).astype(int)
    y_pred = pd.Series(y_pred).astype(int)
    y_prob = pd.Series(y_prob).astype(float)
    if y_true.nunique() < 2:
        return {
            "AUC": np.nan,
            "PR_AUC": np.nan,
            "BACC": np.nan,
            "MCC": np.nan,
            "SEN": np.nan,
            "SPE": np.nan,
            "PRE": np.nan,
        }
    fpr, tpr, _ = roc_curve(y_true, y_prob)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return {
        "AUC": float(auc(fpr, tpr)),
        "PR_AUC": float(average_precision_score(y_true, y_prob)),
        "BACC": float(balanced_accuracy_score(y_true, y_pred)),
        "MCC": float(matthews_corrcoef(y_true, y_pred)),
        "SEN": float(recall_score(y_true, y_pred, zero_division=0)),
        "SPE": float(tn / (tn + fp)) if (tn + fp) else np.nan,
        "PRE": float(precision_score(y_true, y_pred, zero_division=0)),
        "TN": int(tn),
        "FP": int(fp),
        "FN": int(fn),
        "TP": int(tp),
    }


def _person_repeats(data, spec):
    grouped = data.groupby("person_id").agg(
        subject_count=("#Subject", "nunique"),
        subjects=("#Subject", lambda values: ", ".join(sorted(map(str, values)))),
        diagnoses=("diagnosis_label", lambda values: ", ".join(sorted(set(map(str, values))))),
        visits=("visit_index", lambda values: ", ".join(map(str, sorted(set(values))))),
    ).reset_index()
    repeated = grouped[grouped["subject_count"] > 1].copy()
    if repeated.empty:
        return repeated
    repeated.insert(0, "run_label", spec["label"])
    repeated.insert(0, "run_key", spec["run_key"])
    return repeated


def _cohort_distribution(data, spec):
    cohort = (
        data.groupby(["source_cohort", "diagnosis_label"])
        .size()
        .reset_index(name="subject_count")
    )
    cohort.insert(0, "run_label", spec["label"])
    cohort.insert(0, "run_key", spec["run_key"])
    return cohort


def _cohort_metrics(data, spec):
    rows = []
    for cohort, cohort_df in data.groupby("source_cohort"):
        metric = _metrics(
            cohort_df["y_true"],
            cohort_df["y_pred_default"],
            cohort_df["pred_probability_positive"],
        )
        rows.append(
            {
                "run_key": spec["run_key"],
                "run_label": spec["label"],
                "source_cohort": cohort,
                "subject_count": int(len(cohort_df)),
                "hc_count": int((cohort_df["y_true"] == 0).sum()),
                "predlb_count": int((cohort_df["y_true"] == 1).sum()),
                **metric,
            }
        )
    return pd.DataFrame(rows)


def _cohort_chi_square(cohort_table):
    if cohort_table.shape[0] < 2 or cohort_table.shape[1] < 2:
        return {
            "cohort_diagnosis_chi2": np.nan,
            "cohort_diagnosis_p": np.nan,
            "cohort_diagnosis_cramers_v": np.nan,
            "cohort_diagnosis_note": "not estimable",
        }
    chi2, p_value, _, _ = chi2_contingency(cohort_table)
    n = cohort_table.values.sum()
    min_dim = min(cohort_table.shape) - 1
    cramers_v = np.sqrt(chi2 / (n * min_dim)) if n and min_dim else np.nan
    pure_cohorts = []
    for cohort, row in cohort_table.iterrows():
        nonzero = (row > 0).sum()
        if nonzero == 1:
            pure_cohorts.append(str(cohort))
    note = (
        "diagnosis and source cohort are strongly entangled; pure cohorts: "
        + ", ".join(pure_cohorts)
        if pure_cohorts
        else "all observed cohorts contain both classes"
    )
    return {
        "cohort_diagnosis_chi2": float(chi2),
        "cohort_diagnosis_p": float(p_value),
        "cohort_diagnosis_cramers_v": float(cramers_v),
        "cohort_diagnosis_note": note,
    }


def _source_only_ascertainment_benchmark(data):
    if data["source_cohort"].nunique() < 2 or data["y_true"].nunique() < 2:
        return {
            "source_only_auc": np.nan,
            "source_only_bacc": np.nan,
            "source_only_note": "not estimable",
        }
    X = data[["source_cohort"]].copy()
    y = data["y_true"].astype(int).to_numpy()
    model = Pipeline(
        [
            (
                "encode",
                ColumnTransformer(
                    [
                        (
                            "source",
                            OneHotEncoder(handle_unknown="ignore"),
                            ["source_cohort"],
                        )
                    ]
                ),
            ),
            (
                "clf",
                LogisticRegression(
                    solver="liblinear",
                    random_state=17,
                    class_weight="balanced",
                ),
            ),
        ]
    )
    try:
        probabilities = cross_val_predict(
            model,
            X,
            y,
            cv=LeaveOneOut(),
            method="predict_proba",
        )[:, 1]
        predictions = (probabilities >= 0.5).astype(int)
        fpr, tpr, _ = roc_curve(y, probabilities)
        return {
            "source_only_auc": float(auc(fpr, tpr)),
            "source_only_bacc": float(balanced_accuracy_score(y, predictions)),
            "source_only_note": (
                "source-only ascertainment benchmark; high values quantify expected label "
                "enrichment from consortium-defined recruitment and do not by themselves "
                "prove technical acquisition confounding"
            ),
        }
    except Exception as exc:
        logger.warning("Source-only ascertainment benchmark failed", exc_info=True)
        return {
            "source_only_auc": np.nan,
            "source_only_bacc": np.nan,
            "source_only_note": f"failed: {exc}",
        }


def _settings(output_dir):
    return pd.DataFrame(
        [
            {"setting": "output_dir", "value": str(output_dir)},
            {"setting": "scenario", "value": "HC vs preDLB"},
            {
                "setting": "purpose",
                "value": (
                    "Audit existing strict HC-vs-preDLB classifiers for repeated-person leakage, "
                    "first-visit-only performance, cohort/source entanglement, and source-only "
                    "ascertainment predictability. Source is diagnosis-enriched by recruitment "
                    "design, so this is not a negative control for acquisition artefacts."
                ),
            },
            {
                "setting": "person_id_rule",
                "value": "HC2-X maps to HC-X; pre-LBD2-X maps to pre-LBD-X; other codes remain unique.",
            },
            {
                "setting": "first_visit_rule",
                "value": "Exclude HC2-* and pre-LBD2-* second-visit rows.",
            },
        ]
    )


def _default_output_dir():
    return VALIDITY_RESULTS_ROOT / datetime.now().strftime("%Y%m%d_%H%M%S")


def _style_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[:200]
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10),
                60,
            )
