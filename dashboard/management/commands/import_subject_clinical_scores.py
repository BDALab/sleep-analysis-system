from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from dashboard.models import Subject

DEFAULT_KARDIOVIZE_PATH = Path(r"E:\geneactiv-processing-data\KARDIOVIZE.xlsx")
DEFAULT_PREDLB_PATH = Path(r"E:\geneactiv-processing-data\preDLB_shared.xlsx")

# KARDIOVIZE.xlsx -> sheet ID_Clinical_Cognitive
KARDIO_SHEET = "ID_Clinical_Cognitive"
KARDIO_ID_COLUMN = "A"
KARDIO_MAPPING = {
    "rbdq": "G",
    "updrs": "E",
    "mfs": "I",
    "attention": "O",
    "executive": "P",
    "visuospatial": "Q",
    "education_years": "D",
}

# preDLB_shared.xlsx
SCALES_SHEET = "scales_logo_TCS"
SCALES_ID_COLUMN = "A"
SCALES_MAPPING = {
    "rbdq": ("AB", "AT", "BM"),
    "updrs": ("Y", "AQ", "BJ"),
    "mfs": ("AD", "AV", "BO"),
}
PSY_SHEET = "PSY_RAW"
PSY_ID_COLUMN = "A"
PSY_MAPPING = {
    "visuospatial": ("J", "BF", "CI"),
    "attention": ("AH", "BX", "DB"),
    "executive": ("BA", "CE", "DQ"),
    # Education duration is person-level, so the same value applies to each visit.
    "education_years": ("E", "E", "E"),
}

FIELDS = (
    "rbdq",
    "updrs",
    "mfs",
    "visuospatial",
    "attention",
    "executive",
    "education_years",
)


def _to_number(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def _to_text_id(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.replace("_", "-")


def _normalize_prelbd_id(value):
    text = _to_text_id(value)
    if text is None:
        return None
    compact = text.replace(" ", "")

    hc_match = re.fullmatch(r"(?i)^HC(2)?-?(\d+)$", compact)
    if hc_match:
        visit2 = hc_match.group(1) or ""
        return f"HC{visit2}-{int(hc_match.group(2))}"

    prelbd_match = re.fullmatch(r"(?i)^pre-?(?:lbd|dlb)(2)?-?(\d+)$", compact)
    if prelbd_match:
        visit2 = prelbd_match.group(1) or ""
        return f"pre-LBD{visit2}-{int(prelbd_match.group(2))}"

    return compact


def _extract_numeric_suffix(code, prefix):
    part = code[len(prefix):].strip()
    try:
        return str(int(float(part)))
    except ValueError:
        return part


class Command(BaseCommand):
    help = (
        "Import clinical scores into Subject fields "
        "(rbdq, updrs, mfs, visuospatial, attention, executive, education_years) "
        "from KARDIOVIZE.xlsx and preDLB_shared.xlsx."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--kardiovize",
            dest="kardiovize_path",
            default=str(DEFAULT_KARDIOVIZE_PATH),
            help="Path to KARDIOVIZE.xlsx",
        )
        parser.add_argument(
            "--predlb",
            dest="predlb_path",
            default=str(DEFAULT_PREDLB_PATH),
            help="Path to preDLB_shared.xlsx",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Show what would be updated without saving to DB.",
        )

    def handle(self, *args, **options):
        kardiovize_path = Path(options["kardiovize_path"]).expanduser()
        predlb_path = Path(options["predlb_path"]).expanduser()
        dry_run = bool(options["dry_run"])

        if not kardiovize_path.exists():
            raise CommandError(f"KARDIOVIZE file not found: {kardiovize_path}")
        if not predlb_path.exists():
            raise CommandError(f"preDLB file not found: {predlb_path}")

        kardiovize_values = self._load_kardiovize(kardiovize_path)
        predlb_values = self._load_predlb(predlb_path)

        subjects = list(
            Subject.objects.only("id", "code", *FIELDS)
        )

        updated_subjects = []
        changed_cells = 0
        source_hits = {"kardiovize": 0, "predlb_t1": 0, "predlb_t2": 0, "predlb_t3": 0}

        for subject in subjects:
            incoming = self._resolve_subject_values(subject.code, kardiovize_values, predlb_values)
            if incoming is None:
                continue

            source = incoming.pop("__source")
            source_hits[source] += 1

            has_change = False
            for field, value in incoming.items():
                if value is None:
                    continue
                current = getattr(subject, field)
                if current != value:
                    setattr(subject, field, value)
                    changed_cells += 1
                    has_change = True

            if has_change:
                updated_subjects.append(subject)

        if updated_subjects and not dry_run:
            Subject.objects.bulk_update(updated_subjects, list(FIELDS))

        self.stdout.write(
            self.style.SUCCESS(
                "Clinical import complete: "
                f"subjects_total={len(subjects)}, "
                f"matched_kardiovize={source_hits['kardiovize']}, "
                f"matched_predlb_t1={source_hits['predlb_t1']}, "
                f"matched_predlb_t2={source_hits['predlb_t2']}, "
                f"matched_predlb_t3={source_hits['predlb_t3']}, "
                f"subjects_updated={len(updated_subjects)}, "
                f"cells_changed={changed_cells}, "
                f"dry_run={dry_run}"
            )
        )

        unmatched_coben = sorted(
            code for code in (s.code for s in subjects if s.code.startswith("COBEN-"))
            if _extract_numeric_suffix(code, "COBEN-") not in kardiovize_values
        )
        if unmatched_coben:
            preview = ", ".join(unmatched_coben[:10])
            self.stdout.write(
                self.style.WARNING(
                    f"COBEN subjects without KARDIOVIZE row: {len(unmatched_coben)} "
                    f"(first {min(10, len(unmatched_coben))}: {preview})"
                )
            )

    def _load_kardiovize(self, excel_path):
        from openpyxl import load_workbook
        from openpyxl.utils.cell import column_index_from_string

        wb = load_workbook(excel_path, data_only=True, read_only=True)
        if KARDIO_SHEET not in wb.sheetnames:
            raise CommandError(f"Sheet '{KARDIO_SHEET}' not found in {excel_path}")
        ws = wb[KARDIO_SHEET]

        id_idx = column_index_from_string(KARDIO_ID_COLUMN)
        mapping_idx = {field: column_index_from_string(col) for field, col in KARDIO_MAPPING.items()}

        rows = {}
        for row in range(2, ws.max_row + 1):
            raw_id = ws.cell(row=row, column=id_idx).value
            sid = _to_text_id(raw_id)
            if sid is None:
                continue
            try:
                sid = str(int(float(sid)))
            except ValueError:
                pass

            values = {}
            for field, idx in mapping_idx.items():
                values[field] = _to_number(ws.cell(row=row, column=idx).value)

            rows[sid] = values

        return rows

    def _load_predlb(self, excel_path):
        from openpyxl import load_workbook
        from openpyxl.utils.cell import column_index_from_string

        wb = load_workbook(excel_path, data_only=True, read_only=True)

        results = defaultdict(dict)

        def merge_sheet(sheet_name, id_column, mapping, data_start_row):
            if sheet_name not in wb.sheetnames:
                raise CommandError(f"Sheet '{sheet_name}' not found in {excel_path}")
            ws = wb[sheet_name]
            id_idx = column_index_from_string(id_column)
            mapping_idx = {
                field: tuple(column_index_from_string(c) for c in cols)
                for field, cols in mapping.items()
            }

            for row in range(data_start_row, ws.max_row + 1):
                raw_id = ws.cell(row=row, column=id_idx).value
                base_id = _normalize_prelbd_id(raw_id)
                if base_id is None:
                    continue

                subject_data = results[base_id]
                for field, col_indices in mapping_idx.items():
                    for visit_idx, col_idx in enumerate(col_indices, start=1):
                        value = _to_number(ws.cell(row=row, column=col_idx).value)
                        if value is None:
                            continue
                        subject_data[(field, visit_idx)] = value

        merge_sheet(SCALES_SHEET, SCALES_ID_COLUMN, SCALES_MAPPING, data_start_row=3)
        merge_sheet(PSY_SHEET, PSY_ID_COLUMN, PSY_MAPPING, data_start_row=2)

        return results

    def _resolve_subject_values(self, code, kardiovize_values, predlb_values):
        # KARDIOVIZE codes are COBEN-<numeric_id>
        if code.startswith("COBEN-"):
            sid = _extract_numeric_suffix(code, "COBEN-")
            values = kardiovize_values.get(sid)
            if not values:
                return None
            output = {field: values.get(field) for field in FIELDS}
            output["__source"] = "kardiovize"
            return output

        # preDLB / HC: visits are represented by code suffixes (e.g. pre-LBD2-10, HC2-9)
        visit = 1
        base_id = _normalize_prelbd_id(code)
        if base_id is None:
            return None

        if base_id.startswith("pre-LBD2-"):
            visit = 2
            base_id = base_id.replace("pre-LBD2-", "pre-LBD-", 1)
        elif base_id.startswith("pre-LBD3-"):
            visit = 3
            base_id = base_id.replace("pre-LBD3-", "pre-LBD-", 1)
        elif base_id.startswith("HC2-"):
            visit = 2
            base_id = base_id.replace("HC2-", "HC-", 1)
        elif base_id.startswith("HC3-"):
            visit = 3
            base_id = base_id.replace("HC3-", "HC-", 1)

        values = predlb_values.get(base_id)
        if not values:
            return None

        output = {}
        for field in FIELDS:
            output[field] = values.get((field, visit))
        output["__source"] = f"predlb_t{visit}"
        return output
